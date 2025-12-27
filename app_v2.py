# -*- coding: utf-8 -*-
"""
CS Board Hybrid Viewer v2
- Excel upload (template)
- Edit sow/plant dates in Streamlit
- Recompute workload charts (Jun & Month)
- Simple peak detection + shift suggestions (what-if)
- Download updated Excel (same template + updated inputs + output sheets)
"""
import re
import math
import calendar
import datetime as dt
from dataclasses import dataclass
from typing import Dict, Tuple, Optional, List

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ----------------------------
# Helpers: parsing & mappings
# ----------------------------

@dataclass
class Defaults:
    crop_duration_days: Dict[str, int]
    machine_ability: Dict[str, float]
    sowplant_default_period_days: int

JUN_LABELS = ["上旬", "中旬", "下旬"]
# If sow/plant has a long From/To range, auto-distribute S-based ancillary tasks over the range
# to avoid concentrating everything into a single Jun bucket.
AUTO_DIST_THRESHOLD_DAYS = 30
VALID_MACHINE_CATS = {"tractor", "seeder", "transplanter", "sprayer", "combine", "roller"}
ALL_SOURCES = ["template", "sowplant_range", "exception"]

def jun_no_from_date(d: dt.date) -> int:
    """Return 1..36."""
    part = 1 if d.day <= 10 else (2 if d.day <= 20 else 3)
    return (d.month - 1) * 3 + part

def jun_range(year: int, jun_no: int) -> Tuple[dt.date, dt.date]:
    """Return (start_date, end_date) for a given jun_no in the given year."""
    month = (jun_no - 1) // 3 + 1
    part = (jun_no - 1) % 3 + 1
    if part == 1:
        return dt.date(year, month, 1), dt.date(year, month, 10)
    if part == 2:
        return dt.date(year, month, 11), dt.date(year, month, 20)
    last_day = calendar.monthrange(year, month)[1]
    return dt.date(year, month, 21), dt.date(year, month, last_day)

def jun_days(year: int, jun_no: int) -> int:
    s, e = jun_range(year, jun_no)
    return (e - s).days + 1

def overlap_days(a_start: dt.date, a_end: dt.date, b_start: dt.date, b_end: dt.date) -> int:
    s = max(a_start, b_start)
    e = min(a_end, b_end)
    if e < s:
        return 0
    return (e - s).days + 1

def safe_date(x) -> Optional[dt.date]:
    if pd.isna(x):
        return None
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    s = str(x).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def mid_date(s: dt.date, e: dt.date) -> dt.date:
    if e < s:
        s, e = e, s
    return s + dt.timedelta(days=((e - s).days // 2))

JP_RANGE_PATTERN_1 = re.compile(r"(?P<m1>\d{1,2})月(?P<d1>\d{1,2})日\s*[～〜\-]\s*(?P<m2>\d{1,2})月(?P<d2>\d{1,2})日")
JP_RANGE_PATTERN_2 = re.compile(r"(?P<m1>\d{1,2})月(?P<p1>上旬|中旬|下旬)\s*[～〜\-]\s*(?P<m2>\d{1,2})月(?P<p2>上旬|中旬|下旬)")
JP_SINGLE_PART = re.compile(r"(?P<m>\d{1,2})月(?P<p>上旬|中旬|下旬)")

def part_to_day_range(year:int, month:int, part:str) -> Tuple[dt.date, dt.date]:
    if part == "上旬":
        return dt.date(year, month, 1), dt.date(year, month, 10)
    if part == "中旬":
        return dt.date(year, month, 11), dt.date(year, month, 20)
    last_day = calendar.monthrange(year, month)[1]
    return dt.date(year, month, 21), dt.date(year, month, last_day)

def parse_jp_date_range(text: str, year: int) -> Optional[Tuple[dt.date, dt.date]]:
    """Parse strings like '5月1日～6月20日', '3月下旬~4月上旬', '4月上旬'."""
    if text is None:
        return None
    s = str(text).strip()
    if not s or s.lower() in {"nan", "none"}:
        return None

    m = JP_RANGE_PATTERN_1.search(s)
    if m:
        m1, d1, m2, d2 = map(int, [m.group("m1"), m.group("d1"), m.group("m2"), m.group("d2")])
        return dt.date(year, m1, d1), dt.date(year, m2, d2)

    m = JP_RANGE_PATTERN_2.search(s)
    if m:
        m1, p1, m2, p2 = int(m.group("m1")), m.group("p1"), int(m.group("m2")), m.group("p2")
        a1, _ = part_to_day_range(year, m1, p1)
        _, b2 = part_to_day_range(year, m2, p2)
        return a1, b2

    m = JP_SINGLE_PART.search(s)
    if m and "～" not in s and "〜" not in s and "-" not in s:
        month, part = int(m.group("m")), m.group("p")
        return part_to_day_range(year, month, part)

    d = safe_date(s)
    if d:
        return d, d
    return None

def jn_to_label(jn:int) -> str:
    month = (jn - 1) // 3 + 1
    part = (jn - 1) % 3
    return f"{month}月{JUN_LABELS[part]}"

def read_sheet(file, name: str, header: Optional[int]=None) -> pd.DataFrame:
    return pd.read_excel(file, sheet_name=name, header=header, engine="openpyxl")

def parse_defaults(df_pre_raw: pd.DataFrame) -> Defaults:
    crop_duration_days: Dict[str,int] = {}
    machine_ability: Dict[str,float] = {}
    sowplant_default_period_days = 10

    # crop durations
    for i in range(len(df_pre_raw)):
        v0 = df_pre_raw.iloc[i, 0]
        v1 = df_pre_raw.iloc[i, 1] if df_pre_raw.shape[1] > 1 else None
        if str(v0).strip() == "作物" and ("作期" in str(v1)):
            j = i + 1
            while j < len(df_pre_raw):
                crop = df_pre_raw.iloc[j, 0]
                days = df_pre_raw.iloc[j, 1]
                if pd.isna(crop) or str(crop).strip() == "":
                    break
                try:
                    crop_duration_days[str(crop).strip()] = int(days)
                except Exception:
                    pass
                j += 1
            break

    # machine ability table
    for i in range(len(df_pre_raw)):
        if str(df_pre_raw.iloc[i, 0]).strip().startswith("農機カテゴリ"):
            j = i + 1
            while j < len(df_pre_raw):
                cat = df_pre_raw.iloc[j, 0]
                val = df_pre_raw.iloc[j, 1]
                if pd.isna(cat) or str(cat).strip() == "":
                    break
                try:
                    machine_ability[str(cat).split("（")[0].strip()] = float(val)
                except Exception:
                    pass
                j += 1
            break

    # sowplant default period
    for i in range(len(df_pre_raw)):
        if "播種/移植_期間日数" in str(df_pre_raw.iloc[i, 0]):
            try:
                sowplant_default_period_days = int(df_pre_raw.iloc[i, 1])
            except Exception:
                sowplant_default_period_days = 10
            break

    return Defaults(crop_duration_days=crop_duration_days, machine_ability=machine_ability,
                    sowplant_default_period_days=sowplant_default_period_days)

def compute_farm_capacities(df_mach: pd.DataFrame, defaults: Defaults) -> pd.DataFrame:
    """Return long table: farm, machine_category, cap_ha_per_day"""
    df = df_mach.copy()
    df["農家名"] = df["農家名"].astype(str).str.strip()

    map_cols = {
        "tractor": ("トラクタ能力(入力合計)", "トラクタ推定台数(仮)"),
        "seeder": ("播種機能力(入力)", "播種機推定台数(仮)"),
        "transplanter": ("田植機能力(入力なし)", "田植機推定台数(仮)"),
        "sprayer": ("防除(ブーム)能力(入力合計)", "防除推定台数(仮)"),
        "combine": ("コンバイン能力(入力合計)", "コンバイン推定台数(仮)"),
        "roller": ("鎮圧機能力(入力)", "鎮圧推定台数(仮)"),
    }

    out_rows = []
    for cat, (col_in, col_n) in map_cols.items():
        if col_in not in df.columns or col_n not in df.columns:
            continue
        d_ability = float(defaults.machine_ability.get(cat, np.nan))
        for _, r in df.iterrows():
            farm = r["農家名"]
            in_total = r.get(col_in, np.nan)
            n = r.get(col_n, 1)
            try:
                in_total = float(in_total)
            except Exception:
                in_total = 0.0
            try:
                n = float(n)
            except Exception:
                n = 1.0
            if in_total and in_total > 0:
                cap = in_total
            else:
                cap = np.nan if math.isnan(d_ability) else d_ability * (n if n > 0 else 1.0)
            out_rows.append({"農家名": farm, "農機カテゴリ": cat, "能力(ha/日)": cap})
    return pd.DataFrame(out_rows)

def pick_sowplant_machine(crop_name: str) -> str:
    s = str(crop_name)
    if "移植" in s:
        return "transplanter"
    return "seeder"

def _truthy_10(x) -> bool:
    try:
        return float(x) == 1.0
    except Exception:
        return False

def _get_first_present(d: pd.Series, cols: List[str]):
    for c in cols:
        if c in d.index:
            return d.get(c)
    return None

def _parse_int(x) -> Optional[int]:
    if x is None or pd.isna(x):
        return None
    try:
        return int(float(x))
    except Exception:
        return None

def _jun_mid_date(year: int, jun_no: int) -> dt.date:
    s, e = jun_range(year, jun_no)
    return s + dt.timedelta(days=((e - s).days // 2))

def distribute_range_to_jun(
    *,
    farm: str,
    crop: str,
    start_date: dt.date,
    end_date: dt.date,
    base_year: int,
    work_group: str,
    memo: str,
    work: str,
    cat: str,
    area_total: float,
    cap: float,
    mdays_total: float,
    source: str,
    input_row_index: Optional[int],
) -> List[Dict]:
    """Distribute total area/machine-days across Jun buckets proportionally by overlap days."""
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    period_days = (end_date - start_date).days + 1
    if period_days <= 0:
        return []

    year_segments: List[Tuple[int, dt.date, dt.date]] = []
    if start_date.year == end_date.year:
        year_segments = [(start_date.year, start_date, end_date)]
    else:
        year_segments = [
            (start_date.year, start_date, dt.date(start_date.year, 12, 31)),
            (end_date.year, dt.date(end_date.year, 1, 1), end_date),
        ]

    rows: List[Dict] = []
    for year, a_start, a_end in year_segments:
        for jn in range(1, 37):
            b_start, b_end = jun_range(year, jn)
            ov = overlap_days(a_start, a_end, b_start, b_end)
            if ov <= 0:
                continue
            frac = ov / period_days
            rows.append({
                "農家名": farm,
                "年": year,
                "旬番号": jn,
                "旬ラベル": jn_to_label(jn),
                "月": (jn - 1) // 3 + 1,
                "作物": crop,
                "作業グループ": work_group,
                "メモ": memo,
                "作業": work,
                "農機カテゴリ": cat,
                "面積(ha)": area_total * frac,
                "能力(ha/日)": cap,
                "推定機械日数": mdays_total * frac,
                "推定機械日数_有効": mdays_total * frac,
                "source": source,
                "入力行Index": input_row_index,
            })
    return rows

def _days_inclusive(s: dt.date, e: dt.date) -> int:
    if e < s:
        s, e = e, s
    return (e - s).days + 1

def _overlap_range(a_start: dt.date, a_end: dt.date, b_start: dt.date, b_end: dt.date) -> Optional[Tuple[dt.date, dt.date]]:
    s = max(a_start, b_start)
    e = min(a_end, b_end)
    if e < s:
        return None
    return s, e

def _date_mid(s: dt.date, e: dt.date) -> dt.date:
    if e < s:
        s, e = e, s
    return s + dt.timedelta(days=((e - s).days // 2))

def compute_task_events(
    df_sak: pd.DataFrame,
    df_mach: pd.DataFrame,
    df_tpl: pd.DataFrame,
    df_exc: pd.DataFrame,
    defaults: Defaults,
    *,
    farm: str,
    include_sources: List[str],
) -> pd.DataFrame:
    """
    Build date-range events (start/end) with total area and machine-days.
    These are used to compute daily peaks within a Jun bucket.
    """
    cap_long = compute_farm_capacities(df_mach, defaults)
    cap_lu = cap_long.set_index(["農家名", "農機カテゴリ"])["能力(ha/日)"].to_dict()

    sak = df_sak.copy()
    if "農家名" in sak.columns:
        sak["農家名"] = sak["農家名"].astype(str).str.strip()
    if "作物" in sak.columns:
        sak["作物"] = sak["作物"].astype(str).str.strip()
    sak = sak[sak["農家名"] == farm].copy()
    if sak.empty:
        return pd.DataFrame()

    events: List[Dict] = []

    def add_event(
        *,
        crop: str,
        work_group: str,
        memo: str,
        work: str,
        cat: str,
        start_date: dt.date,
        end_date: dt.date,
        area_total: float,
        source: str,
        input_row_index: Optional[int],
    ):
        cat = str(cat).strip()
        if cat not in VALID_MACHINE_CATS:
            return
        cap = cap_lu.get((farm, cat), np.nan)
        mdays_total = np.nan
        if cap and not pd.isna(cap) and cap > 0:
            mdays_total = area_total / cap
        memo_clean = memo if memo and memo.lower() not in {"nan", "none"} else ""
        events.append({
            "農家名": farm,
            "作物": crop,
            "作業グループ": work_group,
            "メモ": memo_clean,
            "作業": work,
            "農機カテゴリ": cat,
            "From": start_date,
            "To": end_date,
            "面積(ha)": float(area_total),
            "能力(ha/日)": cap,
            "推定機械日数_有効": float(mdays_total) if (mdays_total is not None and not pd.isna(mdays_total)) else 0.0,
            "source": source,
            "入力行Index": input_row_index,
        })

    if "sowplant_range" in include_sources:
        for idx, r in sak.reset_index(drop=True).iterrows():
            rng = sowplant_range_for_row(r, defaults)
            if rng is None:
                continue
            s_date, e_date, machine_cat = rng
            crop = str(r.get("作物", "")).strip()
            area = r.get("面積(ha)", np.nan)
            try:
                area = float(area)
            except Exception:
                continue
            if area <= 0:
                continue
            add_event(
                crop=crop,
                work_group="播種/移植",
                memo="",
                work="播種/移植",
                cat=machine_cat,
                start_date=s_date,
                end_date=e_date,
                area_total=area,
                source="sowplant_range",
                input_row_index=idx + 1,
            )

    if "template" in include_sources and df_tpl is not None and not df_tpl.empty:
        tpl = df_tpl.copy()
        tpl = tpl[pd.to_numeric(tpl["有効(1/0)"], errors="coerce").fillna(0).astype(float) == 1.0].copy()
        tpl = tpl[~tpl["作業グループ"].isin(["播種", "移植"])].copy()

        for idx, r in sak.reset_index(drop=True).iterrows():
            crop = str(r.get("作物", "")).strip()
            area = r.get("面積(ha)", np.nan)
            try:
                area = float(area)
            except Exception:
                continue
            if area <= 0:
                continue

            rep = safe_date(r.get("播種/移植日（代表日・調整）"))
            s_date = safe_date(r.get("播種/移植_開始日（参考）"))
            e_date = safe_date(r.get("播種/移植_終了日（参考）"))
            base_year = (rep.year if rep else (s_date.year if s_date else (e_date.year if e_date else dt.date.today().year)))

            dur = r.get("作期日数(上書き)")
            if pd.isna(dur):
                dur = defaults.crop_duration_days.get(crop, defaults.crop_duration_days.get("その他", 120))
            try:
                dur = int(dur)
            except Exception:
                dur = defaults.crop_duration_days.get("その他", 120)

            sow_rng = sowplant_range_for_row(r, defaults)
            sow_period_days = None
            if sow_rng is not None:
                sow_period_days = _days_inclusive(sow_rng[0], sow_rng[1])

            harvest = safe_date(r.get("収穫日(上書き)"))
            harvest_rng = None
            if sow_rng is not None and sow_period_days is not None and sow_period_days >= 2:
                harvest_rng = (sow_rng[0] + dt.timedelta(days=dur), sow_rng[1] + dt.timedelta(days=dur))
                harvest = _date_mid(harvest_rng[0], harvest_rng[1])
            else:
                if harvest is None and rep is not None:
                    harvest = rep + dt.timedelta(days=dur)
                if harvest is not None:
                    harvest_rng = (harvest, harvest)

            lines = tpl[tpl["作物"].astype(str).str.strip() == crop]
            if lines.empty:
                continue

            for _, t in lines.iterrows():
                memo = str(t.get("メモ", "")).strip()
                work_group = str(t.get("作業グループ", "")).strip()
                work = memo if memo and memo.lower() not in {"nan", "none"} else work_group
                cat = str(t.get("農機カテゴリ", "")).strip()
                base = str(t.get("基準(S=代表日 / H=収穫)", "S")).strip()
                offset = _parse_int(t.get("オフセット旬(仮)", 0)) or 0
                area_coef = t.get("面積係数(仮)", 1)
                times_coef = t.get("回数係数(仮)", 1)
                duration_days = _parse_int(_get_first_present(t, ["期間日数(仮)", "期間日数", "作業期間日数(仮)"]))
                dist_on_sowplant = _truthy_10(_get_first_present(t, ["播種/移植期間に按分(1/0)", "期間に按分(1/0)", "期間按分(1/0)"]))
                dist_on_harvest = _truthy_10(_get_first_present(t, ["収穫期間に按分(1/0)", "収穫に按分(1/0)"]))

                if (
                    not dist_on_sowplant
                    and duration_days is None
                    and sow_rng is not None
                    and sow_period_days is not None
                    and sow_period_days >= AUTO_DIST_THRESHOLD_DAYS
                    and str(base).strip().startswith("S")
                ):
                    dist_on_sowplant = True
                if (
                    not dist_on_harvest
                    and duration_days is None
                    and harvest_rng is not None
                    and sow_rng is not None
                    and sow_period_days is not None
                    and sow_period_days >= 2
                    and str(base).strip().startswith("H")
                ):
                    dist_on_harvest = True

                try:
                    area_coef = float(area_coef)
                except Exception:
                    area_coef = 1.0
                try:
                    times_coef = float(times_coef)
                except Exception:
                    times_coef = 1.0

                task_area = area * area_coef * times_coef

                if dist_on_sowplant and sow_rng is not None:
                    s0, e0, _ = sow_rng
                    delta_days = offset * 10
                    add_event(
                        crop=crop,
                        work_group=work_group,
                        memo=memo,
                        work=work,
                        cat=cat,
                        start_date=s0 + dt.timedelta(days=delta_days),
                        end_date=e0 + dt.timedelta(days=delta_days),
                        area_total=task_area,
                        source="template",
                        input_row_index=idx + 1,
                    )
                    continue

                if dist_on_harvest and harvest_rng is not None:
                    s0, e0 = harvest_rng
                    delta_days = offset * 10
                    add_event(
                        crop=crop,
                        work_group=work_group,
                        memo=memo,
                        work=work,
                        cat=cat,
                        start_date=s0 + dt.timedelta(days=delta_days),
                        end_date=e0 + dt.timedelta(days=delta_days),
                        area_total=task_area,
                        source="template",
                        input_row_index=idx + 1,
                    )
                    continue

                # fall back: anchor date (rep/harvest) + offset, optionally with duration
                anchor = rep if str(base).strip().startswith("S") else harvest
                if anchor is None and sow_rng is not None:
                    anchor = _date_mid(sow_rng[0], sow_rng[1])
                if anchor is None:
                    continue
                anchor = anchor + dt.timedelta(days=offset * 10)
                if duration_days is not None and duration_days >= 2:
                    left = (duration_days - 1) // 2
                    right = (duration_days - 1) - left
                    s2 = anchor - dt.timedelta(days=left)
                    e2 = anchor + dt.timedelta(days=right)
                else:
                    s2 = anchor
                    e2 = anchor
                add_event(
                    crop=crop,
                    work_group=work_group,
                    memo=memo,
                    work=work,
                    cat=cat,
                    start_date=s2,
                    end_date=e2,
                    area_total=task_area,
                    source="template",
                    input_row_index=idx + 1,
                )

    if "exception" in include_sources and df_exc is not None and not df_exc.empty:
        exc = df_exc.copy()
        exc = exc[exc["農家名"].notna() & exc["作物"].notna()]
        exc["有効(1/0)"] = pd.to_numeric(exc["有効(1/0)"], errors="coerce").fillna(0)
        exc = exc[exc["有効(1/0)"] == 1]
        if not exc.empty:
            for _, e in exc.iterrows():
                if str(e.get("農家名", "")).strip() != farm:
                    continue
                crop = str(e.get("作物", "")).strip()
                work_group = str(e.get("作業グループ", "")).strip() or "例外"
                memo = str(e.get("メモ", "")).strip()
                work = memo if memo and memo.lower() not in {"nan", "none"} else work_group
                cat = str(e.get("農機カテゴリ", "")).strip()
                base = str(e.get("基準(S/H)", "S")).strip()
                offset = _parse_int(e.get("オフセット旬", 0)) or 0
                area_coef = e.get("面積係数", 1)
                times_coef = e.get("回数係数", 1)
                duration_days = _parse_int(_get_first_present(e, ["期間日数", "期間日数(仮)", "作業期間日数(仮)"]))
                dist_on_sowplant = _truthy_10(_get_first_present(e, ["播種/移植期間に按分(1/0)", "期間に按分(1/0)", "期間按分(1/0)"]))
                dist_on_harvest = _truthy_10(_get_first_present(e, ["収穫期間に按分(1/0)", "収穫に按分(1/0)"]))

                try:
                    area_coef = float(area_coef)
                except Exception:
                    area_coef = 1.0
                try:
                    times_coef = float(times_coef)
                except Exception:
                    times_coef = 1.0

                # base sow/harvest range from sak rows for this crop
                sak_fc = sak[sak["作物"] == crop].copy()
                if sak_fc.empty:
                    continue
                base_row = sak_fc.reset_index(drop=True).iloc[0]
                sow_rng = sowplant_range_for_row(base_row, defaults)
                if sow_rng is None:
                    continue
                sow_period_days = _days_inclusive(sow_rng[0], sow_rng[1])

                dur = defaults.crop_duration_days.get(crop, defaults.crop_duration_days.get("その他", 120))
                try:
                    dur = int(dur)
                except Exception:
                    dur = defaults.crop_duration_days.get("その他", 120)
                harvest_rng = (sow_rng[0] + dt.timedelta(days=dur), sow_rng[1] + dt.timedelta(days=dur)) if sow_period_days >= 2 else None

                if (
                    not dist_on_sowplant
                    and duration_days is None
                    and sow_period_days >= AUTO_DIST_THRESHOLD_DAYS
                    and str(base).strip().startswith("S")
                ):
                    dist_on_sowplant = True
                if (
                    not dist_on_harvest
                    and duration_days is None
                    and harvest_rng is not None
                    and sow_period_days >= 2
                    and str(base).strip().startswith("H")
                ):
                    dist_on_harvest = True

                area_total = float(sak_fc["面積(ha)"].astype(float).sum()) * area_coef * times_coef

                if dist_on_sowplant:
                    s0, e0, _ = sow_rng
                    delta_days = offset * 10
                    add_event(
                        crop=crop,
                        work_group=work_group,
                        memo=memo,
                        work=work,
                        cat=cat,
                        start_date=s0 + dt.timedelta(days=delta_days),
                        end_date=e0 + dt.timedelta(days=delta_days),
                        area_total=area_total,
                        source="exception",
                        input_row_index=None,
                    )
                    continue

                if dist_on_harvest and harvest_rng is not None:
                    s0, e0 = harvest_rng
                    delta_days = offset * 10
                    add_event(
                        crop=crop,
                        work_group=work_group,
                        memo=memo,
                        work=work,
                        cat=cat,
                        start_date=s0 + dt.timedelta(days=delta_days),
                        end_date=e0 + dt.timedelta(days=delta_days),
                        area_total=area_total,
                        source="exception",
                        input_row_index=None,
                    )
                    continue

                anchor = _date_mid(sow_rng[0], sow_rng[1]) if str(base).strip().startswith("S") else _date_mid(harvest_rng[0], harvest_rng[1]) if harvest_rng else None
                if anchor is None:
                    continue
                anchor = anchor + dt.timedelta(days=offset * 10)
                if duration_days is not None and duration_days >= 2:
                    left = (duration_days - 1) // 2
                    right = (duration_days - 1) - left
                    s2 = anchor - dt.timedelta(days=left)
                    e2 = anchor + dt.timedelta(days=right)
                else:
                    s2 = anchor
                    e2 = anchor
                add_event(
                    crop=crop,
                    work_group=work_group,
                    memo=memo,
                    work=work,
                    cat=cat,
                    start_date=s2,
                    end_date=e2,
                    area_total=area_total,
                    source="exception",
                    input_row_index=None,
                )

    ev = pd.DataFrame(events)
    if ev.empty:
        return ev
    ev["From"] = pd.to_datetime(ev["From"]).dt.date
    ev["To"] = pd.to_datetime(ev["To"]).dt.date
    return ev

def daily_load_in_window(
    events: pd.DataFrame,
    *,
    window_start: dt.date,
    window_end: dt.date,
    machine_cat: Optional[str],
    utilization: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Return (daily_totals, daily_contrib) within window for selected machine category."""
    if events is None or events.empty:
        return pd.DataFrame(), pd.DataFrame()

    ev = events.copy()
    if machine_cat:
        ev = ev[ev["農機カテゴリ"] == machine_cat].copy()
    if ev.empty:
        return pd.DataFrame(), pd.DataFrame()

    totals = {}
    contrib_rows = []

    for _, r in ev.iterrows():
        s = r["From"]
        e = r["To"]
        if s is None or e is None:
            continue
        rng = _overlap_range(s, e, window_start, window_end)
        if rng is None:
            continue
        os, oe = rng
        event_days = _days_inclusive(s, e)
        if event_days <= 0:
            continue
        md_per_day = float(r.get("推定機械日数_有効", 0.0)) / event_days
        area_per_day = float(r.get("面積(ha)", 0.0)) / event_days
        d = os
        while d <= oe:
            totals[d] = totals.get(d, 0.0) + md_per_day
            contrib_rows.append({
                "日付": d,
                "作物": r.get("作物", ""),
                "作業": r.get("作業", ""),
                "作業グループ": r.get("作業グループ", ""),
                "メモ": r.get("メモ", ""),
                "source": r.get("source", ""),
                "入力行Index": r.get("入力行Index", None),
                "面積(ha)": area_per_day,
                "推定機械日数_有効": md_per_day,
            })
            d = d + dt.timedelta(days=1)

    if not totals:
        return pd.DataFrame(), pd.DataFrame()

    days = sorted(totals.keys())
    df_day = pd.DataFrame({
        "日付": days,
        "推定機械日数_有効": [totals[d] for d in days],
    })
    df_day["容量(機械日/日)"] = float(utilization)
    df_day["利用率"] = np.where(df_day["容量(機械日/日)"] > 0, df_day["推定機械日数_有効"] / df_day["容量(機械日/日)"], np.nan)
    df_contrib = pd.DataFrame(contrib_rows)
    return df_day, df_contrib

def suggest_shifts_daily_peak(
    df_sak: pd.DataFrame,
    df_mach: pd.DataFrame,
    df_tpl: pd.DataFrame,
    df_exc: pd.DataFrame,
    defaults: Defaults,
    *,
    farm: str,
    window_start: dt.date,
    window_end: dt.date,
    bottleneck_cat: str,
    utilization: float,
    include_sources: List[str],
    top_n: int = 10,
    deltas: List[int] = [-20, -10, 10, 20],
) -> pd.DataFrame:
    """Try shifting sow/plant rows and evaluate max daily utilization in the target window for bottleneck machine category."""
    ev0 = compute_task_events(df_sak, df_mach, df_tpl, df_exc, defaults, farm=farm, include_sources=include_sources)
    day0, _ = daily_load_in_window(ev0, window_start=window_start, window_end=window_end, machine_cat=bottleneck_cat, utilization=utilization)
    if day0 is None or day0.empty:
        return pd.DataFrame()
    peak0 = float(day0["利用率"].max())

    sak_f = df_sak.copy()
    sak_f["農家名"] = sak_f["農家名"].astype(str).str.strip()
    sak_f = sak_f[sak_f["農家名"] == farm].copy()
    sak_f["面積(ha)"] = pd.to_numeric(sak_f["面積(ha)"], errors="coerce")
    sak_f = sak_f[sak_f["面積(ha)"].fillna(0) > 0].copy()
    if sak_f.empty:
        return pd.DataFrame()

    cand = []
    for i, r in sak_f.reset_index(drop=False).iterrows():
        rng = sowplant_range_for_row(r, defaults)
        if rng is None:
            continue
        cand.append((int(r["index"]), float(r["面積(ha)"])))
    cand = sorted(cand, key=lambda x: x[1], reverse=True)[:max(1, int(top_n))]

    out = []
    for idx, area in cand:
        base_row = df_sak.loc[idx].copy()
        crop = str(base_row.get("作物", "")).strip()

        allow_from = safe_date(base_row.get("播種/移植_開始日（参考）"))
        allow_to = safe_date(base_row.get("播種/移植_終了日（参考）"))
        rep = safe_date(base_row.get("播種/移植日（代表日・調整）"))
        if (allow_from is None or allow_to is None) and base_row.get("元入力(参考)") is not None:
            base_year = rep.year if rep else dt.date.today().year
            parsed = parse_jp_date_range(base_row.get("元入力(参考)"), base_year)
            if parsed:
                allow_from, allow_to = parsed
        if allow_from is None or allow_to is None:
            continue

        used_rng = sowplant_range_for_row(base_row, defaults)
        if used_rng is None:
            continue
        used_from, used_to, _ = used_rng

        for delta in deltas:
            new_from = used_from + dt.timedelta(days=delta)
            new_to = used_to + dt.timedelta(days=delta)
            if new_from < allow_from or new_to > allow_to:
                continue

            df_sak2 = df_sak.copy()
            if "播種/移植日（代表日・調整）" in df_sak2.columns:
                df_sak2.loc[idx, "播種/移植日（代表日・調整）"] = mid_date(new_from, new_to)
            if "播種/移植_開始日（参考）" in df_sak2.columns:
                df_sak2.loc[idx, "播種/移植_開始日（参考）"] = new_from
            if "播種/移植_終了日（参考）" in df_sak2.columns:
                df_sak2.loc[idx, "播種/移植_終了日（参考）"] = new_to

            ev1 = compute_task_events(df_sak2, df_mach, df_tpl, df_exc, defaults, farm=farm, include_sources=include_sources)
            day1, _ = daily_load_in_window(ev1, window_start=window_start, window_end=window_end, machine_cat=bottleneck_cat, utilization=utilization)
            if day1 is None or day1.empty:
                continue
            peak1 = float(day1["利用率"].max())
            out.append({
                "入力行Index": idx + 1,
                "作物": crop,
                "面積(ha)": area,
                "シフト(日)": delta,
                "播種/移植_From(現状)": used_from,
                "播種/移植_To(現状)": used_to,
                "播種/移植_From(提案)": new_from,
                "播種/移植_To(提案)": new_to,
                "日単位ピーク利用率(現状)": peak0,
                "日単位ピーク利用率(提案後)": peak1,
                "改善(Δ)": peak0 - peak1,
            })

    if not out:
        return pd.DataFrame()
    return pd.DataFrame(out).sort_values(["改善(Δ)", "面積(ha)"], ascending=[False, False]).reset_index(drop=True)

def sowplant_range_for_row(r: pd.Series, defaults: Defaults) -> Optional[Tuple[dt.date, dt.date, str]]:
    """Return (from,to,machine_cat) used for distribution. None if cannot decide."""
    crop = r.get("作物", None)
    if pd.isna(crop):
        return None
    crop = str(crop).strip()
    rep = safe_date(r.get("播種/移植日（代表日・調整）"))
    s_date = safe_date(r.get("播種/移植_開始日（参考）"))
    e_date = safe_date(r.get("播種/移植_終了日（参考）"))

    # parse from 元入力(参考)
    if (s_date is None or e_date is None) and r.get("元入力(参考)") is not None:
        base_year = rep.year if rep else dt.date.today().year
        parsed = parse_jp_date_range(r.get("元入力(参考)"), base_year)
        if parsed:
            s_date, e_date = parsed

    # build around rep if missing
    if (s_date is None or e_date is None) and rep is not None:
        pdays = defaults.sowplant_default_period_days
        left = (pdays - 1) // 2
        right = (pdays - 1) - left
        s_date = rep - dt.timedelta(days=left)
        e_date = rep + dt.timedelta(days=right)

    if s_date is None or e_date is None:
        return None
    if e_date < s_date:
        s_date, e_date = e_date, s_date

    return s_date, e_date, pick_sowplant_machine(crop)

def compute_sowplant_distribution(df_sak: pd.DataFrame,
                                  cap_long: pd.DataFrame,
                                  defaults: Defaults) -> Tuple[pd.DataFrame, pd.DataFrame]:
    cap_lu = cap_long.set_index(["農家名", "農機カテゴリ"])["能力(ha/日)"].to_dict()
    dist_rows = []
    check_rows = []

    for idx, r in df_sak.reset_index(drop=True).iterrows():
        farm = str(r.get("農家名", "")).strip()
        crop = r.get("作物", None)
        area = r.get("面積(ha)", np.nan)
        if not farm or pd.isna(crop) or pd.isna(area):
            continue
        crop = str(crop).strip()
        try:
            area = float(area)
        except Exception:
            continue
        if area <= 0:
            continue

        rng = sowplant_range_for_row(r, defaults)
        if rng is None:
            continue
        s_date, e_date, machine_cat = rng

        period_days = (e_date - s_date).days + 1
        cap = cap_lu.get((farm, machine_cat), np.nan)

        required_days = np.nan
        shortage_days = np.nan
        status = "不明"
        if cap and not pd.isna(cap) and cap > 0:
            required_days = area / cap
            shortage_days = max(0.0, required_days - period_days)
            status = "OK" if shortage_days <= 1e-9 else "不足"

        check_rows.append({
            "農家名": farm,
            "作物": crop,
            "面積(ha)": area,
            "播種/移植_From": s_date,
            "播種/移植_To": e_date,
            "期間日数": period_days,
            "使用農機カテゴリ": machine_cat,
            "使用能力(ha/日)": cap,
            "必要日数": required_days,
            "不足日数": shortage_days,
            "判定": status,
            "入力行Index": idx + 1,
        })

        # distribute across jun
        year_segments: List[Tuple[int, dt.date, dt.date]] = []
        if s_date.year == e_date.year:
            year_segments = [(s_date.year, s_date, e_date)]
        else:
            year_segments = [
                (s_date.year, s_date, dt.date(s_date.year, 12, 31)),
                (e_date.year, dt.date(e_date.year, 1, 1), e_date),
            ]

        for year, a_start, a_end in year_segments:
            for jn in range(1, 37):
                b_start, b_end = jun_range(year, jn)
                ov = overlap_days(a_start, a_end, b_start, b_end)
                if ov <= 0:
                    continue
                area_j = area * (ov / period_days)
                machine_days = np.nan
                if cap and not pd.isna(cap) and cap > 0:
                    machine_days = area_j / cap
                dist_rows.append({
                    "農家名": farm,
                    "年": year,
                    "旬番号": jn,
                    "旬ラベル": jn_to_label(jn),
                    "月": (jn - 1) // 3 + 1,
                    "作物": crop,
                    "作業グループ": "播種/移植",
                    "メモ": "",
                    "作業": "播種/移植",
                    "農機カテゴリ": machine_cat,
                    "面積(ha)": area_j,
                    "能力(ha/日)": cap,
                    "推定機械日数": machine_days,
                    "推定機械日数_有効": machine_days,
                    "source": "sowplant_range",
                    "入力行Index": idx + 1,
                })

    return pd.DataFrame(dist_rows), pd.DataFrame(check_rows)

def compute_template_tasks(df_sak: pd.DataFrame,
                           df_tpl: pd.DataFrame,
                           cap_long: pd.DataFrame,
                           defaults: Defaults) -> pd.DataFrame:
    cap_lu = cap_long.set_index(["農家名", "農機カテゴリ"])["能力(ha/日)"].to_dict()
    tpl = df_tpl.copy()
    tpl = tpl[pd.to_numeric(tpl["有効(1/0)"], errors="coerce").fillna(0).astype(float) == 1.0].copy()
    tpl = tpl[~tpl["作業グループ"].isin(["播種", "移植"])].copy()

    rows = []
    for idx, r in df_sak.reset_index(drop=True).iterrows():
        farm = str(r.get("農家名", "")).strip()
        crop = r.get("作物", None)
        area = r.get("面積(ha)", np.nan)
        if not farm or pd.isna(crop) or pd.isna(area):
            continue
        crop = str(crop).strip()
        try:
            area = float(area)
        except Exception:
            continue
        if area <= 0:
            continue

        rep_input = safe_date(r.get("播種/移植日（代表日・調整）"))
        s_date = safe_date(r.get("播種/移植_開始日（参考）"))
        e_date = safe_date(r.get("播種/移植_終了日（参考）"))
        sow_rng = sowplant_range_for_row(r, defaults)
        rep = rep_input
        if rep is None and sow_rng is not None:
            rep = mid_date(sow_rng[0], sow_rng[1])
        base_year = (rep.year if rep else (s_date.year if s_date else (e_date.year if e_date else dt.date.today().year)))

        dur = r.get("作期日数(上書き)")
        if pd.isna(dur):
            dur = defaults.crop_duration_days.get(crop, defaults.crop_duration_days.get("その他", 120))
        try:
            dur = int(dur)
        except Exception:
            dur = defaults.crop_duration_days.get("その他", 120)

        sow_period_days = None
        if sow_rng is not None:
            sow_period_days = (sow_rng[1] - sow_rng[0]).days + 1

        # Harvest is derived from crop duration and sow/plant range when available.
        harvest = safe_date(r.get("収穫日(上書き)"))
        harvest_rng = None
        if sow_rng is not None and sow_period_days is not None and sow_period_days >= 2:
            harvest_rng = (sow_rng[0] + dt.timedelta(days=dur), sow_rng[1] + dt.timedelta(days=dur))
            harvest_mid = harvest_rng[0] + dt.timedelta(days=((harvest_rng[1] - harvest_rng[0]).days // 2))
            harvest = harvest_mid
        else:
            if harvest is None and rep is not None:
                harvest = rep + dt.timedelta(days=dur)
            if harvest is not None:
                harvest_rng = (harvest, harvest)

        base_s = jun_no_from_date(rep) if rep else None
        base_h = jun_no_from_date(harvest) if harvest else None

        lines = tpl[tpl["作物"].astype(str).str.strip() == crop]
        if lines.empty:
            continue

        for _, t in lines.iterrows():
            memo = str(t.get("メモ", "")).strip()
            work_group = str(t.get("作業グループ", "")).strip()
            work = memo if memo and memo.lower() not in {"nan", "none"} else work_group
            cat = str(t.get("農機カテゴリ", "")).strip()
            base = str(t.get("基準(S=代表日 / H=収穫)", "S")).strip()
            offset = t.get("オフセット旬(仮)", 0)
            area_coef = t.get("面積係数(仮)", 1)
            times_coef = t.get("回数係数(仮)", 1)
            duration_days = _parse_int(_get_first_present(t, ["期間日数(仮)", "期間日数", "作業期間日数(仮)"]))
            dist_on_sowplant = _truthy_10(_get_first_present(t, ["播種/移植期間に按分(1/0)", "期間に按分(1/0)", "期間按分(1/0)"]))
            dist_on_harvest = _truthy_10(_get_first_present(t, ["収穫期間に按分(1/0)", "収穫に按分(1/0)"]))
            if (
                not dist_on_sowplant
                and duration_days is None
                and sow_rng is not None
                and sow_period_days is not None
                and sow_period_days >= AUTO_DIST_THRESHOLD_DAYS
                and str(base).strip().startswith("S")
            ):
                dist_on_sowplant = True
            if (
                not dist_on_harvest
                and duration_days is None
                and harvest_rng is not None
                and sow_rng is not None
                and sow_period_days is not None
                and sow_period_days >= 2
                and str(base).strip().startswith("H")
            ):
                dist_on_harvest = True

            try:
                offset = int(offset)
            except Exception:
                offset = 0
            try:
                area_coef = float(area_coef)
            except Exception:
                area_coef = 1.0
            try:
                times_coef = float(times_coef)
            except Exception:
                times_coef = 1.0

            base_jn = base_s if base.startswith("S") else base_h
            if base_jn is None:
                continue
            jn = ((base_jn - 1 + offset) % 36) + 1
            cap = cap_lu.get((farm, cat), np.nan)

            task_area = area * area_coef
            mdays = np.nan
            if cap and not pd.isna(cap) and cap > 0:
                mdays = (task_area / cap) * times_coef

            memo_clean = memo if memo.lower() not in {"nan", "none"} else ""

            if dist_on_sowplant:
                if sow_rng is not None:
                    s_date, e_date, _ = sow_rng
                    # offset is in Jun; approximate as 10 days shift (consistent with suggest_shifts)
                    delta_days = int(offset) * 10
                    s2 = s_date + dt.timedelta(days=delta_days)
                    e2 = e_date + dt.timedelta(days=delta_days)
                    rows.extend(distribute_range_to_jun(
                        farm=farm,
                        crop=crop,
                        start_date=s2,
                        end_date=e2,
                        base_year=base_year,
                        work_group=work_group,
                        memo=memo_clean,
                        work=work,
                        cat=cat,
                        area_total=task_area,
                        cap=cap,
                        mdays_total=float(mdays) if not pd.isna(mdays) else 0.0,
                        source="template",
                        input_row_index=idx + 1,
                    ))
                    continue

            if dist_on_harvest:
                if harvest_rng is not None:
                    s_date, e_date = harvest_rng
                    delta_days = int(offset) * 10
                    s2 = s_date + dt.timedelta(days=delta_days)
                    e2 = e_date + dt.timedelta(days=delta_days)
                    rows.extend(distribute_range_to_jun(
                        farm=farm,
                        crop=crop,
                        start_date=s2,
                        end_date=e2,
                        base_year=base_year,
                        work_group=work_group,
                        memo=memo_clean,
                        work=work,
                        cat=cat,
                        area_total=task_area,
                        cap=cap,
                        mdays_total=float(mdays) if not pd.isna(mdays) else 0.0,
                        source="template",
                        input_row_index=idx + 1,
                    ))
                    continue

            if duration_days is not None and duration_days >= 2:
                rep_date = _jun_mid_date(base_year, jn)
                left = (duration_days - 1) // 2
                right = (duration_days - 1) - left
                s2 = rep_date - dt.timedelta(days=left)
                e2 = rep_date + dt.timedelta(days=right)
                rows.extend(distribute_range_to_jun(
                    farm=farm,
                    crop=crop,
                    start_date=s2,
                    end_date=e2,
                    base_year=base_year,
                    work_group=work_group,
                    memo=memo_clean,
                    work=work,
                    cat=cat,
                    area_total=task_area,
                    cap=cap,
                    mdays_total=float(mdays) if not pd.isna(mdays) else 0.0,
                    source="template",
                    input_row_index=idx + 1,
                ))
                continue

            rows.append({
                "農家名": farm,
                "年": base_year,
                "旬番号": jn,
                "旬ラベル": jn_to_label(jn),
                "月": (jn - 1) // 3 + 1,
                "作物": crop,
                "作業グループ": work_group,
                "メモ": memo_clean,
                "作業": work,
                "農機カテゴリ": cat,
                "面積(ha)": task_area,
                "能力(ha/日)": cap,
                "推定機械日数": mdays,
                "推定機械日数_有効": mdays,
                "source": "template",
                "入力行Index": idx + 1,
            })

    return pd.DataFrame(rows)

def compute_exceptions(df_sak: pd.DataFrame,
                       df_exc: pd.DataFrame,
                       cap_long: pd.DataFrame,
                       defaults: Defaults) -> pd.DataFrame:
    if df_exc is None or df_exc.empty:
        return pd.DataFrame(columns=["農家名","年","旬番号","旬ラベル","月","作物","作業グループ","メモ","作業","農機カテゴリ",
                                     "面積(ha)","能力(ha/日)","推定機械日数","推定機械日数_有効","source","入力行Index"])

    exc = df_exc.copy()
    exc = exc[exc["農家名"].notna() & exc["作物"].notna()]
    exc["有効(1/0)"] = pd.to_numeric(exc["有効(1/0)"], errors="coerce").fillna(0)
    exc = exc[exc["有効(1/0)"] == 1]
    if exc.empty:
        return pd.DataFrame(columns=["農家名","年","旬番号","旬ラベル","月","作物","作業グループ","メモ","作業","農機カテゴリ",
                                     "面積(ha)","能力(ha/日)","推定機械日数","推定機械日数_有効","source","入力行Index"])

    cap_lu = cap_long.set_index(["農家名","農機カテゴリ"])["能力(ha/日)"].to_dict()

    sak = df_sak.copy()
    sak = sak[sak["農家名"].notna() & sak["作物"].notna() & sak["面積(ha)"].notna()]
    sak["農家名"] = sak["農家名"].astype(str).str.strip()
    sak["作物"] = sak["作物"].astype(str).str.strip()
    sak["面積(ha)"] = pd.to_numeric(sak["面積(ha)"], errors="coerce")
    sak = sak[sak["面積(ha)"] > 0]
    if sak.empty:
        return pd.DataFrame()

    def median_date(series):
        ds = [safe_date(x) for x in series]
        ds = [d for d in ds if d is not None]
        if not ds:
            return None
        ds_sorted = sorted(ds)
        return ds_sorted[len(ds_sorted)//2]

    grouped = sak.reset_index(drop=True).groupby(["農家名","作物"], as_index=False).agg({
        "面積(ha)":"sum",
        "播種/移植日（代表日・調整）": median_date,
        "収穫日(上書き)": median_date,
        "播種/移植_開始日（参考）": median_date,
        "播種/移植_終了日（参考）": median_date,
    })

    rows = []
    for _, e in exc.iterrows():
        farm = str(e["農家名"]).strip()
        crop = str(e["作物"]).strip()
        memo = str(e.get("メモ", "")).strip()
        work_group = str(e.get("作業グループ","")).strip()
        work = memo if memo and memo.lower() not in {"nan", "none"} else work_group
        cat = str(e.get("農機カテゴリ","")).strip()
        base = str(e.get("基準(S/H)","S")).strip()
        offset = e.get("オフセット旬", 0)
        area_coef = e.get("面積係数", 1)
        times_coef = e.get("回数係数", 1)
        jun_no = e.get("旬No", np.nan)
        duration_days = _parse_int(_get_first_present(e, ["期間日数", "期間日数(仮)", "作業期間日数(仮)"]))
        dist_on_sowplant = _truthy_10(_get_first_present(e, ["播種/移植期間に按分(1/0)", "期間に按分(1/0)", "期間按分(1/0)"]))
        dist_on_harvest = _truthy_10(_get_first_present(e, ["収穫期間に按分(1/0)", "収穫に按分(1/0)"]))

        try:
            offset = int(offset)
        except Exception:
            offset = 0
        try:
            area_coef = float(area_coef)
        except Exception:
            area_coef = 1.0
        try:
            times_coef = float(times_coef)
        except Exception:
            times_coef = 1.0

        g = grouped[(grouped["農家名"]==farm) & (grouped["作物"]==crop)]
        if g.empty:
            continue
        area_total = float(g.iloc[0]["面積(ha)"])
        rep = g.iloc[0]["播種/移植日（代表日・調整）"]
        rep = rep if isinstance(rep, dt.date) else safe_date(rep)
        # if representative date is missing, derive it from sow/plant range for this crop
        rep_eff = rep
        base_row_for_rng = pd.Series({
            "作物": crop,
            "播種/移植日（代表日・調整）": rep,
            "播種/移植_開始日（参考）": g.iloc[0].get("播種/移植_開始日（参考）"),
            "播種/移植_終了日（参考）": g.iloc[0].get("播種/移植_終了日（参考）"),
            "元入力(参考)": g.iloc[0].get("元入力(参考)"),
        })
        sow_rng0 = sowplant_range_for_row(base_row_for_rng, defaults)
        if rep_eff is None and sow_rng0 is not None:
            rep_eff = mid_date(sow_rng0[0], sow_rng0[1])
        base_year = rep_eff.year if rep_eff else dt.date.today().year

        harvest = safe_date(g.iloc[0].get("収穫日(上書き)"))
        if harvest is None and rep_eff is not None:
            dur = defaults.crop_duration_days.get(crop, defaults.crop_duration_days.get("その他",120))
            harvest = rep_eff + dt.timedelta(days=int(dur))

        base_s = jun_no_from_date(rep_eff) if rep_eff else None
        base_h = jun_no_from_date(harvest) if harvest else None

        if pd.isna(jun_no) or jun_no is None:
            base_jn = base_s if base.startswith("S") else base_h
            if base_jn is None:
                continue
            jun_no = ((base_jn - 1 + offset) % 36) + 1
        else:
            try:
                jun_no = int(jun_no)
            except Exception:
                continue

        cap = cap_lu.get((farm, cat), np.nan)
        task_area = area_total * area_coef
        mdays = np.nan
        if cap and not pd.isna(cap) and cap > 0:
            mdays = (task_area / cap) * times_coef

        memo_clean = memo if memo.lower() not in {"nan", "none"} else ""
        work_group_clean = work_group if work_group else "例外"
        work_clean = work if work else work_group_clean

        base_row = pd.Series({
            "作物": crop,
            "播種/移植日（代表日・調整）": rep_eff,
            "播種/移植_開始日（参考）": g.iloc[0].get("播種/移植_開始日（参考）"),
            "播種/移植_終了日（参考）": g.iloc[0].get("播種/移植_終了日（参考）"),
            "元入力(参考)": g.iloc[0].get("元入力(参考)"),
            "収穫日(上書き)": g.iloc[0].get("収穫日(上書き)"),
        })
        sow_rng = sowplant_range_for_row(base_row, defaults)
        sow_period_days = None
        if sow_rng is not None:
            sow_period_days = (sow_rng[1] - sow_rng[0]).days + 1
        dur = defaults.crop_duration_days.get(crop, defaults.crop_duration_days.get("その他", 120))
        try:
            dur = int(dur)
        except Exception:
            dur = defaults.crop_duration_days.get("その他", 120)

        harvest_rng = None
        if sow_rng is not None and sow_period_days is not None and sow_period_days >= 2:
            harvest_rng = (sow_rng[0] + dt.timedelta(days=int(dur)), sow_rng[1] + dt.timedelta(days=int(dur)))
        elif harvest is not None:
            harvest_rng = (harvest, harvest)
        if (
            not dist_on_sowplant
            and duration_days is None
            and sow_rng is not None
            and sow_period_days is not None
            and sow_period_days >= AUTO_DIST_THRESHOLD_DAYS
            and str(base).strip().startswith("S")
        ):
            dist_on_sowplant = True
        if (
            not dist_on_harvest
            and duration_days is None
            and harvest_rng is not None
            and sow_rng is not None
            and sow_period_days is not None
            and sow_period_days >= 2
            and str(base).strip().startswith("H")
        ):
            dist_on_harvest = True

        if dist_on_sowplant:
            if sow_rng is not None:
                s_date, e_date, _ = sow_rng
                delta_days = int(offset) * 10
                s2 = s_date + dt.timedelta(days=delta_days)
                e2 = e_date + dt.timedelta(days=delta_days)
                rows.extend(distribute_range_to_jun(
                    farm=farm,
                    crop=crop,
                    start_date=s2,
                    end_date=e2,
                    base_year=base_year,
                    work_group=work_group_clean,
                    memo=memo_clean,
                    work=work_clean,
                    cat=cat,
                    area_total=task_area,
                    cap=cap,
                    mdays_total=float(mdays) if not pd.isna(mdays) else 0.0,
                    source="exception",
                    input_row_index=None,
                ))
                continue

        if dist_on_harvest:
            if harvest_rng is not None:
                s_date, e_date = harvest_rng
                delta_days = int(offset) * 10
                s2 = s_date + dt.timedelta(days=delta_days)
                e2 = e_date + dt.timedelta(days=delta_days)
                rows.extend(distribute_range_to_jun(
                    farm=farm,
                    crop=crop,
                    start_date=s2,
                    end_date=e2,
                    base_year=base_year,
                    work_group=work_group_clean,
                    memo=memo_clean,
                    work=work_clean,
                    cat=cat,
                    area_total=task_area,
                    cap=cap,
                    mdays_total=float(mdays) if not pd.isna(mdays) else 0.0,
                    source="exception",
                    input_row_index=None,
                ))
                continue

        if duration_days is not None and duration_days >= 2:
            rep_date = _jun_mid_date(base_year, int(jun_no))
            left = (duration_days - 1) // 2
            right = (duration_days - 1) - left
            s2 = rep_date - dt.timedelta(days=left)
            e2 = rep_date + dt.timedelta(days=right)
            rows.extend(distribute_range_to_jun(
                farm=farm,
                crop=crop,
                start_date=s2,
                end_date=e2,
                base_year=base_year,
                work_group=work_group_clean,
                memo=memo_clean,
                work=work_clean,
                cat=cat,
                area_total=task_area,
                cap=cap,
                mdays_total=float(mdays) if not pd.isna(mdays) else 0.0,
                source="exception",
                input_row_index=None,
            ))
            continue

        rows.append({
            "農家名": farm,
            "年": base_year,
            "旬番号": jun_no,
            "旬ラベル": jn_to_label(jun_no),
            "月": (jun_no - 1) // 3 + 1,
            "作物": crop,
            "作業グループ": work_group_clean,
            "メモ": memo_clean,
            "作業": work_clean,
            "農機カテゴリ": cat,
            "面積(ha)": task_area,
            "能力(ha/日)": cap,
            "推定機械日数": mdays,
            "推定機械日数_有効": mdays,
            "source": "exception",
            "入力行Index": None,
        })

    return pd.DataFrame(rows)

def compute_all(df_sak: pd.DataFrame, df_mach: pd.DataFrame, df_tpl: pd.DataFrame, df_exc: pd.DataFrame, defaults: Defaults):
    cap_long = compute_farm_capacities(df_mach, defaults)
    sow_dist, sow_check = compute_sowplant_distribution(df_sak, cap_long, defaults)
    tpl_tasks = compute_template_tasks(df_sak, df_tpl, cap_long, defaults)
    exc_tasks = compute_exceptions(df_sak, df_exc, cap_long, defaults)

    tasks = pd.concat([tpl_tasks, sow_dist, exc_tasks], ignore_index=True)
    if "農機カテゴリ" in tasks.columns:
        tasks["農機カテゴリ"] = tasks["農機カテゴリ"].astype(str).str.strip()
        tasks = tasks[tasks["農機カテゴリ"].isin(VALID_MACHINE_CATS)].copy()
    tasks["推定機械日数_有効"] = pd.to_numeric(tasks["推定機械日数_有効"], errors="coerce").fillna(0.0)
    tasks["旬番号"] = pd.to_numeric(tasks["旬番号"], errors="coerce").astype("Int64")
    tasks = tasks.dropna(subset=["旬番号"]).copy()
    tasks["旬番号"] = tasks["旬番号"].astype(int)
    tasks["旬ラベル"] = tasks["旬番号"].map(lambda x: jn_to_label(int(x)))
    tasks["月"] = ((tasks["旬番号"] - 1) // 3 + 1).astype(int)
    return cap_long, tasks, sow_check

def build_load_table(tasks: pd.DataFrame, farm: str, group_by: str, utilization: float, year: Optional[int]=None, detail: bool=False) -> pd.DataFrame:
    """Compute required machine-days per Jun/Month by machine category and compare to available days."""
    df = tasks[tasks["農家名"] == farm].copy()
    if year is not None and "年" in df.columns:
        df = df[df["年"] == year].copy()
    if df.empty:
        return pd.DataFrame()

    if "面積(ha)" not in df.columns:
        df["面積(ha)"] = 0.0
    df["面積(ha)"] = pd.to_numeric(df["面積(ha)"], errors="coerce").fillna(0.0)

    if group_by == "旬":
        if detail:
            for c in ["作業グループ", "メモ", "作業"]:
                if c not in df.columns:
                    df[c] = ""
            g_detail = df.groupby(["年","旬番号","旬ラベル","農機カテゴリ","作業グループ","メモ"], as_index=False).agg(
                {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
            )
            g_total = df.groupby(["年","旬番号","旬ラベル","農機カテゴリ"], as_index=False).agg(
                {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
            )
            g_total["期間日数"] = g_total.apply(lambda r: jun_days(int(r["年"]), int(r["旬番号"])), axis=1)
            g_total["容量(機械日)"] = g_total["期間日数"] * float(utilization)
            g_total["利用率"] = np.where(g_total["容量(機械日)"] > 0, g_total["推定機械日数_有効"] / g_total["容量(機械日)"], np.nan)
            g = g_detail.merge(
                g_total[["年","旬番号","旬ラベル","農機カテゴリ","期間日数","容量(機械日)","利用率","推定機械日数_有効","面積(ha)"]].rename(
                    columns={"推定機械日数_有効": "推定機械日数_有効(合計)", "面積(ha)": "面積(ha)(合計)"}
                ),
                on=["年","旬番号","旬ラベル","農機カテゴリ"],
                how="left",
            )
            g["構成比"] = np.where(g["推定機械日数_有効(合計)"] > 0, g["推定機械日数_有効"] / g["推定機械日数_有効(合計)"], np.nan)
            return g.sort_values(["年","旬番号","農機カテゴリ","推定機械日数_有効"], ascending=[True, True, True, False]).reset_index(drop=True)

        g = df.groupby(["年","旬番号","旬ラベル","農機カテゴリ"], as_index=False).agg(
            {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
        )
        g["期間日数"] = g.apply(lambda r: jun_days(int(r["年"]), int(r["旬番号"])), axis=1)
        g["容量(機械日)"] = g["期間日数"] * float(utilization)
        g["利用率"] = np.where(g["容量(機械日)"] > 0, g["推定機械日数_有効"] / g["容量(機械日)"], np.nan)
        return g.sort_values(["年","旬番号","農機カテゴリ"]).reset_index(drop=True)

    # 月
    if detail:
        for c in ["作業グループ", "メモ", "作業"]:
            if c not in df.columns:
                df[c] = ""
        g_detail = df.groupby(["年","月","農機カテゴリ","作業グループ","メモ"], as_index=False).agg(
            {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
        )
        g_total = df.groupby(["年","月","農機カテゴリ"], as_index=False).agg(
            {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
        )
        g_total["期間日数"] = g_total.apply(lambda r: calendar.monthrange(int(r["年"]), int(r["月"]))[1], axis=1)
        g_total["容量(機械日)"] = g_total["期間日数"] * float(utilization)
        g_total["利用率"] = np.where(g_total["容量(機械日)"] > 0, g_total["推定機械日数_有効"] / g_total["容量(機械日)"], np.nan)
        g = g_detail.merge(
            g_total[["年","月","農機カテゴリ","期間日数","容量(機械日)","利用率","推定機械日数_有効","面積(ha)"]].rename(
                columns={"推定機械日数_有効": "推定機械日数_有効(合計)", "面積(ha)": "面積(ha)(合計)"}
            ),
            on=["年","月","農機カテゴリ"],
            how="left",
        )
        g["構成比"] = np.where(g["推定機械日数_有効(合計)"] > 0, g["推定機械日数_有効"] / g["推定機械日数_有効(合計)"], np.nan)
        return g.sort_values(["年","月","農機カテゴリ","推定機械日数_有効"], ascending=[True, True, True, False]).reset_index(drop=True)

    g = df.groupby(["年","月","農機カテゴリ"], as_index=False).agg(
        {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
    )
    g["期間日数"] = g.apply(lambda r: calendar.monthrange(int(r["年"]), int(r["月"]))[1], axis=1)
    g["容量(機械日)"] = g["期間日数"] * float(utilization)
    g["利用率"] = np.where(g["容量(機械日)"] > 0, g["推定機械日数_有効"] / g["容量(機械日)"], np.nan)
    return g.sort_values(["年","月","農機カテゴリ"]).reset_index(drop=True)

def build_capacity_check(tasks: pd.DataFrame, farm: str, group_by: str, utilization: float, year: Optional[int]=None) -> pd.DataFrame:
    """Return capacity check table across all tasks (template/sowplant/exception)."""
    load = build_load_table(tasks, farm=farm, group_by=group_by, utilization=utilization, year=year, detail=False)
    if load is None or load.empty:
        return pd.DataFrame()

    load = load.copy()
    load["不足(機械日)"] = np.where(
        load["容量(機械日)"] > 0,
        np.maximum(0.0, load["推定機械日数_有効"] - load["容量(機械日)"]),
        np.nan,
    )
    load["判定"] = np.where(load["不足(機械日)"].fillna(0.0) > 1e-9, "不足", "OK")
    return load

def suggest_shifts(
    df_sak: pd.DataFrame,
    df_mach: pd.DataFrame,
    df_tpl: pd.DataFrame,
    df_exc: pd.DataFrame,
    defaults: Defaults,
    farm: str,
    utilization: float,
    top_n: int = 10,
    deltas: Optional[List[int]] = None,
    group_by: str = "旬",
) -> pd.DataFrame:
    """
    Simple heuristic:
    - Compute current peak required rate (bottleneck across machine categories).
    - For top N sow/plant rows by area in the selected farm, try shifting rep/from/to by given deltas (days).
    - Keep shift within allowable range if 元入力(参考) gives range or if From/To are explicitly set.
    - Return best candidates that reduce the peak required rate.
    """
    if deltas is None:
        deltas = [-20, -10, 10, 20]

    cap_long, tasks0, _ = compute_all(df_sak, df_mach, df_tpl, df_exc, defaults)
    load0 = build_load_table(tasks0, farm=farm, group_by=group_by, utilization=utilization, detail=False)
    if load0.empty:
        return pd.DataFrame()
    load0 = load0.copy()
    load0["必要稼働率"] = np.where(
        pd.to_numeric(load0.get("期間日数"), errors="coerce") > 0,
        pd.to_numeric(load0.get("推定機械日数_有効"), errors="coerce") / pd.to_numeric(load0.get("期間日数"), errors="coerce"),
        np.nan,
    )
    peak0 = float(load0["必要稼働率"].max())
    peak0_row = load0.loc[load0["必要稼働率"].idxmax()]
    if group_by == "月":
        peak0_when = f'{int(peak0_row["年"])}年 {int(peak0_row["月"])}月'
    else:
        peak0_when = f'{int(peak0_row["年"])}年 {jn_to_label(int(peak0_row["旬番号"]))}'
    peak0_cat = str(peak0_row.get("農機カテゴリ", "")).strip()
    peak0_mdays = float(pd.to_numeric(peak0_row.get("推定機械日数_有効"), errors="coerce"))
    peak0_cap = float(pd.to_numeric(peak0_row.get("容量(機械日)"), errors="coerce"))
    peak0_days = float(pd.to_numeric(peak0_row.get("期間日数"), errors="coerce"))

    sak_f = df_sak.copy()
    sak_f["農家名"] = sak_f["農家名"].astype(str).str.strip()
    sak_f = sak_f[sak_f["農家名"] == farm].copy()
    sak_f["面積(ha)"] = pd.to_numeric(sak_f["面積(ha)"], errors="coerce")
    sak_f = sak_f[sak_f["面積(ha)"].fillna(0) > 0].copy()
    if sak_f.empty:
        return pd.DataFrame()

    # choose candidates: sow/plant meaningful rows (rep or range exists)
    cand = []
    for i, r in sak_f.reset_index(drop=False).iterrows():
        rng = sowplant_range_for_row(r, defaults)
        if rng is None:
            continue
        cand.append((int(r["index"]), float(r["面積(ha)"])))
    cand = sorted(cand, key=lambda x: x[1], reverse=True)[:max(1, int(top_n))]

    out = []
    for idx, area in cand:
        base_row = df_sak.loc[idx].copy()
        crop = str(base_row.get("作物","")).strip()

        # determine "used range" & "allowed range"
        rng = sowplant_range_for_row(base_row, defaults)
        if rng is None:
            continue
        used_from, used_to, _ = rng

        # allowed range: parse 元入力(参考) if available; otherwise no hard constraint
        # (播種/移植_開始日/終了日は編集対象なので、ここでは「固定制約」とみなさない)
        allow_from = None
        allow_to = None
        if base_row.get("元入力(参考)") is not None:
            base_year = used_from.year if used_from is not None else dt.date.today().year
            parsed = parse_jp_date_range(base_row.get("元入力(参考)"), base_year)
            if parsed:
                allow_from, allow_to = parsed

        for delta in deltas:
            # shift the whole window; if allowable range exists, keep within it
            new_from = used_from + dt.timedelta(days=delta)
            new_to = used_to + dt.timedelta(days=delta)
            if allow_from is not None and new_from < allow_from:
                continue
            if allow_to is not None and new_to > allow_to:
                continue

            df_sak2 = df_sak.copy()
            if "播種/移植日（代表日・調整）" in df_sak2.columns:
                df_sak2.loc[idx, "播種/移植日（代表日・調整）"] = mid_date(new_from, new_to)
            if "播種/移植_開始日（参考）" in df_sak2.columns:
                df_sak2.loc[idx, "播種/移植_開始日（参考）"] = new_from
            if "播種/移植_終了日（参考）" in df_sak2.columns:
                df_sak2.loc[idx, "播種/移植_終了日（参考）"] = new_to

            _, tasks1, _ = compute_all(df_sak2, df_mach, df_tpl, df_exc, defaults)
            load1 = build_load_table(tasks1, farm=farm, group_by=group_by, utilization=utilization, detail=False)
            if load1.empty:
                continue
            load1 = load1.copy()
            load1["必要稼働率"] = np.where(
                pd.to_numeric(load1.get("期間日数"), errors="coerce") > 0,
                pd.to_numeric(load1.get("推定機械日数_有効"), errors="coerce") / pd.to_numeric(load1.get("期間日数"), errors="coerce"),
                np.nan,
            )
            peak1 = float(load1["必要稼働率"].max())
            peak1_row = load1.loc[load1["必要稼働率"].idxmax()]
            if group_by == "月":
                peak1_when = f'{int(peak1_row["年"])}年 {int(peak1_row["月"])}月'
            else:
                peak1_when = f'{int(peak1_row["年"])}年 {jn_to_label(int(peak1_row["旬番号"]))}'
            peak1_cat = str(peak1_row.get("農機カテゴリ", "")).strip()
            peak1_mdays = float(pd.to_numeric(peak1_row.get("推定機械日数_有効"), errors="coerce"))
            peak1_cap = float(pd.to_numeric(peak1_row.get("容量(機械日)"), errors="coerce"))
            peak1_days = float(pd.to_numeric(peak1_row.get("期間日数"), errors="coerce"))
            out.append({
                "入力行Index": idx + 1,
                "作物": crop,
                "面積(ha)": area,
                "シフト(日)": delta,
                "播種/移植_From(現状)": used_from,
                "播種/移植_To(現状)": used_to,
                "播種/移植_From(提案)": new_from,
                "播種/移植_To(提案)": new_to,
                "ピーク必要稼働率(現状)": peak0,
                "ピーク必要稼働率(提案後)": peak1,
                "ピーク時期(現状)": peak0_when,
                "ボトルネック農機(現状)": peak0_cat,
                "ピーク時期(提案後)": peak1_when,
                "ボトルネック農機(提案後)": peak1_cat,
                "ピーク必要日数(現状,機械日)": peak0_mdays,
                "ピーク回せる日数(現状,機械日)": peak0_cap,
                "ピーク超過(現状,機械日)": peak0_mdays - peak0_cap,
                "ピーク期間日数(現状)": peak0_days,
                "ピーク必要日数(提案後,機械日)": peak1_mdays,
                "ピーク回せる日数(提案後,機械日)": peak1_cap,
                "ピーク超過(提案後,機械日)": peak1_mdays - peak1_cap,
                "ピーク期間日数(提案後)": peak1_days,
                "改善(Δ)": peak0 - peak1,
            })

    if not out:
        return pd.DataFrame()
    res = pd.DataFrame(out).sort_values(["改善(Δ)","面積(ha)"], ascending=[False, False]).reset_index(drop=True)
    return res

def write_back_to_excel(original_bytes: bytes, df_sak_edited: pd.DataFrame, extra_sheets: Dict[str, pd.DataFrame]) -> bytes:
    """
    Update the uploaded workbook:
    - overwrite '入力_作付' key input columns (same row alignment as pandas read header=3)
    - add extra output sheets
    """
    from io import BytesIO
    bio = BytesIO(original_bytes)
    wb = load_workbook(bio)
    if "入力_作付" not in wb.sheetnames:
        raise ValueError("入力_作付 シートが見つかりません。")

    ws = wb["入力_作付"]

    # find header row index: we assume header at row 4 (header=3 in pandas)
    header_row = 4
    # map header text -> column index
    header_map = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        header_map[str(v).strip()] = col

    target_cols = [
        "農家名", "作物", "面積(ha)",
        "播種/移植_開始日（参考）", "播種/移植_終了日（参考）", "播種/移植日（代表日・調整）",
        "作期日数(上書き)", "収穫日(上書き)",
        "元入力(参考)",
    ]
    for c in target_cols:
        if c not in header_map:
            # ignore missing, but keep going
            pass

    df = df_sak_edited.reset_index(drop=True).copy()
    # If representative date is missing but From/To exist, fill it with midpoint for compatibility.
    if (
        "播種/移植日（代表日・調整）" in df.columns
        and "播種/移植_開始日（参考）" in df.columns
        and "播種/移植_終了日（参考）" in df.columns
    ):
        rep = df["播種/移植日（代表日・調整）"].apply(safe_date)
        sdt = df["播種/移植_開始日（参考）"].apply(safe_date)
        edt = df["播種/移植_終了日（参考）"].apply(safe_date)
        for i in range(len(df)):
            if rep.iloc[i] is None and sdt.iloc[i] is not None and edt.iloc[i] is not None:
                df.loc[i, "播種/移植日（代表日・調整）"] = mid_date(sdt.iloc[i], edt.iloc[i])

    # write rows starting at row 5
    start_row = header_row + 1
    for i in range(len(df)):
        r_excel = start_row + i
        for c in target_cols:
            if c not in df.columns or c not in header_map:
                continue
            col = header_map[c]
            val = df.loc[i, c]
            # keep dates as datetime for excel
            if isinstance(val, dt.date) and not isinstance(val, dt.datetime):
                val = dt.datetime(val.year, val.month, val.day)
            ws.cell(row=r_excel, column=col, value=val)

    # add extra sheets
    for name, df_out in extra_sheets.items():
        if name in wb.sheetnames:
            del wb[name]
        ws2 = wb.create_sheet(title=name[:31])
        # write header
        for j, colname in enumerate(df_out.columns, start=1):
            ws2.cell(row=1, column=j, value=str(colname))
        # write data
        for i, row in enumerate(df_out.itertuples(index=False), start=2):
            for j, v in enumerate(row, start=1):
                if isinstance(v, dt.date) and not isinstance(v, dt.datetime):
                    v = dt.datetime(v.year, v.month, v.day)
                ws2.cell(row=i, column=j, value=v)
        # basic column widths
        for j in range(1, min(25, len(df_out.columns)) + 1):
            ws2.column_dimensions[get_column_letter(j)].width = 14

    out_bio = BytesIO()
    wb.save(out_bio)
    return out_bio.getvalue()

# ----------------------------
# Streamlit UI
# ----------------------------

st.set_page_config(page_title="CS Board Hybrid Viewer v2", layout="wide")
st.title("CS Board（Excel × Streamlit）労働負荷の可視化・調整・根拠づくり")

def _date_col_config(cols: List[str]):
    cfg = {}
    for c in cols:
        cfg[c] = st.column_config.DateColumn(format="YYYY-MM-DD")
    return cfg

def _format_evidence_table(
    df_in: pd.DataFrame,
    *,
    gran: str,
    utilization: float,
    beginner_mode: bool,
    detail: bool,
) -> Tuple[pd.DataFrame, Dict]:
    if df_in is None or df_in.empty:
        return df_in, {}

    df = df_in.copy()

    if "推定機械日数_有効" in df.columns:
        df["必要日数(機械日)"] = pd.to_numeric(df["推定機械日数_有効"], errors="coerce")
    if "容量(機械日)" in df.columns:
        df["回せる日数(機械日)"] = pd.to_numeric(df["容量(機械日)"], errors="coerce")
    if "期間日数" in df.columns:
        df["期間日数"] = pd.to_numeric(df["期間日数"], errors="coerce")

    df["必要稼働率"] = np.where(
        pd.to_numeric(df.get("期間日数"), errors="coerce") > 0,
        pd.to_numeric(df.get("必要日数(機械日)"), errors="coerce") / pd.to_numeric(df.get("期間日数"), errors="coerce"),
        np.nan,
    )
    df["超過(機械日)"] = pd.to_numeric(df.get("必要日数(機械日)"), errors="coerce") - pd.to_numeric(df.get("回せる日数(機械日)"), errors="coerce")
    df["超過(必要稼働率-稼働率)"] = pd.to_numeric(df.get("必要稼働率"), errors="coerce") - float(utilization)
    df["判定"] = np.where(pd.to_numeric(df["超過(機械日)"], errors="coerce") > 0, "超過", "OK")

    # labels
    if gran == "旬" and "旬ラベル" in df.columns:
        df["期間"] = df["旬ラベル"].astype(str)
    elif gran == "月" and "月" in df.columns:
        df["期間"] = df["月"].astype(int).astype(str) + "月"

    if detail and ("作業グループ" in df.columns or "メモ" in df.columns):
        g = df.get("作業グループ", "").fillna("").astype(str).str.strip()
        m = df.get("メモ", "").fillna("").astype(str).str.strip()
        m = m.where(~m.str.lower().isin({"nan", "none"}), "")
        df["作業（作業グループ/メモ）"] = np.where(m != "", g + "(" + m + ")", g)

    colcfg: Dict = {}
    if beginner_mode:
        df["必要稼働率(%)"] = pd.to_numeric(df["必要稼働率"], errors="coerce") * 100.0
        df["超過(%)"] = pd.to_numeric(df["超過(必要稼働率-稼働率)"], errors="coerce") * 100.0

        # keep only easy columns
        cols = []
        for c in ["年", "期間", "農機カテゴリ"]:
            if c in df.columns:
                cols.append(c)
        if detail and "作業（作業グループ/メモ）" in df.columns:
            cols.append("作業（作業グループ/メモ）")
        for c in ["面積(ha)", "能力(ha/日)", "必要日数(機械日)", "回せる日数(機械日)", "期間日数", "必要稼働率(%)", "超過(機械日)", "超過(%)", "判定"]:
            if c in df.columns:
                cols.append(c)
        for c in ["source", "入力行Index"]:
            if detail and c in df.columns:
                cols.append(c)
        df = df[cols].copy()

        colcfg.update({
            "面積(ha)": st.column_config.NumberColumn(format="%.2f"),
            "能力(ha/日)": st.column_config.NumberColumn(format="%.2f"),
            "必要日数(機械日)": st.column_config.NumberColumn(format="%.3f"),
            "回せる日数(機械日)": st.column_config.NumberColumn(format="%.2f"),
            "必要稼働率(%)": st.column_config.NumberColumn(format="%.0f"),
            "超過(機械日)": st.column_config.NumberColumn(format="%.2f"),
            "超過(%)": st.column_config.NumberColumn(format="%.0f"),
        })
    return df, colcfg

st.subheader("アップロード")
st.caption("まずはCS_boardのExcel（.xlsx）をアップロードしてください。")

uploaded = st.file_uploader("CS_boardのExcelをアップロード（.xlsx）", type=["xlsx"])

@st.cache_data(show_spinner=False)
def load_raw(file_bytes: bytes):
    from io import BytesIO
    bio = BytesIO(file_bytes)
    pre_raw = read_sheet(bio, "前提(仮置き)", header=None)
    defaults = parse_defaults(pre_raw)

    bio.seek(0)
    df_sak = read_sheet(bio, "入力_作付", header=3)
    bio.seek(0)
    df_mach = read_sheet(bio, "入力_農機", header=3)
    bio.seek(0)
    df_tpl = read_sheet(bio, "作業テンプレ(編集)", header=3)
    bio.seek(0)
    try:
        df_exc = read_sheet(bio, "例外入力(ここだけ)", header=2)
    except Exception:
        df_exc = pd.DataFrame()

    # normalize
    if "農家名" in df_sak.columns:
        df_sak["農家名"] = df_sak["農家名"].astype(str).str.strip()
    if "作物" in df_sak.columns:
        df_sak["作物"] = df_sak["作物"].astype(str).str.strip()

    return defaults, df_sak, df_mach, df_tpl, df_exc

if not uploaded:
    st.info("まずはExcelをアップロードしてください。")
    st.stop()

defaults, df_sak_raw, df_mach, df_tpl, df_exc = load_raw(uploaded.getvalue())

def _input_diagnostics(df_sak: pd.DataFrame, df_mach_in: pd.DataFrame, df_tpl_in: pd.DataFrame, df_exc_in: pd.DataFrame):
    required_sak = ["農家名", "作物", "面積(ha)"]
    required_mach = ["農家名"]
    required_tpl = ["作物", "作業グループ", "農機カテゴリ", "有効(1/0)"]
    critical = []
    warn = []

    if df_sak is None or df_sak.empty:
        critical.append("`入力_作付` が空です（作業を配置できません）。")
    else:
        missing = [c for c in required_sak if c not in df_sak.columns]
        if missing:
            critical.append(f"`入力_作付` の必須列が見つかりません: {', '.join(missing)}")

    if df_mach_in is None or df_mach_in.empty:
        warn.append("`入力_農機` が空です（能力がデフォルトに寄る可能性があります）。")
    else:
        missing = [c for c in required_mach if c not in df_mach_in.columns]
        if missing:
            critical.append(f"`入力_農機` の必須列が見つかりません: {', '.join(missing)}")

    if df_tpl_in is None or df_tpl_in.empty:
        warn.append("`作業テンプレ(編集)` が空です（付随作業が出ません）。")
    else:
        missing = [c for c in required_tpl if c not in df_tpl_in.columns]
        if missing:
            critical.append(f"`作業テンプレ(編集)` の必須列が見つかりません: {', '.join(missing)}")

    if df_exc_in is not None and not df_exc_in.empty:
        if "有効(1/0)" not in df_exc_in.columns:
            warn.append("`例外入力(ここだけ)` に `有効(1/0)` 列がないため、一部の例外作業が反映されない可能性があります。")

    summary = {}
    try:
        if df_sak is not None and not df_sak.empty and "農家名" in df_sak.columns:
            farms = df_sak["農家名"].dropna().astype(str).str.strip()
            summary["農家数"] = int(farms.nunique())
        if df_sak is not None and not df_sak.empty and "面積(ha)" in df_sak.columns:
            area = pd.to_numeric(df_sak["面積(ha)"], errors="coerce").fillna(0.0)
            summary["合計面積(ha)"] = float(area.sum())
        if df_sak is not None and not df_sak.empty and "作物" in df_sak.columns:
            crops = df_sak["作物"].dropna().astype(str).str.strip()
            summary["作物数"] = int(crops.nunique())
        if df_sak is not None and not df_sak.empty and ("播種/移植_開始日（参考）" in df_sak.columns) and ("播種/移植_終了日（参考）" in df_sak.columns):
            s0 = pd.to_datetime(df_sak["播種/移植_開始日（参考）"], errors="coerce").dt.date
            e0 = pd.to_datetime(df_sak["播種/移植_終了日（参考）"], errors="coerce").dt.date
            smin = s0.dropna().min() if not s0.dropna().empty else None
            emax = e0.dropna().max() if not e0.dropna().empty else None
            if smin and emax:
                summary["播種/移植(参考)の全体レンジ"] = f"{smin}〜{emax}"
    except Exception:
        pass
    return critical, warn, summary

crit_msgs, warn_msgs, summary = _input_diagnostics(df_sak_raw, df_mach, df_tpl, df_exc)

if crit_msgs:
    st.error("Excelの形式が想定と一致しないため停止しました。")
    for m in crit_msgs:
        st.write(f"- {m}")
    st.stop()
if warn_msgs:
    with st.expander("入力チェック（注意）", expanded=False):
        for m in warn_msgs:
            st.write(f"- {m}")

with st.expander("このアプリの前提・計算ロジック（まとめ）", expanded=False):
    st.markdown(
        """
### このアプリで分かること（目的）
- 旬別/⽉別に「農機が回せるか？」を見える化します。
- 「この旬は手一杯なので、播種/移植の期間をずらそう」と判断するための根拠（数値と内訳）を作ります。

### 使い方（最短）
1. サイドバーで農家を選ぶ
2. ピーク確認で「赤線（農機稼働率％）を超える旬/月」を探す
3. 播種/移植の From/To を動かして、ピークが下がるか比較する（現状 vs 編集後）

### 対象（重要）
- **農機を使う作業のみ**が対象です（手作業・資材準備などはここでは扱いません）。
- 対象農機カテゴリ：`tractor / seeder / transplanter / sprayer / combine / roller`

### 入力（Excel）
- `入力_作付`：作物・面積・播種/移植のFrom/To（または元入力レンジ）
- `入力_農機`：農機能力（ha/日）。未入力は `前提(仮置き)` × 推定台数で補完します
- `作業テンプレ(編集)`：付随作業（どの作業を、播種/収穫のどの時期に置くか）
- `例外入力(ここだけ)`：テンプレに無い作業の追加

### 計算のルール（ここが根拠）
- **推定機械日数**（機械を何日動かす必要があるか）
  - `推定機械日数 = 面積(ha) ÷ 農機能力(ha/日)`
- **必要稼働率**（その旬/月のうち、何%の日数で動かす必要があるか）
  - `必要稼働率 = 推定機械日数 ÷ 期間日数`
- **From/Toがある作業は、期間内に日数比で均等配分**して旬/月へ割り当てます（平均化の仮定）。
- 播種/移植がFrom/Toの範囲なら、収穫も同じ幅の範囲として扱います（作期日数で前後にずれます）。

### グラフの見方
- y軸：必要稼働率（%）
- 赤線：農機稼働率％（「現実に動かせる割合」の見積り）
  - `必要稼働率 > 赤線` は「その旬/月に回し切れない」目安です
- 農機カテゴリが「全て」のとき：各旬/月で **ボトルネック農機だけ**を表示します（全量の合算ではありません）

### 面積(ha)の注意（よく混乱します）
- 散布など「同じ圃場を複数回」作業する場合があります。
- このとき `面積(ha)` は **延べ面積（作業量）**として増えることがあります（作付面積そのものとは別物です）。
        """
    )

# Session state for edits
if "df_sak_edit" not in st.session_state:
    st.session_state.df_sak_edit = df_sak_raw.copy()

# Sidebar
st.sidebar.header("設定（必要最低限）")
farms_all = sorted(st.session_state.df_sak_edit["農家名"].dropna().unique().tolist()) if "農家名" in st.session_state.df_sak_edit.columns else []
farm_sel = st.sidebar.selectbox("農家", farms_all) if farms_all else None

gran = st.sidebar.radio("集計粒度", ["旬", "月"], horizontal=True)
breakdown = st.sidebar.radio("内訳", ["作業別", "作物別"], horizontal=True)
st.sidebar.caption("※ 作業別は『作業グループ(メモ)』単位で、作物をまたいで同じ作業を合算します（例：直播2作物の播種は足し算）。")
compare_mode = st.sidebar.checkbox("比較表示", value=True)
beginner_mode = True

utilization = st.sidebar.slider("農機稼働率％（天候・段取りロス込み）", min_value=40, max_value=100, value=70, step=5) / 100.0
with st.sidebar.expander("稼働率とは？（注意点）", expanded=False):
    st.caption("目安：週休2日制で農機も止まる前提なら上限は概ね `5/7 ≒ 71%`。天候・段取り・移動・整備も考慮して下げて設定します。")
    st.caption("※ 本アプリは「農機を使う作業のみ」を対象に「機械稼働（ha/日）→ 推定機械日数」を計算します（手作業・資材準備などは除外）。ピークは農機カテゴリ別のボトルネックで判定します。")

# Recompute baseline and edited
cap_long, tasks_edit, sow_check = compute_all(st.session_state.df_sak_edit, df_mach, df_tpl, df_exc, defaults)
tasks_edit = tasks_edit[tasks_edit["source"].isin(ALL_SOURCES)].copy()

_, tasks_base, _ = compute_all(df_sak_raw, df_mach, df_tpl, df_exc, defaults)
tasks_base = tasks_base[tasks_base["source"].isin(ALL_SOURCES)].copy()

st.subheader("生産者")
st.caption("左のサイドバーで農家を選択し、ピーク確認→調整→ダウンロードの順に進めます。")
if not farm_sel:
    st.info("左のサイドバーから農家を選択してください。")
else:
    st.success(f"選択中の農家：{farm_sel}")
    df_sak_f = st.session_state.df_sak_edit.copy()
    if "農家名" in df_sak_f.columns:
        df_sak_f = df_sak_f[df_sak_f["農家名"].astype(str).str.strip() == str(farm_sel).strip()].copy()

    # ---- Crop summary
    crop_sum = pd.DataFrame(columns=["作物", "面積(ha)"])
    farm_area_total = np.nan
    farm_crops_n = np.nan
    top_crop_text = "—"
    try:
        if ("作物" in df_sak_f.columns) and ("面積(ha)" in df_sak_f.columns) and (not df_sak_f.empty):
            df_sak_f["作物"] = df_sak_f["作物"].astype(str).str.strip()
            df_sak_f["面積(ha)"] = pd.to_numeric(df_sak_f["面積(ha)"], errors="coerce").fillna(0.0)
            farm_area_total = float(df_sak_f["面積(ha)"].sum())
            farm_crops_n = int(df_sak_f["作物"].dropna().astype(str).str.strip().nunique())

            crop_sum = (
                df_sak_f.dropna(subset=["作物"])
                .groupby("作物", as_index=False)["面積(ha)"]
                .sum()
                .sort_values("面積(ha)", ascending=False)
            )
            crop_sum = crop_sum[crop_sum["面積(ha)"].fillna(0) > 0].copy()
            if not crop_sum.empty:
                top_crop_text = f'{crop_sum.iloc[0]["作物"]}（{float(crop_sum.iloc[0]["面積(ha)"]):.1f}ha）'
    except Exception:
        pass

    # ---- Machine profile
    mach_profile = pd.DataFrame(columns=["農機", "推定台数(仮)", "1日あたりの作業面積(ha/日)"])
    mach_n = 0
    weakest_text = "—"
    try:
        mach_rows = []
        mach_map = {
            "tractor": ("トラクタ", "トラクタ推定台数(仮)"),
            "seeder": ("播種機", "播種機推定台数(仮)"),
            "transplanter": ("田植機", "田植機推定台数(仮)"),
            "sprayer": ("防除機", "防除推定台数(仮)"),
            "combine": ("コンバイン", "コンバイン推定台数(仮)"),
            "roller": ("鎮圧機", "鎮圧推定台数(仮)"),
        }

        cap_f = cap_long[
            cap_long["農家名"].astype(str).str.strip() == str(farm_sel).strip()
        ].copy() if cap_long is not None else pd.DataFrame()
        cap_f = cap_f.dropna(subset=["農機カテゴリ"]).copy() if not cap_f.empty else cap_f
        if not cap_f.empty:
            cap_f["農機カテゴリ"] = cap_f["農機カテゴリ"].astype(str).str.strip()
        cap_map = cap_f.set_index("農機カテゴリ")["能力(ha/日)"].to_dict() if not cap_f.empty else {}

        df_mach_f = df_mach.copy() if df_mach is not None else pd.DataFrame()
        if "農家名" in df_mach_f.columns:
            df_mach_f["農家名"] = df_mach_f["農家名"].astype(str).str.strip()
            df_mach_f = df_mach_f[df_mach_f["農家名"] == str(farm_sel).strip()].copy()
        df_mach_row = df_mach_f.iloc[0] if df_mach_f is not None and not df_mach_f.empty else None

        for cat in sorted(list(VALID_MACHINE_CATS)):
            label, n_col = mach_map.get(cat, (cat, None))
            cap_val = cap_map.get(cat, np.nan)
            if pd.isna(cap_val) or float(cap_val) <= 0:
                continue

            n_val = np.nan
            if df_mach_row is not None and n_col and n_col in df_mach_row.index:
                n_val = pd.to_numeric(df_mach_row.get(n_col), errors="coerce")

            mach_rows.append(
                {
                    "農機": label,
                    "推定台数(仮)": (float(n_val) if not pd.isna(n_val) else np.nan),
                    "1日あたりの作業面積(ha/日)": float(cap_val),
                }
            )

        if mach_rows:
            mach_profile = pd.DataFrame(mach_rows).sort_values("1日あたりの作業面積(ha/日)", ascending=True)
            mach_n = int(len(mach_profile))
            weakest_text = f'{mach_profile.iloc[0]["農機"]}（{float(mach_profile.iloc[0]["1日あたりの作業面積(ha/日)"]):.2f}ha/日）'
    except Exception:
        pass

    # ---- Summary cards (scan first, details later)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("合計面積(ha)", f"{farm_area_total:.1f}" if not pd.isna(farm_area_total) else "—")
    c2.metric("作物数", f"{farm_crops_n:d}" if not pd.isna(farm_crops_n) else "—")
    c3.metric("最大の作物", top_crop_text)
    c4.metric("農機の種類", f"{mach_n:d}" if mach_n else "—")
    st.caption(f"弱い農機の目安（能力が小さい）: {weakest_text}")

    tab_crop, tab_mach = st.tabs(["作物サマリー", "農機プロフィール"])
    with tab_crop:
        if crop_sum is None or crop_sum.empty:
            st.info("作物別の面積サマリーを表示できません（`入力_作付` の作物/面積をご確認ください）。")
        else:
            st.dataframe(
                crop_sum,
                use_container_width=True,
                height=320,
                hide_index=True,
                column_config={"面積(ha)": st.column_config.NumberColumn(format="%.2f")},
            )

    with tab_mach:
        if mach_profile is None or mach_profile.empty:
            st.info("農機プロフィールを表示できません（`入力_農機` の農機能力をご確認ください）。")
        else:
            mp = mach_profile.copy().sort_values("1日あたりの作業面積(ha/日)", ascending=False)
            st.caption("推定台数(仮)が空の場合は未入力です。能力(ha/日)は入力値が無い場合、前提(仮置き)×推定台数で補完した値になります。")
            st.dataframe(
                mp,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "推定台数(仮)": st.column_config.NumberColumn(format="%.0f"),
                    "1日あたりの作業面積(ha/日)": st.column_config.NumberColumn(format="%.2f"),
                },
            )

# ----------------------------
# 2) Capacity check & peak detection
# ----------------------------
st.subheader("ピーク確認（ピーク検出／根拠の見える化）")
st.caption("赤線（農機稼働率％）を超える旬/月があるかを確認します。必要なら③で播種/移植日を動かします。")

if not farm_sel:
    st.info("農家を選択してください。")
else:
    load_edit_total = build_load_table(tasks_edit, farm=farm_sel, group_by=gran, utilization=utilization, detail=False)
    if load_edit_total.empty:
        st.warning("データがありません。入力やフィルター条件をご確認ください。")
    else:
        load_edit_total = load_edit_total.copy()
        load_edit_total["必要稼働率"] = np.where(
            pd.to_numeric(load_edit_total.get("期間日数"), errors="coerce") > 0,
            pd.to_numeric(load_edit_total.get("推定機械日数_有効"), errors="coerce") / pd.to_numeric(load_edit_total.get("期間日数"), errors="coerce"),
            np.nan,
        )
        peak = float(load_edit_total["必要稼働率"].max())
        peak_row = load_edit_total.loc[load_edit_total["必要稼働率"].idxmax()]
        peak_cat = str(peak_row.get("農機カテゴリ", "")).strip()
        if gran == "旬":
            peak_label = f'{int(peak_row["年"])}年 {jn_to_label(int(peak_row["旬番号"]))}'
        else:
            peak_label = f'{int(peak_row["年"])}年 {int(peak_row["月"])}月'

        if compare_mode:
            load_base_total = build_load_table(tasks_base, farm=farm_sel, group_by=gran, utilization=utilization, detail=False)
            if not load_base_total.empty:
                load_base_total = load_base_total.copy()
                load_base_total["必要稼働率"] = np.where(
                    pd.to_numeric(load_base_total.get("期間日数"), errors="coerce") > 0,
                    pd.to_numeric(load_base_total.get("推定機械日数_有効"), errors="coerce") / pd.to_numeric(load_base_total.get("期間日数"), errors="coerce"),
                    np.nan,
                )
                peak0 = float(load_base_total["必要稼働率"].max())
                peak1 = float(load_edit_total["必要稼働率"].max())
                col_m1, col_m2, col_m3 = st.columns([1, 1, 1])
                col_m1.metric("ピーク必要稼働率（現在）", f"{peak0:.0%}")
                col_m2.metric("ピーク必要稼働率（編集後）", f"{peak1:.0%}")
                col_m3.metric("改善（編集前後）", f"{(peak0 - peak1):.0%}")
            else:
                st.metric("ピーク必要稼働率（現在）", f"{peak:.0%}")
        else:
            st.metric(
                "ピーク必要稼働率（現在）",
                f"{peak:.0%}",
                help="赤線（農機稼働率％）を超えると『その粒度で回し切れない』目安（稼働率の仮定に依存）",
            )

        if float(peak) > float(utilization):
            try:
                # 具体例（試し）：ピーク期間・ボトルネック農機への寄与が大きい作付を1つ選び、前後どちらかにずらす
                if gran == "旬":
                    period_name = "旬"
                    key_now = (int(peak_row["年"]), int(peak_row["旬番号"]))
                    period_max = 36
                    t_peak = tasks_edit[
                        (tasks_edit["農家名"].astype(str).str.strip() == str(farm_sel).strip())
                        & (tasks_edit["農機カテゴリ"].astype(str).str.strip() == str(peak_cat).strip())
                        & (tasks_edit["年"] == int(peak_row["年"]))
                        & (tasks_edit["旬番号"] == int(peak_row["旬番号"]))
                    ].copy()
                    shift_days = 7
                else:
                    period_name = "月"
                    key_now = (int(peak_row["年"]), int(peak_row["月"]))
                    period_max = 12
                    t_peak = tasks_edit[
                        (tasks_edit["農家名"].astype(str).str.strip() == str(farm_sel).strip())
                        & (tasks_edit["農機カテゴリ"].astype(str).str.strip() == str(peak_cat).strip())
                        & (tasks_edit["年"] == int(peak_row["年"]))
                        & (tasks_edit["月"] == int(peak_row["月"]))
                    ].copy()
                    shift_days = 14

                def _shift_key(key: Tuple[int, int], delta: int) -> Tuple[int, int]:
                    y, p = key
                    p2 = p + delta
                    y2 = y
                    while p2 < 1:
                        y2 -= 1
                        p2 += period_max
                    while p2 > period_max:
                        y2 += 1
                        p2 -= period_max
                    return y2, p2

                # 前後の余裕から「前倒し/後ろ倒し」を決める（両方なければ前倒し）
                load_cat = load_edit_total[load_edit_total["農機カテゴリ"].astype(str).str.strip() == str(peak_cat).strip()].copy()
                rates: Dict[Tuple[int, int], float] = {}
                if not load_cat.empty:
                    if gran == "旬":
                        for _, rr in load_cat.iterrows():
                            rates[(int(rr["年"]), int(rr["旬番号"]))] = float(rr["必要稼働率"])
                    else:
                        for _, rr in load_cat.iterrows():
                            rates[(int(rr["年"]), int(rr["月"]))] = float(rr["必要稼働率"])

                k_prev = _shift_key(key_now, -1)
                k_next = _shift_key(key_now, +1)
                r_prev = rates.get(k_prev, None)
                r_next = rates.get(k_next, None)
                spare_prev = (float(utilization) - float(r_prev)) if r_prev is not None else None
                spare_next = (float(utilization) - float(r_next)) if r_next is not None else None
                suggest = "前倒し"
                if spare_prev is not None and spare_next is not None:
                    suggest = "前倒し" if spare_prev >= spare_next else "後ろ倒し"
                elif spare_next is not None and spare_prev is None:
                    suggest = "後ろ倒し"

                if t_peak is not None and not t_peak.empty and "入力行Index" in t_peak.columns:
                    top = (
                        t_peak.groupby("入力行Index", as_index=False)["推定機械日数_有効"]
                        .sum()
                        .sort_values("推定機械日数_有効", ascending=False)
                        .head(1)
                    )
                    if not top.empty:
                        input_idx = int(top.iloc[0]["入力行Index"])
                        df_sak_all = st.session_state.df_sak_edit.copy().reset_index(drop=True)
                        if 1 <= input_idx <= len(df_sak_all):
                            row_sak = df_sak_all.iloc[input_idx - 1].copy()
                            crop = str(row_sak.get("作物", "")).strip()
                            rng = sowplant_range_for_row(row_sak, defaults)
                            if rng is not None:
                                cur_from, cur_to, _ = rng
                                delta = -shift_days if suggest == "前倒し" else shift_days
                                rec_from = cur_from + dt.timedelta(days=delta)
                                rec_to = cur_to + dt.timedelta(days=delta)
                                label = f"{input_idx}行目 {crop}" if crop else f"{input_idx}行目"
                                st.info(
                                    f"ヒント：{label} の From/To を {cur_from}〜{cur_to} → {rec_from}〜{rec_to}（{suggest}）にして「プレビュー更新」を押してみてください。"
                                    f"（ピークの前後{period_name}の余裕から方向を推定）"
                                )
            except Exception:
                pass
        else:
            st.success("赤線オーバーは見つかりませんでした（必要稼働率が農機稼働率％以下）。")

        # Add top contributing work_group/memo into the evidence (total) table
        load_tbl_detail = build_load_table(tasks_edit, farm=farm_sel, group_by=gran, utilization=utilization, detail=True)
        if not load_tbl_detail.empty:
            if gran == "旬":
                key_cols = ["年", "旬番号", "農機カテゴリ"]
            else:
                key_cols = ["年", "月", "農機カテゴリ"]
            top_reason = (
                load_tbl_detail.sort_values("推定機械日数_有効", ascending=False)
                .dropna(subset=key_cols)
                .groupby(key_cols, as_index=False)
                .first()[key_cols + ["作業グループ", "メモ"]]
                .rename(columns={"作業グループ": "主な作業グループ", "メモ": "主なメモ"})
            )
            load_edit_total = load_edit_total.merge(top_reason, on=key_cols, how="left")

        with st.expander("根拠テーブル", expanded=False):
            df_show, cfg = _format_evidence_table(
                load_edit_total,
                gran=gran,
                utilization=utilization,
                beginner_mode=beginner_mode,
                detail=False,
            )
            if beginner_mode:
                st.caption("見方：赤線（農機稼働率%）より `必要稼働率(%)` が大きい期間が『超過』です。")
            st.dataframe(df_show, use_container_width=True, height=260, column_config=cfg, hide_index=True)

        peak_year = int(peak_row["年"])
        if gran == "旬":
            peak_jun_no = int(peak_row["旬番号"])
        else:
            peak_jun_no = None

        # Manual shift tool (meter)
        with st.expander("播種/移植日をメーターで調整（手動）", expanded=(float(peak) > float(utilization))):
                col_s1, col_s2, col_s3, col_s4 = st.columns([1.2, 1, 1, 1])
                with col_s1:
                    st.metric("現在のボトルネック", peak_label)
                with col_s2:
                    st.metric("農機（ボトルネック）", peak_cat if peak_cat else "-")
                with col_s3:
                    st.metric("ピーク必要稼働率", f"{peak:.0%}")
                with col_s4:
                    st.metric("農機稼働率（設定）", f"{utilization:.0%}")

                st.caption("作付行を選んで、シフト量（メーター）を動かします。プレビューで効果を確認してから適用します。")
                if peak_cat:
                    # --- Hint: direction (earlier/later) based on spare capacity around the peak period ---
                    try:
                        if gran == "旬":
                            key_now = (int(peak_year), int(peak_jun_no))
                            period_max = 36
                            period_name = "旬"
                        else:
                            key_now = (int(peak_year), int(peak_row.get("月")))
                            period_max = 12
                            period_name = "月"

                        def _shift_key(key: Tuple[int, int], delta: int) -> Tuple[int, int]:
                            y, p = key
                            p2 = p + delta
                            y2 = y
                            while p2 < 1:
                                y2 -= 1
                                p2 += period_max
                            while p2 > period_max:
                                y2 += 1
                                p2 -= period_max
                            return y2, p2

                        load_cat = load_edit_total[load_edit_total["農機カテゴリ"].astype(str).str.strip() == str(peak_cat).strip()].copy()
                        rates = {}
                        if not load_cat.empty:
                            if gran == "旬":
                                for _, rr in load_cat.iterrows():
                                    rates[(int(rr["年"]), int(rr["旬番号"]))] = float(rr["必要稼働率"])
                            else:
                                for _, rr in load_cat.iterrows():
                                    rates[(int(rr["年"]), int(rr["月"]))] = float(rr["必要稼働率"])

                        k_prev = _shift_key(key_now, -1)
                        k_next = _shift_key(key_now, +1)
                        r_prev = rates.get(k_prev, None)
                        r_next = rates.get(k_next, None)
                        spare_prev = (float(utilization) - float(r_prev)) if r_prev is not None else None
                        spare_next = (float(utilization) - float(r_next)) if r_next is not None else None

                        if spare_prev is not None or spare_next is not None:
                            suggest = None
                            if spare_prev is not None and spare_next is not None:
                                suggest = "前倒し" if spare_prev >= spare_next else "後ろ倒し"
                            elif spare_prev is not None:
                                suggest = "前倒し"
                            else:
                                suggest = "後ろ倒し"

                            parts = []
                            if spare_prev is not None:
                                parts.append(f"前の{period_name}：必要稼働率 {r_prev:.0%}（余裕 {spare_prev:.0%}）")
                            if spare_next is not None:
                                parts.append(f"次の{period_name}：必要稼働率 {r_next:.0%}（余裕 {spare_next:.0%}）")
                            st.info(
                                f"ヒント：ピーク（{peak_label}／{peak_cat}）を下げるには、From/To を **{suggest}** 方向に動かすと効きやすい場合があります。"
                                + (" / " + " / ".join(parts) if parts else "")
                            )
                    except Exception:
                        pass

                    # --- Hint: which sow/plant rows are impactful in this peak (by machine-days) ---
                    try:
                        t_peak = tasks_edit[
                            (tasks_edit["農家名"].astype(str).str.strip() == str(farm_sel).strip())
                            & (tasks_edit["農機カテゴリ"].astype(str).str.strip() == str(peak_cat).strip())
                        ].copy()
                        if gran == "旬":
                            t_peak = t_peak[(t_peak["年"] == int(peak_year)) & (t_peak["旬番号"] == int(peak_jun_no))].copy()
                        else:
                            t_peak = t_peak[(t_peak["年"] == int(peak_year)) & (t_peak["月"] == int(peak_row.get("月")))].copy()

                        if (t_peak is not None) and (not t_peak.empty) and ("入力行Index" in t_peak.columns):
                            g = (
                                t_peak.groupby("入力行Index", as_index=False)["推定機械日数_有効"]
                                .sum()
                                .sort_values("推定機械日数_有効", ascending=False)
                                .head(3)
                            )
                            if not g.empty:
                                df_all = st.session_state.df_sak_edit.copy().reset_index(drop=True)
                                df_all["入力行Index"] = df_all.index + 1
                                base = df_all[["入力行Index"] + [c for c in ["作物", "面積(ha)"] if c in df_all.columns]].copy()
                                g = g.merge(base, on="入力行Index", how="left")
                                hints = []
                                for _, rr in g.iterrows():
                                    crop = str(rr.get("作物", "")).strip()
                                    md = float(pd.to_numeric(rr.get("推定機械日数_有効"), errors="coerce"))
                                    ii = int(rr["入力行Index"])
                                    if crop:
                                        hints.append(f"{ii}行目 {crop}（{md:.2f}機械日）")
                                    else:
                                        hints.append(f"{ii}行目（{md:.2f}機械日）")
                    except Exception:
                        pass

                # 1) candidate rows (sow/plant range exists)
                cand_src = st.session_state.df_sak_edit.copy().reset_index(drop=True)
                cand_src = cand_src[cand_src["農家名"].astype(str).str.strip() == str(farm_sel).strip()].copy()
                cand_src["面積(ha)"] = pd.to_numeric(cand_src.get("面積(ha)"), errors="coerce")
                cand_src = cand_src[cand_src["面積(ha)"].fillna(0) > 0].copy()

                cand_rows = []
                for _, r in cand_src.reset_index(drop=False).iterrows():
                    rng = sowplant_range_for_row(r, defaults)
                    if rng is None:
                        continue
                    used_from, used_to, _ = rng
                    cand_rows.append(
                        {
                            "対象": False,
                            "_row_idx": int(r["index"]),
                            "入力行Index": int(r["index"]) + 1,
                            "作物": str(r.get("作物", "")).strip(),
                            "面積(ha)": float(r.get("面積(ha)", np.nan)),
                            "播種/移植_From": used_from,
                            "播種/移植_To": used_to,
                        }
                    )

                if not cand_rows:
                    st.warning("候補がありません（播種/移植の範囲がない、または面積が0の可能性があります）。")
                else:
                    cand_df = pd.DataFrame(cand_rows).sort_values("面積(ha)", ascending=False).reset_index(drop=True)
                    # keep date as dt.date for display
                    cand_df["播種/移植_From"] = pd.to_datetime(cand_df["播種/移植_From"], errors="coerce").dt.date
                    cand_df["播種/移植_To"] = pd.to_datetime(cand_df["播種/移植_To"], errors="coerce").dt.date

                st.markdown("**① 動かす作付行を選ぶ**")
                # single-select "radio-like" in the table (Streamlit doesn't support radio cells)
                sel_key = f"manual_shift_selected_row_{farm_sel}"
                prev_selected_row_idx = st.session_state.get(sel_key, None)
                if prev_selected_row_idx is not None:
                    cand_df["対象"] = cand_df["_row_idx"] == int(prev_selected_row_idx)

                edited_cand = st.data_editor(
                    cand_df[["対象", "入力行Index", "作物", "面積(ha)", "播種/移植_From", "播種/移植_To"]],
                    use_container_width=True,
                    height=260,
                    key=f"manual_shift_select_{farm_sel}",
                    disabled=["入力行Index", "作物", "面積(ha)", "播種/移植_From", "播種/移植_To"],
                    column_config=_date_col_config(["播種/移植_From", "播種/移植_To"]),
                    hide_index=True,
                )

                picked = edited_cand.loc[edited_cand["対象"] == True, "入力行Index"].tolist() if "対象" in edited_cand.columns else []
                # Enforce single selection: keep the newly selected row if multiple are checked.
                if len(picked) > 1:
                    prev_input_idx = (int(prev_selected_row_idx) + 1) if prev_selected_row_idx is not None else None
                    new_picked = [x for x in picked if prev_input_idx is None or x != prev_input_idx]
                    keep_input_idx = int(new_picked[0]) if new_picked else int(picked[0])
                    # update session and rerun to reflect auto-uncheck
                    st.session_state[sel_key] = int(keep_input_idx) - 1
                    st.warning("1つだけ選べます（最後に選んだ行だけを対象にしました）。")
                    st.rerun()
                elif len(picked) == 1:
                    st.session_state[sel_key] = int(picked[0]) - 1
                else:
                    st.session_state[sel_key] = None

                sel_row_idx = st.session_state.get(sel_key, None)
                if sel_row_idx is None:
                    st.info("①で『対象』を1つだけ選んでください。")
                    sel_base = cand_df.iloc[0:0].copy()
                else:
                    sel_base = cand_df[cand_df["_row_idx"] == int(sel_row_idx)].copy()

                st.markdown("**② 播種/移植の範囲をバーで上書き**")
                if sel_base.empty:
                    new_from = None
                    new_to = None
                else:
                    ref_row = sel_base.iloc[0]
                    ref_from = pd.to_datetime(ref_row["播種/移植_From"], errors="coerce").date()
                    ref_to = pd.to_datetime(ref_row["播種/移植_To"], errors="coerce").date()

                    def _month_start(d: dt.date) -> dt.date:
                        return dt.date(d.year, d.month, 1)

                    def _month_end(d: dt.date) -> dt.date:
                        return dt.date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])

                    buf_days = 60
                    min_d = _month_start(ref_from - dt.timedelta(days=buf_days))
                    max_d = _month_end(ref_to + dt.timedelta(days=buf_days))

                    new_from, new_to = st.slider(
                        "新しい播種/移植の範囲（From/To）",
                        min_value=min_d,
                        max_value=max_d,
                        value=(ref_from, ref_to),
                        key=f"manual_shift_range_{farm_sel}",
                    )
                    if isinstance(new_from, dt.date) and isinstance(new_to, dt.date) and new_to < new_from:
                        st.error("To が From より前です。範囲を選び直してください。")
                    st.caption(f"この作付行の From/To を {ref_from}〜{ref_to} → {new_from}〜{new_to} に上書きします。")

                def _build_shifted_df(new_from_d: Optional[dt.date], new_to_d: Optional[dt.date]) -> Tuple[pd.DataFrame, pd.DataFrame]:
                    df_all = st.session_state.df_sak_edit.copy().reset_index(drop=True)
                    if sel_row_idx is None:
                        return df_all, pd.DataFrame()
                    sel_idxs = [int(sel_row_idx)]
                    if not isinstance(new_from_d, dt.date) or not isinstance(new_to_d, dt.date):
                        return df_all, pd.DataFrame()
                    if new_to_d < new_from_d:
                        return df_all, pd.DataFrame()

                    preview_rows = []
                    for ridx in sel_idxs:
                        row = df_all.loc[ridx].copy()
                        rng = sowplant_range_for_row(row, defaults)
                        if rng is None:
                            continue
                        used_from, used_to, _ = rng
                        new_from = new_from_d
                        new_to = new_to_d

                        preview_rows.append(
                            {
                                "入力行Index": int(ridx) + 1,
                                "作物": str(row.get("作物", "")).strip(),
                                "面積(ha)": float(pd.to_numeric(row.get("面積(ha)"), errors="coerce")) if not pd.isna(pd.to_numeric(row.get("面積(ha)"), errors="coerce")) else 0.0,
                                "From(現状)": used_from,
                                "To(現状)": used_to,
                                "From(変更後)": new_from,
                                "To(変更後)": new_to,
                            }
                        )
                        if "播種/移植_開始日（参考）" in df_all.columns:
                            df_all.loc[ridx, "播種/移植_開始日（参考）"] = new_from
                        if "播種/移植_終了日（参考）" in df_all.columns:
                            df_all.loc[ridx, "播種/移植_終了日（参考）"] = new_to
                        if "播種/移植日（代表日・調整）" in df_all.columns:
                            df_all.loc[ridx, "播種/移植日（代表日・調整）"] = mid_date(new_from, new_to)
                    prev = pd.DataFrame(preview_rows)
                    if not prev.empty:
                        prev["From(現状)"] = pd.to_datetime(prev["From(現状)"], errors="coerce").dt.date
                        prev["To(現状)"] = pd.to_datetime(prev["To(現状)"], errors="coerce").dt.date
                        prev["From(変更後)"] = pd.to_datetime(prev["From(変更後)"], errors="coerce").dt.date
                        prev["To(変更後)"] = pd.to_datetime(prev["To(変更後)"], errors="coerce").dt.date
                    return df_all, prev

                def _peak_summary(df_sak_any: pd.DataFrame) -> Optional[Dict]:
                    _, tasks_any, _ = compute_all(df_sak_any, df_mach, df_tpl, df_exc, defaults)
                    tasks_any = tasks_any[tasks_any["source"].isin(ALL_SOURCES)].copy()
                    load_any = build_load_table(tasks_any, farm=farm_sel, group_by=gran, utilization=utilization, detail=False)
                    if load_any is None or load_any.empty:
                        return None
                    load_any = load_any.copy()
                    load_any["必要稼働率"] = np.where(
                        pd.to_numeric(load_any.get("期間日数"), errors="coerce") > 0,
                        pd.to_numeric(load_any.get("推定機械日数_有効"), errors="coerce") / pd.to_numeric(load_any.get("期間日数"), errors="coerce"),
                        np.nan,
                    )
                    row = load_any.loc[load_any["必要稼働率"].idxmax()]
                    if gran == "月":
                        when = f'{int(row["年"])}年 {int(row["月"])}月'
                    else:
                        when = f'{int(row["年"])}年 {jn_to_label(int(row["旬番号"]))}'
                    md = float(pd.to_numeric(row.get("推定機械日数_有効"), errors="coerce"))
                    cap = float(pd.to_numeric(row.get("容量(機械日)"), errors="coerce"))
                    return {
                        "when": when,
                        "cat": str(row.get("農機カテゴリ", "")).strip(),
                        "rate": float(pd.to_numeric(row.get("必要稼働率"), errors="coerce")),
                        "need": md,
                        "cap": cap,
                        "over": md - cap,
                    }

                # ③ preview
                st.markdown("**③ 効果プレビュー**")
                btn_preview = st.button("プレビュー更新", key=f"manual_shift_preview_{farm_sel}", type="secondary")
                if btn_preview:
                    df_after, prev_tbl = _build_shifted_df(new_from, new_to)
                    st.session_state[f"manual_shift_prev_tbl_{farm_sel}"] = prev_tbl
                    st.session_state[f"manual_shift_peak_after_{farm_sel}"] = _peak_summary(df_after)

                prev_tbl = st.session_state.get(f"manual_shift_prev_tbl_{farm_sel}", pd.DataFrame())
                peak_after = st.session_state.get(f"manual_shift_peak_after_{farm_sel}", None)

                peak_upload = None
                if compare_mode:
                    try:
                        load_base_p = build_load_table(tasks_base, farm=farm_sel, group_by=gran, utilization=utilization, detail=False)
                        if load_base_p is not None and not load_base_p.empty:
                            load_base_p = load_base_p.copy()
                            load_base_p["必要稼働率"] = np.where(
                                pd.to_numeric(load_base_p.get("期間日数"), errors="coerce") > 0,
                                pd.to_numeric(load_base_p.get("推定機械日数_有効"), errors="coerce") / pd.to_numeric(load_base_p.get("期間日数"), errors="coerce"),
                                np.nan,
                            )
                            peak_upload = float(load_base_p["必要稼働率"].max())
                    except Exception:
                        peak_upload = None

                colP1, colP2, colP3, colP4 = st.columns([1, 1, 1, 2])
                with colP1:
                    if peak_upload is not None:
                        st.metric("ピーク必要稼働率（アップロード直後）", f"{peak_upload:.0%}")
                    else:
                        st.metric("ピーク必要稼働率（アップロード直後）", "-")
                with colP2:
                    st.metric("ピーク必要稼働率（編集前）", f"{peak:.0%}")
                with colP3:
                    if isinstance(peak_after, dict) and peak_after.get("rate") is not None:
                        st.metric("変更後ピーク必要稼働率", f"{float(peak_after['rate']):.0%}")
                    else:
                        st.metric("変更後ピーク必要稼働率", "-")
                with colP4:
                    if isinstance(peak_after, dict):
                        st.caption(f"変更後ピーク: {peak_after.get('when')}（{peak_after.get('cat')}）｜超過 {float(peak_after.get('over', 0.0)):.2f}機械日")
                    else:
                        st.caption("プレビュー更新を押すと、変更後ピークが表示されます。")

                if isinstance(prev_tbl, pd.DataFrame) and not prev_tbl.empty:
                    prev_tbl_disp = prev_tbl.copy()
                    # show "after" columns first to make it obvious what to look at
                    preferred = [c for c in ["入力行Index", "作物", "面積(ha)", "From(変更後)", "To(変更後)", "From(現状)", "To(現状)"] if c in prev_tbl_disp.columns]
                    rest = [c for c in prev_tbl_disp.columns if c not in preferred]
                    prev_tbl_disp = prev_tbl_disp[preferred + rest]
                    st.dataframe(
                        prev_tbl_disp,
                        use_container_width=True,
                        height=240,
                        column_config=_date_col_config(["From(現状)", "To(現状)", "From(変更後)", "To(変更後)"]),
                        hide_index=True,
                    )

                # ④ apply
                st.markdown("**④ 適用**")
                if st.button("このシフトを適用して再計算", key=f"manual_shift_apply_{farm_sel}", type="primary"):
                    df_after, prev_tbl2 = _build_shifted_df(new_from, new_to)
                    st.session_state.df_sak_edit = df_after
                    st.rerun()

# ----------------------------
# 3) Main chart
# ----------------------------
st.subheader("労働負荷グラフ")
st.caption("棒：必要稼働率（=推定機械日数÷期間日数）。赤線：農機稼働率％。赤線オーバーの旬/月がボトルネック候補です。")
if not farm_sel:
    st.stop()

df = tasks_edit[tasks_edit["農家名"] == farm_sel].copy()
if df.empty:
    st.warning("フィルター条件でデータがありません。")
    st.stop()

def _top_crops_text(d: pd.DataFrame, n: int = 6) -> str:
    if d is None or d.empty or "作物" not in d.columns:
        return ""
    g = d.groupby("作物", as_index=False)["推定機械日数_有効"].sum().sort_values("推定機械日数_有効", ascending=False)
    crops = g["作物"].astype(str).tolist()
    crops = [c.strip() for c in crops if c and c.strip() and c.strip().lower() not in {"nan", "none"}]
    if not crops:
        return ""
    if len(crops) <= n:
        return " / ".join(crops)
    return " / ".join(crops[:n]) + f" (+{len(crops) - n})"
def _work_label_col(df_in: pd.DataFrame) -> pd.Series:
    g = df_in.get("作業グループ", "").fillna("").astype(str).str.strip()
    m = df_in.get("メモ", "").fillna("").astype(str).str.strip()
    m = m.where(~m.str.lower().isin({"nan", "none"}), "")
    return np.where(m != "", g + "(" + m + ")", g)

def _top_cats_text(d: pd.DataFrame, n: int = 6) -> str:
    if d is None or d.empty or "農機カテゴリ" not in d.columns:
        return ""
    g = d.groupby("農機カテゴリ", as_index=False)["推定機械日数_有効"].sum().sort_values("推定機械日数_有効", ascending=False)
    cats = g["農機カテゴリ"].astype(str).tolist()
    cats = [c.strip() for c in cats if c and c.strip() and c.strip().lower() not in {"nan", "none"}]
    if not cats:
        return ""
    if len(cats) <= n:
        return " / ".join(cats)
    return " / ".join(cats[:n]) + f" (+{len(cats) - n})"

def _required_rate(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype=float)
    d = pd.to_numeric(df.get("期間日数"), errors="coerce")
    m = pd.to_numeric(df.get("推定機械日数_有効"), errors="coerce")
    return np.where(d > 0, m / d, np.nan)

color_col = "作業表示" if breakdown == "作業別" else "作物"

cats_in_scope = sorted(
    [
        c
        for c in pd.concat(
            [
                tasks_edit.loc[tasks_edit["農家名"] == farm_sel, "農機カテゴリ"],
                tasks_base.loc[tasks_base["農家名"] == farm_sel, "農機カテゴリ"] if compare_mode else pd.Series([], dtype=object),
            ],
            ignore_index=True,
        )
        .dropna()
        .astype(str)
        .unique()
        .tolist()
        if str(c).strip()
    ]
)
default_cat = None
if "peak_cat" in locals() and isinstance(peak_cat, str) and peak_cat in cats_in_scope:
    default_cat = f"ボトルネック（{peak_cat}）"
cat_options = ["全て"]
if default_cat:
    cat_options.append(default_cat)
cat_options.extend([c for c in cats_in_scope if default_cat is None or c != peak_cat])
cat_sel_key = "workload_machine_cat_sel"
cat_sel = st.selectbox("労働負荷グラフの農機カテゴリ", cat_options, index=0, key=cat_sel_key)

cap_line = None
cap_cat = None
cap_multiplier = 1
cap_ha_day = None
if cat_sel.startswith("ボトルネック（") and "peak_cat" in locals():
    cap_cat = peak_cat
elif cat_sel != "全て":
    cap_cat = cat_sel

st.caption("y軸は『必要稼働率（=推定機械日数÷期間日数）』です。赤線は設定した『農機稼働率％』で、これを超えると回し切れない目安です。")
if cat_sel == "全て":
    st.info("注意：『全て』は各旬/月で“ボトルネック農機カテゴリ”だけを表示します。作業の全量や分散の詳細は、農機カテゴリ（sprayer/tractor…）を選んで確認してください。")

if cap_cat:
    cap_s = cap_long[(cap_long["農家名"].astype(str).str.strip() == farm_sel) & (cap_long["農機カテゴリ"] == cap_cat)]["能力(ha/日)"]
    if not cap_s.empty and not pd.isna(cap_s.iloc[0]):
        cap_ha_day = float(cap_s.iloc[0])

if gran == "旬":
    group_by = "旬"
else:
    group_by = "月"

scenario_tasks = []
for scen_name, tasks_src in [("編集後", tasks_edit), ("現状", tasks_base)]:
    if scen_name == "現状" and not compare_mode:
        continue
    t = tasks_src[tasks_src["農家名"] == farm_sel].copy()
    if t.empty:
        continue
    t["シナリオ"] = scen_name
    t["作業表示"] = _work_label_col(t)
    scenario_tasks.append(t)

def _period_days_table(keys_df: pd.DataFrame) -> pd.DataFrame:
    out = keys_df.drop_duplicates().copy()
    if group_by == "旬":
        out["期間日数"] = out.apply(lambda r: jun_days(int(r["年"]), int(r["旬番号"])), axis=1)
        out["ラベル"] = out["旬ラベル"].astype(str)
    else:
        out["期間日数"] = out.apply(lambda r: calendar.monthrange(int(r["年"]), int(r["月"]))[1], axis=1)
        out["ラベル"] = out["月"].astype(int).astype(str) + "月"
    return out

def _bottleneck_by_period(t: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
    g = t.groupby(keys + ["農機カテゴリ"], as_index=False).agg({"推定機械日数_有効": "sum"})
    pdays = _period_days_table(g[keys])
    g = g.merge(pdays[keys + ["期間日数", "ラベル"]], on=keys, how="left")
    g["必要稼働率"] = np.where(pd.to_numeric(g["期間日数"], errors="coerce") > 0, g["推定機械日数_有効"] / g["期間日数"], np.nan)
    idx = g.groupby(keys)["必要稼働率"].idxmax()
    bn = g.loc[idx, keys + ["農機カテゴリ", "必要稼働率", "期間日数", "ラベル"]].copy()
    bn = bn.rename(columns={"農機カテゴリ": "ボトルネック農機カテゴリ"})
    return bn

if scenario_tasks:
    t_all = pd.concat(scenario_tasks, ignore_index=True)
    if group_by == "旬":
        period_keys = ["年", "旬番号", "旬ラベル"]
        sort_cols = ["年", "旬番号"]
    else:
        period_keys = ["年", "月"]
        sort_cols = ["年", "月"]

    plot_rows = []
    for scen_name, t in t_all.groupby("シナリオ"):
        t = t.copy()
        if breakdown == "作業別":
            for c in ["作業グループ", "メモ"]:
                if c not in t.columns:
                    t[c] = ""
        if cap_cat:
            # single machine category
            t = t[t["農機カテゴリ"] == cap_cat].copy()
            if t.empty:
                continue
            grp_cols = period_keys + ["農機カテゴリ", color_col]
            if breakdown == "作業別":
                grp_cols += ["作業グループ", "メモ"]
            g = t.groupby(grp_cols, as_index=False).agg(
                {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
            )
        else:
            # 全て: 各旬/月ごとの「ボトルネック農機カテゴリ」だけに絞って表示（= 稼働限界に対する利用率が意味を持つ）
            if t.empty:
                continue
            bn = _bottleneck_by_period(t, period_keys)
            t = t.merge(bn[period_keys + ["ボトルネック農機カテゴリ"]], on=period_keys, how="inner")
            t = t[t["農機カテゴリ"] == t["ボトルネック農機カテゴリ"]].copy()
            if t.empty:
                continue
            grp_cols = period_keys + ["農機カテゴリ", color_col]
            if breakdown == "作業別":
                grp_cols += ["作業グループ", "メモ"]
            g = t.groupby(grp_cols, as_index=False).agg(
                {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
            )

        # Tooltip: show crop names even in 作業別 (work breakdown)
        if breakdown == "作業別" and "作物" in t.columns:
            crop_cols = [c for c in grp_cols if c in t.columns]
            top = (
                t.groupby(crop_cols + ["作物"], as_index=False)["推定機械日数_有効"]
                .sum()
                .sort_values("推定機械日数_有効", ascending=False)
            )

            def _top_crops_per_group(d: pd.DataFrame) -> str:
                crops = d["作物"].astype(str).tolist()
                crops = [c.strip() for c in crops if c and c.strip() and c.strip().lower() not in {"nan", "none"}]
                crops = crops[:3]
                return " / ".join(crops)

            top_crops = top.groupby(crop_cols, as_index=False).apply(_top_crops_per_group)
            top_crops = top_crops.rename(columns={0: "作物(上位3)"}).reset_index(drop=True)
            g = g.merge(top_crops, on=crop_cols, how="left")

        pdays = _period_days_table(g[period_keys])
        g = g.merge(pdays[period_keys + ["期間日数", "ラベル"]], on=period_keys, how="left")
        g["必要稼働率"] = np.where(
            pd.to_numeric(g["期間日数"], errors="coerce") > 0,
            g["推定機械日数_有効"] / g["期間日数"],
            np.nan,
        )
        g["シナリオ"] = scen_name
        plot_rows.append(g)

    if plot_rows:
        plot_df = pd.concat(plot_rows, ignore_index=True).sort_values(["シナリオ"] + sort_cols)
        cap_map = cap_long.set_index(["農家名", "農機カテゴリ"])["能力(ha/日)"].to_dict() if not cap_long.empty else {}
        plot_df["農機能力(ha/日)"] = plot_df.apply(lambda r: cap_map.get((farm_sel, str(r.get("農機カテゴリ", "")).strip())), axis=1)
        ymax = float(np.nanmax(pd.to_numeric(plot_df["必要稼働率"], errors="coerce").to_numpy())) if not plot_df["必要稼働率"].dropna().empty else 0.0
        title_cat = cap_cat if cap_cat else "全て（ボトルネック農機）"
        hover_data = {
            "農機カテゴリ": True,
            "農機能力(ha/日)": ":.2f",
            "推定機械日数_有効": ":.2f",
            "面積(ha)": ":.2f",
            "期間日数": ":.0f",
        }
        # Crop name in tooltip (作物別は作物、作業別は作物(上位3))
        if "作物" in plot_df.columns:
            hover_data["作物"] = True
        if "作物(上位3)" in plot_df.columns:
            hover_data["作物(上位3)"] = True
        # Hide noisy fields from tooltip
        if "ラベル" in plot_df.columns:
            hover_data["ラベル"] = False
        if "シナリオ" in plot_df.columns:
            hover_data["シナリオ"] = False

        fig = px.bar(
            plot_df,
            x="ラベル",
            y="必要稼働率",
            color=color_col,
            barmode="stack",
            facet_col="シナリオ" if compare_mode else None,
            hover_data=hover_data,
            title=f"{farm_sel}｜{gran}別｜必要稼働率（{title_cat}｜{breakdown}）",
        )
        fig.update_yaxes(tickformat=".0%", rangemode="tozero", range=[0, max(ymax * 1.1, float(utilization) * 1.2, 0.1)])
        fig.update_layout(
            template="plotly_white",
            height=620,
            bargap=0.25,
            title=dict(x=0, xanchor="left"),
            legend=dict(orientation="h", yanchor="top", y=1.20, xanchor="left", x=0, title_text=""),
            margin=dict(t=190, r=20, b=110, l=60),
        )
        # Red line: keep it simple (annotations tend to overlap when comparing scenarios)
        fig.add_hline(y=float(utilization), line_dash="dash", line_color="red")
        if group_by == "旬":
            cat_arr = [jn_to_label(i) for i in range(1, 37)]
            for ax_name in [k for k in fig.layout if str(k).startswith("xaxis")]:
                fig.layout[ax_name].update({"categoryorder": "array", "categoryarray": cat_arr})
        fig.update_xaxes(tickangle=-45, automargin=True, title_text="")
        st.plotly_chart(fig, use_container_width=True)

        # Quick understanding aid for "全て（ボトルネック）"
        if cat_sel == "全て" and cap_cat is None:
            bn_tbl = (
                plot_df.groupby(period_keys + ["ラベル", "農機カテゴリ", "シナリオ"], as_index=False)
                .agg({"推定機械日数_有効": "sum", "面積(ha)": "sum", "期間日数": "max"})
                .sort_values(["シナリオ"] + sort_cols)
            )
            bn_tbl["必要稼働率(%)"] = np.where(
                pd.to_numeric(bn_tbl["期間日数"], errors="coerce") > 0,
                pd.to_numeric(bn_tbl["推定機械日数_有効"], errors="coerce") / pd.to_numeric(bn_tbl["期間日数"], errors="coerce") * 100.0,
                np.nan,
            )
            bn_tbl["農機稼働率(%)"] = float(utilization) * 100.0
            bn_tbl["超過(%)"] = bn_tbl["必要稼働率(%)"] - bn_tbl["農機稼働率(%)"]
            bn_tbl = bn_tbl.rename(columns={"農機カテゴリ": "ボトルネック農機カテゴリ"})
            show_cols = []
            for c in ["シナリオ", "年"]:
                if c in bn_tbl.columns:
                    show_cols.append(c)
            if group_by == "旬":
                for c in ["旬ラベル"]:
                    if c in bn_tbl.columns:
                        show_cols.append(c)
            else:
                for c in ["月"]:
                    if c in bn_tbl.columns:
                        show_cols.append(c)
            for c in ["ボトルネック農機カテゴリ", "面積(ha)", "推定機械日数_有効", "期間日数", "必要稼働率(%)", "農機稼働率(%)", "超過(%)"]:
                if c in bn_tbl.columns:
                    show_cols.append(c)
            with st.expander("各旬/月のボトルネック農機（一覧）", expanded=False):
                st.caption("『全て』表示は、各旬/月で最も詰まっている（=必要稼働率が最大の）農機カテゴリだけを抜き出しています。")
                st.dataframe(
                    bn_tbl[show_cols],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "面積(ha)": st.column_config.NumberColumn(format="%.2f"),
                        "推定機械日数_有効": st.column_config.NumberColumn(format="%.2f"),
                        "期間日数": st.column_config.NumberColumn(format="%.0f"),
                        "必要稼働率(%)": st.column_config.NumberColumn(format="%.0f"),
                        "農機稼働率(%)": st.column_config.NumberColumn(format="%.0f"),
                        "超過(%)": st.column_config.NumberColumn(format="%.0f"),
                    },
                )

        with st.expander("グラフ値（テーブル表示）", expanded=False):
            st.caption("下の表は、上の労働負荷グラフに実際に描画されている集計値です。")
            st.markdown("**合計（旬/月×農機×シナリオ）**")
            totals_tbl = (
                plot_df.groupby(period_keys + ["ラベル", "農機カテゴリ", "シナリオ"], as_index=False)
                .agg({"推定機械日数_有効": "sum", "面積(ha)": "sum", "期間日数": "max"})
            )
            totals_tbl["必要稼働率(%)"] = np.where(
                pd.to_numeric(totals_tbl["期間日数"], errors="coerce") > 0,
                pd.to_numeric(totals_tbl["推定機械日数_有効"], errors="coerce") / pd.to_numeric(totals_tbl["期間日数"], errors="coerce") * 100.0,
                np.nan,
            )
            total_cols = [c for c in [
                "シナリオ",
                "年",
                "ラベル",
                "農機カテゴリ",
                "面積(ha)",
                "推定機械日数_有効",
                "期間日数",
                "必要稼働率(%)",
            ] if c in totals_tbl.columns]
            st.dataframe(
                totals_tbl.sort_values(["シナリオ"] + sort_cols + (["農機カテゴリ"] if "農機カテゴリ" in totals_tbl.columns else []))[total_cols],
                use_container_width=True,
                height=240,
                hide_index=True,
                column_config={
                    "必要稼働率(%)": st.column_config.NumberColumn(format="%.0f"),
                    "推定機械日数_有効": st.column_config.NumberColumn(format="%.2f"),
                    "面積(ha)": st.column_config.NumberColumn(format="%.2f"),
                    "期間日数": st.column_config.NumberColumn(format="%.0f"),
                },
            )

            st.markdown("**内訳（積み上げの各色）**")
            plot_disp = plot_df.copy()
            plot_disp["必要稼働率(%)"] = pd.to_numeric(plot_disp["必要稼働率"], errors="coerce") * 100.0
            show_cols = [c for c in [
                "シナリオ",
                "年",
                "ラベル",
                "農機カテゴリ",
                "作業グループ",
                "メモ",
                color_col,
                "面積(ha)",
                "推定機械日数_有効",
                "期間日数",
                "必要稼働率(%)",
            ] if c in plot_disp.columns]
            st.dataframe(
                plot_disp.sort_values(["シナリオ"] + sort_cols + (["農機カテゴリ"] if "農機カテゴリ" in plot_disp.columns else []) + [color_col])[show_cols],
                use_container_width=True,
                height=320,
                hide_index=True,
                column_config={
                    "必要稼働率(%)": st.column_config.NumberColumn(format="%.0f"),
                    "推定機械日数_有効": st.column_config.NumberColumn(format="%.2f"),
                    "面積(ha)": st.column_config.NumberColumn(format="%.2f"),
                    "期間日数": st.column_config.NumberColumn(format="%.0f"),
                },
            )

        if cat_sel == "全て":
            st.caption("※『全て』表示は、各旬/月で最も詰まっている農機（ボトルネック農機カテゴリ）だけに絞って内訳を表示します（稼働限界に対する利用率が意味を持つため）。")

            with st.expander("参考：農機カテゴリ別の必要稼働率（全て）", expanded=False):
                rows_cat = []
                for scen_name, tasks_src in [("編集後", tasks_edit), ("現状", tasks_base)]:
                    if scen_name == "現状" and not compare_mode:
                        continue
                    load_s = build_load_table(tasks_src, farm=farm_sel, group_by=gran, utilization=utilization, detail=False)
                    if load_s is None or load_s.empty:
                        continue
                    load_s = load_s.copy()
                    load_s["必要稼働率"] = np.where(
                        pd.to_numeric(load_s.get("期間日数"), errors="coerce") > 0,
                        pd.to_numeric(load_s.get("推定機械日数_有効"), errors="coerce") / pd.to_numeric(load_s.get("期間日数"), errors="coerce"),
                        np.nan,
                    )
                    if gran == "旬":
                        load_s["ラベル"] = load_s["旬ラベル"].astype(str)
                    else:
                        load_s["ラベル"] = load_s["月"].astype(int).astype(str) + "月"
                    load_s["シナリオ"] = scen_name
                    rows_cat.append(load_s[["年", "農機カテゴリ", "ラベル", "必要稼働率", "推定機械日数_有効", "期間日数", "シナリオ"]])

                if rows_cat:
                    cat_df = pd.concat(rows_cat, ignore_index=True)
                    top_cats = (
                        cat_df[cat_df["シナリオ"] == "編集後"]
                        .groupby("農機カテゴリ", as_index=False)["必要稼働率"]
                        .max()
                        .sort_values("必要稼働率", ascending=False)
                        .head(6)["農機カテゴリ"]
                        .tolist()
                    )
                    sel_cats = st.multiselect(
                        "表示する農機カテゴリ（上位を初期選択）",
                        options=sorted(cat_df["農機カテゴリ"].dropna().astype(str).unique().tolist()),
                        default=top_cats,
                    )
                    if sel_cats:
                        cat_df = cat_df[cat_df["農機カテゴリ"].isin(sel_cats)].copy()
                    ymax2 = float(np.nanmax(pd.to_numeric(cat_df["必要稼働率"], errors="coerce").to_numpy())) if not cat_df["必要稼働率"].dropna().empty else 0.0
                    fig_cat = px.bar(
                        cat_df.sort_values(["シナリオ", "年", "ラベル"]),
                        x="ラベル",
                        y="必要稼働率",
                        color="農機カテゴリ",
                        barmode="group",
                        facet_col="シナリオ" if compare_mode else None,
                        hover_data={"推定機械日数_有効": ":.2f", "期間日数": ":.0f"},
                        title=f"{farm_sel}｜{gran}別｜農機カテゴリ別 必要稼働率",
                    )
                    fig_cat.update_yaxes(tickformat=".0%", rangemode="tozero", range=[0, max(ymax2 * 1.1, float(utilization) * 1.2, 1.0)])
                    fig_cat.add_hline(y=float(utilization), line_dash="dash", line_color="red")
                    if gran == "旬":
                        cat_arr = [jn_to_label(i) for i in range(1, 37)]
                        for ax_name in [k for k in fig_cat.layout if str(k).startswith("xaxis")]:
                            fig_cat.layout[ax_name].update({"categoryorder": "array", "categoryarray": cat_arr})
                    st.plotly_chart(fig_cat, use_container_width=True)

                    cat_tbl = (
                        cat_df.groupby(["シナリオ", "農機カテゴリ"], as_index=False)["必要稼働率"]
                        .max()
                        .sort_values(["シナリオ", "必要稼働率"], ascending=[True, False])
                    )
                    cat_tbl["必要稼働率(%)"] = pd.to_numeric(cat_tbl["必要稼働率"], errors="coerce") * 100.0
                    st.dataframe(
                        cat_tbl[["シナリオ", "農機カテゴリ", "必要稼働率(%)"]],
                        use_container_width=True,
                        height=220,
                        column_config={"必要稼働率(%)": st.column_config.NumberColumn(format="%.0f")},
                        hide_index=True,
                    )
                else:
                    st.caption("農機カテゴリ別の集計データがありません。")

        # ----------------------------
        # Overflow summary (actionable list)
        # ----------------------------
        st.markdown("**超過期間サマリ（必要稼働率 > 農機稼働率％）**")
        st.caption("どの旬/月で、どの農機が詰まっていて、内訳（作業/作物）の何が効いているかを上位から並べます。")

        totals = (
            plot_df.groupby(period_keys + ["シナリオ", "農機カテゴリ", "ラベル"], as_index=False)
            .agg({"推定機械日数_有効": "sum", "期間日数": "max"})
        )
        totals["必要稼働率"] = np.where(
            pd.to_numeric(totals["期間日数"], errors="coerce") > 0,
            pd.to_numeric(totals["推定機械日数_有効"], errors="coerce") / pd.to_numeric(totals["期間日数"], errors="coerce"),
            np.nan,
        )
        totals["超過(必要稼働率-稼働率)"] = totals["必要稼働率"] - float(utilization)

        # top contributors text for each period/scenario
        contrib = plot_df.copy()
        contrib["寄与(必要稼働率)"] = pd.to_numeric(contrib["必要稼働率"], errors="coerce")
        def _top3_text(d: pd.DataFrame) -> str:
            if d is None or d.empty:
                return ""
            return " / ".join(
                d.sort_values("寄与(必要稼働率)", ascending=False)[color_col].astype(str).head(3).tolist()
            )

        top_contrib = (
            contrib.groupby(period_keys + ["シナリオ", "農機カテゴリ", "ラベル"])
            .apply(_top3_text)
            .rename(f"{color_col}(上位3)")
            .reset_index()
        )
        totals = totals.merge(top_contrib, on=period_keys + ["シナリオ", "農機カテゴリ", "ラベル"], how="left")

        show = totals.copy()
        if compare_mode:
            show = show.pivot_table(
                index=period_keys + ["ラベル", "農機カテゴリ", f"{color_col}(上位3)"],
                columns="シナリオ",
                values=["必要稼働率", "超過(必要稼働率-稼働率)"],
                aggfunc="max",
            )
            show.columns = [f"{a}（{b}）" for a, b in show.columns]
            show = show.reset_index()
            # sort by edited overflow desc
            sort_key = "超過(必要稼働率-稼働率)（編集後）" if "超過(必要稼働率-稼働率)（編集後）" in show.columns else None
            if sort_key:
                show = show.sort_values(sort_key, ascending=False)
        else:
            show = show.sort_values(["超過(必要稼働率-稼働率)"], ascending=False)

        # only periods where edited is overflow (or single scenario overflow)
        if compare_mode and "超過(必要稼働率-稼働率)（編集後）" in show.columns:
            show = show[show["超過(必要稼働率-稼働率)（編集後）"] > 0].copy()
        elif not compare_mode:
            show = show[show["超過(必要稼働率-稼働率)"] > 0].copy()

        if show.empty:
            st.success("超過は見つかりませんでした（必要稼働率が農機稼働率％以下）。")
        else:
            # show as percent numbers (0-100) for readability
            show_disp = show.copy()
            pct_cols = [c for c in show_disp.columns if ("必要稼働率" in c) or ("超過(必要稼働率-稼働率)" in c)]
            for c in pct_cols:
                show_disp[c] = pd.to_numeric(show_disp[c], errors="coerce") * 100.0

            col_list, col_hint = st.columns([2, 1])
            with col_hint:
                st.caption("Tip: まずは上位の旬/月の『上位3内訳』に対応する作付のFrom/Toを動かすのが近道です。")
            with col_list:
                st.dataframe(
                    show_disp.head(15),
                    use_container_width=True,
                    height=300,
                    column_config={
                        "必要稼働率": st.column_config.NumberColumn(format="%.0f"),
                        "必要稼働率（編集後）": st.column_config.NumberColumn(format="%.0f"),
                        "必要稼働率（現状）": st.column_config.NumberColumn(format="%.0f"),
                        "超過(必要稼働率-稼働率)": st.column_config.NumberColumn(format="%.0f"),
                        "超過(必要稼働率-稼働率)（編集後）": st.column_config.NumberColumn(format="%.0f"),
                        "超過(必要稼働率-稼働率)（現状）": st.column_config.NumberColumn(format="%.0f"),
                    },
                    hide_index=True,
                )

                with st.expander("ボトルネック農機カテゴリ（旬/月ごとの一覧）", expanded=False):
                    bn_disp = totals.copy()
                    bn_disp["必要稼働率(%)"] = pd.to_numeric(bn_disp["必要稼働率"], errors="coerce") * 100.0
                    bn_disp["超過(%)"] = pd.to_numeric(bn_disp["超過(必要稼働率-稼働率)"], errors="coerce") * 100.0
                    bn_cols = [c for c in [
                        "ラベル", "シナリオ", "農機カテゴリ", "必要稼働率(%)", "超過(%)", f"{color_col}(上位3)",
                    ] if c in bn_disp.columns]
                    sort_by = ["シナリオ"] + [c for c in sort_cols if c in bn_disp.columns]
                    if not sort_by:
                        sort_by = [c for c in ["シナリオ", "ラベル"] if c in bn_disp.columns]
                    st.dataframe(
                        bn_disp.sort_values(sort_by)[bn_cols],
                        use_container_width=True,
                        height=260,
                        hide_index=True,
                    )
    else:
        st.info("データがありません。入力をご確認ください。")
else:
    st.info("データがありません。入力をご確認ください。")

if cat_sel == "全て" and farm_sel:
    with st.expander("全て表示：ピーク必要稼働率（最大）の根拠（どの農機が詰まっているか）", expanded=False):
        rows = []
        for scen_name, tasks_src in [("編集後", tasks_edit), ("現状", tasks_base)]:
            if scen_name == "現状" and not compare_mode:
                continue
            load_s = build_load_table(tasks_src, farm=farm_sel, group_by=gran, utilization=utilization, detail=False)
            if load_s is None or load_s.empty:
                continue
            if gran == "旬":
                grp_cols = ["年", "旬番号", "旬ラベル"]
            else:
                grp_cols = ["年", "月"]
            load_s = load_s.copy()
            load_s["必要稼働率"] = _required_rate(load_s)
            idx = load_s.groupby(grp_cols)["必要稼働率"].idxmax()
            mx = load_s.loc[idx].copy()
            mx["シナリオ"] = scen_name
            rows.append(mx)

        if not rows:
            st.caption("対象データがありません。")
        else:
            mx_df = pd.concat(rows, ignore_index=True)
            show_cols = [c for c in [
                "年", "旬ラベル" if gran == "旬" else "月", "農機カテゴリ",
                "必要稼働率", "推定機械日数_有効", "期間日数", "容量(機械日)", "面積(ha)", "シナリオ",
            ] if c in mx_df.columns]
            mx_show = mx_df.copy()
            if gran == "月" and "月" in mx_show.columns:
                mx_show["月"] = mx_show["月"].astype(int).astype(str) + "月"
            mx_show = mx_show.sort_values(["シナリオ", "年", "旬ラベル" if gran == "旬" else "月"])
            if "必要稼働率" in mx_show.columns:
                mx_show["必要稼働率"] = pd.to_numeric(mx_show["必要稼働率"], errors="coerce")
            st.dataframe(mx_show[show_cols], use_container_width=True, height=260, hide_index=True)

# ----------------------------
# 6) Downloads
# ----------------------------
st.subheader("ダウンロード")
with st.expander("レポート（HTML → 印刷でPDF保存）", expanded=False):
    st.caption("HTMLをダウンロード → ブラウザで開く → 印刷（PDF保存）でレポート化できます。グラフは旬別のみを掲載します。")
    if not farm_sel:
        st.info("農家を選択してください。")
    else:
        import plotly.io as pio

        def _df_to_html_table(df_in: pd.DataFrame) -> str:
            if df_in is None or df_in.empty:
                return "<p class='muted'>（データなし）</p>"
            return df_in.to_html(index=False, escape=True, classes="table", border=0)

        def _fig_to_html(fig_in) -> str:
            return pio.to_html(
                fig_in,
                include_plotlyjs="inline",
                full_html=False,
                config={"displayModeBar": False, "responsive": True},
            )

        # ---- Report data (旬固定)
        report_breakdown = "作業別"
        report_gran = "旬"
        report_util = float(utilization)

        load_edit_r = build_load_table(tasks_edit, farm=farm_sel, group_by=report_gran, utilization=report_util, detail=False)
        load_base_r = build_load_table(tasks_base, farm=farm_sel, group_by=report_gran, utilization=report_util, detail=False) if compare_mode else pd.DataFrame()
        for _d in [load_edit_r, load_base_r]:
            if _d is None or _d.empty:
                continue
            _d["必要稼働率"] = np.where(
                pd.to_numeric(_d.get("期間日数"), errors="coerce") > 0,
                pd.to_numeric(_d.get("推定機械日数_有効"), errors="coerce") / pd.to_numeric(_d.get("期間日数"), errors="coerce"),
                np.nan,
            )

        peak_edit = float(load_edit_r["必要稼働率"].max()) if load_edit_r is not None and not load_edit_r.empty else np.nan
        peak_base = float(load_base_r["必要稼働率"].max()) if load_base_r is not None and not load_base_r.empty else np.nan
        peak_row_edit = load_edit_r.loc[load_edit_r["必要稼働率"].idxmax()] if load_edit_r is not None and not load_edit_r.empty else None
        peak_when = ""
        peak_cat_r = ""
        if peak_row_edit is not None:
            peak_when = f'{int(peak_row_edit["年"])}年 {jn_to_label(int(peak_row_edit["旬番号"]))}'
            peak_cat_r = str(peak_row_edit.get("農機カテゴリ", "")).strip()

        # ---- Workload chart (旬 / 全て=ボトルネック / 作業別)
        def _work_label_col(df_in: pd.DataFrame) -> pd.Series:
            g = df_in.get("作業グループ", "").fillna("").astype(str).str.strip()
            m = df_in.get("メモ", "").fillna("").astype(str).str.strip()
            m = m.where(~m.str.lower().isin({"nan", "none"}), "")
            return np.where(m != "", g + "(" + m + ")", g)

        def _period_days_jun(keys_df: pd.DataFrame) -> pd.DataFrame:
            out = keys_df.drop_duplicates().copy()
            out["期間日数"] = out.apply(lambda r: jun_days(int(r["年"]), int(r["旬番号"])), axis=1)
            out["ラベル"] = out["旬ラベル"].astype(str)
            return out

        def _bottleneck_by_period(t: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
            g = t.groupby(keys + ["農機カテゴリ"], as_index=False).agg({"推定機械日数_有効": "sum"})
            pdays = _period_days_jun(g[keys])
            g = g.merge(pdays[keys + ["期間日数", "ラベル"]], on=keys, how="left")
            g["必要稼働率"] = np.where(pd.to_numeric(g["期間日数"], errors="coerce") > 0, g["推定機械日数_有効"] / g["期間日数"], np.nan)
            idx = g.groupby(keys)["必要稼働率"].idxmax()
            bn = g.loc[idx, keys + ["農機カテゴリ", "必要稼働率", "期間日数", "ラベル"]].copy()
            bn = bn.rename(columns={"農機カテゴリ": "ボトルネック農機カテゴリ"})
            return bn

        period_keys = ["年", "旬番号", "旬ラベル"]
        sort_cols = ["年", "旬番号"]
        scenario_tasks_r = []
        for scen_name, tasks_src in [("編集後", tasks_edit), ("現状", tasks_base)]:
            if scen_name == "現状" and not compare_mode:
                continue
            t = tasks_src[tasks_src["農家名"] == farm_sel].copy()
            if t.empty:
                continue
            for c in ["作業グループ", "メモ"]:
                if c not in t.columns:
                    t[c] = ""
            t["作業表示"] = _work_label_col(t)
            t["シナリオ"] = scen_name
            scenario_tasks_r.append(t)

        plot_df_r = pd.DataFrame()
        fig_html = "<p class='muted'>（グラフなし）</p>"
        bn_html = "<p class='muted'>（データなし）</p>"
        if scenario_tasks_r:
            t_all = pd.concat(scenario_tasks_r, ignore_index=True)
            plot_rows = []
            for scen_name, t in t_all.groupby("シナリオ"):
                bn = _bottleneck_by_period(t, period_keys)
                t2 = t.merge(bn[period_keys + ["ボトルネック農機カテゴリ"]], on=period_keys, how="inner")
                t2 = t2[t2["農機カテゴリ"] == t2["ボトルネック農機カテゴリ"]].copy()
                if t2.empty:
                    continue
                g = t2.groupby(period_keys + ["農機カテゴリ", "作業表示"], as_index=False).agg(
                    {"推定機械日数_有効": "sum", "面積(ha)": "sum"}
                )
                pdays = _period_days_jun(g[period_keys])
                g = g.merge(pdays[period_keys + ["期間日数", "ラベル"]], on=period_keys, how="left")
                g["必要稼働率"] = np.where(
                    pd.to_numeric(g["期間日数"], errors="coerce") > 0,
                    g["推定機械日数_有効"] / g["期間日数"],
                    np.nan,
                )
                g["シナリオ"] = scen_name
                plot_rows.append(g)

            if plot_rows:
                plot_df_r = pd.concat(plot_rows, ignore_index=True).sort_values(["シナリオ"] + sort_cols)
                cap_map = cap_long.set_index(["農家名", "農機カテゴリ"])["能力(ha/日)"].to_dict() if cap_long is not None and not cap_long.empty else {}
                plot_df_r["農機能力(ha/日)"] = plot_df_r.apply(lambda r: cap_map.get((farm_sel, str(r.get("農機カテゴリ", "")).strip())), axis=1)
                ymax = float(np.nanmax(pd.to_numeric(plot_df_r["必要稼働率"], errors="coerce").to_numpy())) if not plot_df_r["必要稼働率"].dropna().empty else 0.0

                fig_r = px.bar(
                    plot_df_r,
                    x="ラベル",
                    y="必要稼働率",
                    color="作業表示",
                    color_discrete_sequence=px.colors.qualitative.Safe,
                    barmode="stack",
                    facet_col="シナリオ" if compare_mode else None,
                    hover_data={
                        "農機カテゴリ": True,
                        "農機能力(ha/日)": ":.2f",
                        "推定機械日数_有効": ":.2f",
                        "面積(ha)": ":.2f",
                        "期間日数": ":.0f",
                    },
                    title=None,
                )
                fig_r.update_traces(marker_line_width=0)
                fig_r.update_yaxes(tickformat=".0%", rangemode="tozero", range=[0, max(ymax * 1.1, report_util * 1.2, 0.1)])
                fig_r.update_layout(
                    template="plotly_white",
                    height=520,
                    bargap=0.25,
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="top", y=-0.25, xanchor="left", x=0, title_text=""),
                    margin=dict(t=30, r=20, b=150, l=60),
                )
                fig_r.add_hline(y=report_util, line_dash="dash", line_color="red")
                cat_arr = [jn_to_label(i) for i in range(1, 37)]
                for ax_name in [k for k in fig_r.layout if str(k).startswith("xaxis")]:
                    fig_r.layout[ax_name].update({"categoryorder": "array", "categoryarray": cat_arr})
                fig_r.update_xaxes(tickangle=-45, automargin=True, title_text="")
                fig_html = _fig_to_html(fig_r)

                # bottleneck table by period (report)
                bn_tbl_r = (
                    plot_df_r.groupby(period_keys + ["ラベル", "農機カテゴリ", "シナリオ"], as_index=False)
                    .agg({"推定機械日数_有効": "sum", "面積(ha)": "sum", "期間日数": "max"})
                    .sort_values(["シナリオ"] + sort_cols)
                )
                bn_tbl_r["必要稼働率(%)"] = np.where(
                    pd.to_numeric(bn_tbl_r["期間日数"], errors="coerce") > 0,
                    pd.to_numeric(bn_tbl_r["推定機械日数_有効"], errors="coerce") / pd.to_numeric(bn_tbl_r["期間日数"], errors="coerce") * 100.0,
                    np.nan,
                )
                bn_tbl_r["農機稼働率(%)"] = report_util * 100.0
                bn_tbl_r["超過(%)"] = bn_tbl_r["必要稼働率(%)"] - bn_tbl_r["農機稼働率(%)"]
                bn_tbl_r = bn_tbl_r.rename(columns={"農機カテゴリ": "ボトルネック農機カテゴリ"})
                bn_tbl_r["超過(%)"] = pd.to_numeric(bn_tbl_r["超過(%)"], errors="coerce")
                bn_tbl_r = bn_tbl_r[bn_tbl_r["超過(%)"] > 0].copy()
                if bn_tbl_r.empty:
                    bn_html = "<p class='muted'>（超過はありません）</p>"
                else:
                    bn_tbl_r = bn_tbl_r[["シナリオ", "年", "旬ラベル", "ボトルネック農機カテゴリ", "必要稼働率(%)", "農機稼働率(%)", "超過(%)"]].copy()
                    bn_tbl_r["必要稼働率(%)"] = pd.to_numeric(bn_tbl_r["必要稼働率(%)"], errors="coerce").round(0)
                    bn_tbl_r["農機稼働率(%)"] = pd.to_numeric(bn_tbl_r["農機稼働率(%)"], errors="coerce").round(0)
                    bn_tbl_r["超過(%)"] = pd.to_numeric(bn_tbl_r["超過(%)"], errors="coerce").round(0)
                    bn_html = _df_to_html_table(bn_tbl_r)

        # ---- Profile tables (reuse app logic, but for report we keep it compact)
        crop_sum = pd.DataFrame()
        mach_profile = pd.DataFrame()
        try:
            df_sak_f = st.session_state.df_sak_edit.copy()
            if "農家名" in df_sak_f.columns:
                df_sak_f = df_sak_f[df_sak_f["農家名"].astype(str).str.strip() == str(farm_sel).strip()].copy()
            if ("作物" in df_sak_f.columns) and ("面積(ha)" in df_sak_f.columns) and (not df_sak_f.empty):
                df_sak_f["作物"] = df_sak_f["作物"].astype(str).str.strip()
                df_sak_f["面積(ha)"] = pd.to_numeric(df_sak_f["面積(ha)"], errors="coerce").fillna(0.0)
                crop_sum = (
                    df_sak_f.dropna(subset=["作物"])
                    .groupby("作物", as_index=False)["面積(ha)"]
                    .sum()
                    .sort_values("面積(ha)", ascending=False)
                )
                crop_sum = crop_sum[crop_sum["面積(ha)"].fillna(0) > 0].copy()
                crop_sum["面積(ha)"] = pd.to_numeric(crop_sum["面積(ha)"], errors="coerce").round(2)
        except Exception:
            crop_sum = pd.DataFrame()

        try:
            mach_map = {
                "tractor": ("トラクタ", "トラクタ推定台数(仮)"),
                "seeder": ("播種機", "播種機推定台数(仮)"),
                "transplanter": ("田植機", "田植機推定台数(仮)"),
                "sprayer": ("防除機", "防除推定台数(仮)"),
                "combine": ("コンバイン", "コンバイン推定台数(仮)"),
                "roller": ("鎮圧機", "鎮圧推定台数(仮)"),
            }
            cap_f = cap_long[cap_long["農家名"].astype(str).str.strip() == str(farm_sel).strip()].copy() if cap_long is not None else pd.DataFrame()
            cap_f = cap_f.dropna(subset=["農機カテゴリ"]).copy() if not cap_f.empty else cap_f
            if not cap_f.empty:
                cap_f["農機カテゴリ"] = cap_f["農機カテゴリ"].astype(str).str.strip()
            cap_map = cap_f.set_index("農機カテゴリ")["能力(ha/日)"].to_dict() if not cap_f.empty else {}

            df_mach_f = df_mach.copy() if df_mach is not None else pd.DataFrame()
            if "農家名" in df_mach_f.columns:
                df_mach_f["農家名"] = df_mach_f["農家名"].astype(str).str.strip()
                df_mach_f = df_mach_f[df_mach_f["農家名"] == str(farm_sel).strip()].copy()
            df_mach_row = df_mach_f.iloc[0] if df_mach_f is not None and not df_mach_f.empty else None

            rows = []
            for cat in sorted(list(VALID_MACHINE_CATS)):
                label, n_col = mach_map.get(cat, (cat, None))
                cap_val = cap_map.get(cat, np.nan)
                if pd.isna(cap_val) or float(cap_val) <= 0:
                    continue
                n_val = np.nan
                if df_mach_row is not None and n_col and n_col in df_mach_row.index:
                    n_val = pd.to_numeric(df_mach_row.get(n_col), errors="coerce")
                rows.append(
                    {
                        "農機": label,
                        "推定台数(仮)": ("" if pd.isna(n_val) else f"{float(n_val):.0f}"),
                        "1日あたりの作業面積(ha/日)": f"{float(cap_val):.2f}",
                    }
                )
            mach_profile = pd.DataFrame(rows)
        except Exception:
            mach_profile = pd.DataFrame()

        # ---- Build HTML
        today_s = dt.date.today().isoformat()
        peak_base_s = "—" if pd.isna(peak_base) else f"{peak_base*100:.0f}%"
        peak_edit_s = "—" if pd.isna(peak_edit) else f"{peak_edit*100:.0f}%"
        util_s = f"{report_util*100:.0f}%"
        peak_cat_s = peak_cat_r if peak_cat_r else "—"
        peak_when_s = peak_when if peak_when else "—"

        crop_html = _df_to_html_table(crop_sum)
        mach_html = _df_to_html_table(mach_profile)

        report_html = f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>農機負荷レポート_{farm_sel}_{today_s}</title>
  <style>
    :root {{
      --text:#111;
      --muted:#666;
      --border:#e5e7eb;
      --bg:#f1f5f9;
      --paper:#fff;
      --card:#f8fafc;
      --danger:#dc2626;
      --ok:#16a34a;
      --shadow: 0 10px 30px rgba(15, 23, 42, 0.08);
    }}
    html,body{{background:var(--bg); color:var(--text); font-family:-apple-system,BlinkMacSystemFont,\"Hiragino Sans\",\"Noto Sans JP\",Segoe UI,Roboto,Helvetica,Arial,sans-serif;}}
    .page{{max-width:980px; margin:24px auto; padding:22px 22px; background:var(--paper); border:1px solid var(--border); border-radius:14px; box-shadow:var(--shadow);}}
    h1{{font-size:22px; margin:0 0 4px; letter-spacing:0.02em;}}
    h2{{font-size:16px; margin:22px 0 10px; padding-top:4px; border-top:1px solid var(--border);}}
    h3{{font-size:13px; margin:12px 0 6px; color:#0f172a;}}
    .meta{{color:var(--muted); font-size:12px; margin-bottom:14px;}}
    .grid{{display:grid; grid-template-columns: repeat(4, 1fr); gap:10px;}}
    .card{{background:var(--card); border:1px solid var(--border); border-radius:12px; padding:10px 12px;}}
    .k{{font-size:12px; color:var(--muted);}}
    .v{{font-size:18px; font-weight:700; margin-top:4px;}}
    .note{{color:var(--muted); font-size:12px; margin-top:8px;}}
    .danger{{color:var(--danger); font-weight:700;}}
    .table{{width:100%; border-collapse:collapse; font-size:12px;}}
    .table th,.table td{{border:1px solid var(--border); padding:6px 8px; vertical-align:top;}}
    .table th{{background:#f3f4f6; text-align:left;}}
    .table tr:nth-child(even) td{{background:#fafafa;}}
    .muted{{color:var(--muted);}}
    .section{{page-break-inside: avoid;}}
    .headerbar{{height:6px; border-radius:999px; background:linear-gradient(90deg,#2563eb,#22c55e,#f59e0b,#ef4444); margin-bottom:12px;}}
    @media print {{
      html,body{{background:#fff;}}
      .page{{max-width:100%; margin:0; padding:0; background:#fff; border:none; border-radius:0; box-shadow:none;}}
      @page {{ size: A4; margin: 12mm; }}
      h1,h2{{page-break-after: avoid;}}
    }}
  </style>
</head>
<body>
  <div class="page">
    <div class="section">
      <div class="headerbar"></div>
      <h1>農機負荷レポート（旬別）</h1>
      <div class="meta">農家: {farm_sel} / 作成日: {today_s} / 農機稼働率（赤線）: {util_s}</div>
      <div class="grid">
        <div class="card"><div class="k">ピーク必要稼働率（現在）</div><div class="v">{peak_base_s}</div></div>
        <div class="card"><div class="k">ピーク必要稼働率（編集後）</div><div class="v">{peak_edit_s}</div></div>
        <div class="card"><div class="k">編集後ピーク時期</div><div class="v">{peak_when_s}</div></div>
        <div class="card"><div class="k">編集後ボトルネック農機</div><div class="v">{peak_cat_s}</div></div>
      </div>
      <div class="note">必要稼働率 = 推定機械日数 ÷ 期間日数。赤線（農機稼働率%）を超える旬は「回し切れない」目安です。</div>
    </div>

    <div class="section">
      <h2>プロフィール</h2>
      <h3>作物別の合計面積（ha）</h3>
      {crop_html}
      <h3>農機プロフィール（1日あたりの作業面積）</h3>
      {mach_html}
      <div class="note">推定台数(仮)が空の場合は未入力です。能力(ha/日)は入力が無い場合、前提(仮置き)×推定台数で補完した値になります。</div>
    </div>

    <div class="section">
      <h2>労働負荷グラフ（旬別）</h2>
      <div class="note">注意：このグラフの「全て」は各旬の“ボトルネック農機”だけを表示します（全量の合算ではありません）。詳細は農機カテゴリ別で確認します。</div>
      {fig_html}
    </div>

    <div class="section">
      <h2>根拠（各旬のボトルネック）</h2>
      {bn_html}
    </div>
  </div>
</body>
</html>
"""

        st.download_button(
            "HTMLレポートをダウンロード",
            data=report_html.encode("utf-8"),
            file_name=f"{farm_sel}_report_jun.html",
            mime="text/html",
        )

with st.expander("計算済み明細（CSV）", expanded=False):
    csv = df.to_csv(index=False).encode("utf-8-sig")
    st.download_button("CSVダウンロード（計算済み明細）", data=csv, file_name=f"{farm_sel}_tasks.csv", mime="text/csv")

with st.expander("更新したExcelをダウンロード", expanded=False):
    extra = {
        "streamlit_tasks": df.copy(),
    }
    if st.button("Excelを生成", key="btn_build_excel"):
        try:
            out_bytes = write_back_to_excel(uploaded.getvalue(), st.session_state.df_sak_edit, extra)
            st.download_button(
                "Excelダウンロード（入力更新 + 出力シート追加）",
                data=out_bytes,
                file_name="CS_board_streamlit_updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Excel生成に失敗しました: {e}")
