import io
import os
import re
import math
import asyncio
from collections import defaultdict
import csv
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
import streamlit as st

try:
    import httpx
    CONCURRENT_FETCH_ENABLED = True
except ImportError:
    httpx = None
    CONCURRENT_FETCH_ENABLED = False

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference


DEFAULT_GIGYA_BASE = os.getenv("GIGYA_BASE", "https://accounts.eu1.gigya.com")
DEFAULT_GIGYA_API_KEY = os.getenv("GIGYA_API_KEY", "3_W-AXsoj7TvX-9gi7S-IGxXfLWVkEbnGSl57M7t49GN538umaKs2EID8hyipAux2y")
DEFAULT_TOKEN_URL = os.getenv("XARVIO_TOKEN_API_URL", "https://fm-api.xarvio.com/api/users/tokens")
DEFAULT_GRAPHQL_URL = os.getenv("XARVIO_GRAPHQL_ENDPOINT", "https://fm-api.xarvio.com/api/graphql/data")
REQUEST_TIMEOUT_SEC = 60
SCAN_CHUNK_SIZE = 50
JST = ZoneInfo("Asia/Tokyo")

FARMS_OVERVIEW = """
query FarmsOverview {
  farms: farmsV2(uuids: []) {
    uuid
    name
  }
}
"""

FIELDS_BY_FARMS = """
query FieldsByFarms($farmUuids: [UUID!]!) {
  fieldsV2(farmUuids: $farmUuids) {
    uuid
    name
    farmV2 { uuid name }
  }
}
"""

FDL_CLASSIFICATION_QUERY = """
query UseGetFieldDataLayerClassification($fieldDataLayerUuid: UUID!, $magnitude: MagnitudeType, $elevation: [String], $locale: String!) {
  fieldDataLayerClassifications(
    fieldDataLayerUuid: $fieldDataLayerUuid
    magnitudes: [$magnitude]
    elevation: $elevation
    locale: $locale
  ) {
    fieldDataLayerUuid
    processingType
    magnitudeType
    minElevation
    maxElevation
    unit
    elevation
    classificationEntries {
      color
      value
      minValue
      maxValue
      average
      label
      area
      percentage
      featureId
      __typename
    }
    __typename
  }
}
"""

ANALYTICS_MAP_QUERY = """
query AnalyticsMap($fieldUuid: UUID!) {
  fieldV2(uuid: $fieldUuid) {
    fieldDataLayers {
      uuid
      date
      type
      magnitudes {
        type
        fieldDataLayer {
          uuid
          type
          date
        }
      }
    }
  }
}
"""


def make_payload(operation_name: str, query: str, variables: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "operationName": operation_name,
        "query": query,
        "variables": variables,
    }


def chunked(items: List[str], size: int) -> Iterable[List[str]]:
    for i in range(0, len(items), size):
        yield items[i:i + size]


def gigya_login(gigya_base: str, gigya_api_key: str, email: str, password: str) -> Dict[str, str]:
    if not gigya_api_key:
        raise RuntimeError("GIGYA_API_KEY が未設定です。.env に設定してください。")

    resp = requests.post(
        f"{gigya_base.rstrip('/')}/accounts.login",
        data={"apiKey": gigya_api_key, "loginID": email, "password": password, "format": "json"},
        timeout=REQUEST_TIMEOUT_SEC,
    )
    resp.raise_for_status()
    out = resp.json()
    if (out.get("errorCode") or 0) != 0:
        error_code = out.get("errorCode")
        # 認証エラー（メールアドレスまたはパスワード不正）
        if error_code in (403042, 403047, 400006):
            raise RuntimeError("メールアドレスまたはパスワードが正しくありません。確認してください。")
        raise RuntimeError(
            f"ログイン失敗: errorCode={error_code} errorMessage={out.get('errorMessage')}"
        )

    session_info = out.get("sessionInfo") or {}
    return {
        "login_token": str(session_info.get("cookieValue") or ""),
        "gigya_uuid": str(out.get("UID") or ""),
        "gigya_uuid_signature": str(out.get("UIDSignature") or ""),
        "gigya_signature_timestamp": str(out.get("signatureTimestamp") or ""),
    }


def issue_xarvio_token(token_url: str, login_data: Dict[str, str]) -> str:
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Cookie": f"LOGIN_TOKEN={login_data['login_token']}",
        "Origin": "https://fm.xarvio.com",
        "Referer": "https://fm.xarvio.com/",
        "User-Agent": "xarvio-ndvi-lai-tool/1.3",
    }
    payload = {
        "gigyaUuid": login_data["gigya_uuid"],
        "gigyaUuidSignature": login_data["gigya_uuid_signature"],
        "gigyaSignatureTimestamp": login_data["gigya_signature_timestamp"],
    }
    resp = requests.post(token_url, json=payload, headers=headers, timeout=REQUEST_TIMEOUT_SEC)
    resp.raise_for_status()
    out = resp.json()
    token = str(out.get("token") or "")
    if not token:
        raise RuntimeError(f"DF token missing in response: {out}")
    return token


def call_xarvio_graphql(
    graphql_url: str,
    login_token: str,
    api_token: str,
    operation_name: str,
    query: str,
    variables: Dict[str, Any],
) -> Dict[str, Any]:
    payload = make_payload(operation_name, query, variables)
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Cookie": f"LOGIN_TOKEN={login_token}; DF_TOKEN={api_token}",
        "Origin": "https://fm.xarvio.com",
        "Referer": "https://fm.xarvio.com/",
        "User-Agent": "xarvio-ndvi-lai-tool/1.3",
    }
    resp = requests.post(graphql_url, json=payload, headers=headers, timeout=REQUEST_TIMEOUT_SEC)
    resp.raise_for_status()
    out = resp.json()
    if out.get("errors"):
        raise RuntimeError(f"GraphQL errors ({operation_name}): {out.get('errors')}")
    return out


async def async_call_xarvio_graphql(
    client,
    graphql_url: str,
    login_token: str,
    api_token: str,
    operation_name: str,
    query: str,
    variables: Dict[str, Any],
    context: Dict[str, Any],
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    payload = make_payload(operation_name, query, variables)
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Cookie": f"LOGIN_TOKEN={login_token}; DF_TOKEN={api_token}",
        "Origin": "https://fm.xarvio.com",
        "Referer": "https://fm.xarvio.com/",
        "User-Agent": "xarvio-ndvi-lai-tool/1.3",
    }
    resp = await client.post(graphql_url, json=payload, headers=headers, timeout=REQUEST_TIMEOUT_SEC)
    resp.raise_for_status()
    out = resp.json()
    if out.get("errors"):
        raise RuntimeError(f"GraphQL errors ({operation_name}) for {context}: {out.get('errors')}")
    return out, context


def parse_farms(gql_out: Dict[str, Any]) -> List[Dict[str, str]]:
    farms = ((gql_out.get("data") or {}).get("farms") or [])
    result: List[Dict[str, str]] = []
    for farm in farms:
        if not isinstance(farm, dict):
            continue
        uuid = str(farm.get("uuid") or "").strip()
        if not uuid:
            continue
        result.append({"uuid": uuid, "name": str(farm.get("name") or "")})
    return result


def parse_fields(gql_out: Dict[str, Any]) -> List[Dict[str, str]]:
    fields = ((gql_out.get("data") or {}).get("fieldsV2") or [])
    result: List[Dict[str, str]] = []
    for field in fields:
        if not isinstance(field, dict):
            continue
        field_uuid = str(field.get("uuid") or "").strip()
        if not field_uuid:
            continue
        farm = field.get("farmV2") or {}
        result.append(
            {
                "field_uuid": field_uuid,
                "field_name": str(field.get("name") or ""),
                "farm_uuid": str(farm.get("uuid") or ""),
                "farm_name": str(farm.get("name") or ""),
            }
        )
    return result


def parse_fdl_uuid_lines(text: str) -> List[str]:
    out: List[str] = []
    seen: set[str] = set()
    for raw in text.replace(",", "\n").splitlines():
        uuid = raw.strip()
        if not uuid or uuid in seen:
            continue
        seen.add(uuid)
        out.append(uuid)
    return out


def extract_fdl_entry_value(entry: Dict[str, Any]) -> Optional[float]:
    for key in ("average", "value", "minValue", "maxValue"):
        value = to_float_or_none(entry.get(key))
        if value is not None:
            return value
    return None


def metric_key_from_magnitude(magnitude: str) -> Optional[str]:
    mapping = {
        "LAI": "lai",
        "LAI_CONTRAST": "lai_contrast",
        "NDVI": "ndvi",
        "AVERAGE_NDVI": "ndvi",
        "NDVI_CONTRAST": "ndvi_contrast",
    }
    return mapping.get(magnitude.strip().upper())


def normalize_magnitude_type(mt: str) -> str:
    mt = mt.strip().upper()
    alias_map = {
        "AVERAGE_NDVI": "NDVI",
    }
    return alias_map.get(mt, mt)


def detect_map_type(layer_type: str, magnitude: Optional[str] = None) -> str:
    lt = (layer_type or "").strip().upper()
    mg = (magnitude or "").strip().upper()

    if lt == "BIOMASS_MULTI_IMAGE_LAI":
        return "地力マップ"
    if lt == "BIOMASS_SINGLE_IMAGE_LAI":
        if mg == "LAI_CONTRAST":
            return "生育マップ（相対表示）"
        return "生育マップ（絶対表示）"
    if lt == "BIOMASS_NDVI":
        if mg == "AVERAGE_NDVI":
            return "NDVI平均植生"
        if mg == "NDVI_CONTRAST":
            return "NDVIマップ（相対表示）"
        return "NDVIマップ（絶対表示）"
    if lt == "BIOMASS_SINGLE_IMAGE_NDVI_FIELD_AVERAGE":
        return "平均植生のマップ"

    return layer_type or "不明"


def derive_map_type_label(row: Dict[str, Any], fallback: str = "") -> str:
    has_lai = row.get("lai") is not None
    has_lai_contrast = row.get("lai_contrast") is not None
    has_ndvi = row.get("ndvi") is not None
    has_ndvi_contrast = row.get("ndvi_contrast") is not None

    layer_type = row.get("fdl_layer_type")
    if layer_type:
        if layer_type == "BIOMASS_MULTI_IMAGE_LAI":
            return "地力マップ"
        if layer_type == "BIOMASS_SINGLE_IMAGE_LAI":
            if has_lai and has_lai_contrast:
                return "生育マップ（絶対＋相対）"
            if has_lai_contrast:
                return detect_map_type(layer_type, "LAI_CONTRAST")
            if has_lai:
                return detect_map_type(layer_type, "LAI")
        if layer_type == "BIOMASS_NDVI":
            if has_ndvi and has_ndvi_contrast:
                return "NDVIマップ（絶対＋相対）"
            if has_ndvi_contrast:
                return detect_map_type(layer_type, "NDVI_CONTRAST")
            if has_ndvi:
                return detect_map_type(layer_type, "NDVI")
        if layer_type == "BIOMASS_SINGLE_IMAGE_NDVI_FIELD_AVERAGE":
            return detect_map_type(layer_type)

    if fallback == "全マップ（LAI/NDVI すべて）":
        if has_lai and has_lai_contrast:
            return "生育マップ（絶対＋相対）"
        if has_ndvi and has_ndvi_contrast:
            return "NDVIマップ（絶対＋相対）"
        if has_lai_contrast:
            return "生育マップ（相対表示）"
        if has_ndvi_contrast:
            return "NDVIマップ（相対表示）"
        if has_lai:
            return "生育マップ（絶対表示）"
        if has_ndvi:
            return "NDVIマップ（絶対表示）"
        return fallback

    if (has_lai or has_lai_contrast) and (has_ndvi or has_ndvi_contrast):
        return "全マップ（LAI/NDVI）"
    if has_ndvi and has_ndvi_contrast:
        return "植生マップ（NDVI + NDVI_CONTRAST）"
    if has_ndvi:
        return "植生マップ（NDVIのみ）"
    if has_lai and has_lai_contrast:
        return "生育マップ（LAI + LAI_CONTRAST）"
    if has_lai:
        return "地力マップ（LAIのみ）"
    return fallback


def target_magnitudes_by_fdl_map_type(fdl_map_type: str) -> Tuple[str, Optional[str]]:
    if fdl_map_type == "生育マップ（LAI + LAI_CONTRAST）":
        return "LAI", "LAI_CONTRAST"
    if fdl_map_type == "地力マップ（LAIのみ）":
        return "LAI", None
    if fdl_map_type == "植生マップ（NDVI + NDVI_CONTRAST）":
        return "NDVI", "NDVI_CONTRAST"
    if fdl_map_type == "植生マップ（NDVIのみ）":
        return "NDVI", None
    return "LAI", "LAI_CONTRAST"


def target_magnitudes_all(fdl_map_type: str) -> List[str]:
    if fdl_map_type == "全マップ（LAI/NDVI すべて）":
        return ["LAI", "LAI_CONTRAST", "NDVI", "NDVI_CONTRAST"]
    primary_magnitude, contrast_magnitude = target_magnitudes_by_fdl_map_type(fdl_map_type)
    magnitudes = [primary_magnitude]
    if contrast_magnitude:
        magnitudes.append(contrast_magnitude)
    return magnitudes


def is_target_magnitude(magnitude_type: str, target_set: set[str]) -> bool:
    mt_raw = magnitude_type.strip().upper()
    if not mt_raw:
        return False
    mt = normalize_magnitude_type(mt_raw)
    normalized_targets = {normalize_magnitude_type(t) for t in target_set}
    if mt in normalized_targets:
        return True
    return any(mt.startswith(t + "_") for t in normalized_targets)


def extract_fdl_uuids_from_analytics_map(
    gql_out: Dict[str, Any], target_magnitudes: Sequence[str]
) -> Tuple[List[Tuple[str, Optional[date], Optional[str], str]], List[str]]:
    target_set = {t.strip().upper() for t in target_magnitudes if t}
    seen: set[str] = set()
    out: List[Tuple[str, Optional[date], Optional[str], str]] = []
    observed_types: set[str] = set()
    field = (gql_out.get("data") or {}).get("fieldV2") or {}

    for layer in field.get("fieldDataLayers") or []:
        layer_uuid = str(layer.get("uuid") or "")
        layer_date_raw = layer.get("date")
        layer_date = normalize_iso_date(layer_date_raw)
        layer_type_fallback = str(layer.get("type") or "")
        magnitudes = layer.get("magnitudes") or []
        matched = False
        matched_layer_specific = False

        for m in magnitudes:
            m_type = str(m.get("type") or "")
            if m_type:
                observed_types.add(m_type)

            if is_target_magnitude(m_type, target_set):
                matched = True
                fdl_node = m.get("fieldDataLayer") or {}
                mag_layer_uuid = str(fdl_node.get("uuid") or "").strip()
                layer_type = str(fdl_node.get("type") or layer_type_fallback)
                mag_layer_date_raw = fdl_node.get("date") or layer_date_raw

                if mag_layer_uuid and mag_layer_uuid not in seen:
                    matched_layer_specific = True
                    seen.add(mag_layer_uuid)
                    out.append((mag_layer_uuid, normalize_iso_date(mag_layer_date_raw), mag_layer_date_raw, layer_type))

        if matched and not matched_layer_specific and layer_uuid and layer_uuid not in seen:
            seen.add(layer_uuid)
            out.append((layer_uuid, layer_date, layer_date_raw, layer_type_fallback))

    return out, sorted(observed_types)


def extract_fdl_rows(
    classification_outputs: Dict[str, Dict[str, Any]],
    field_data_layer_uuid: str,
    observed_at: date,
    observed_at_raw: Optional[str],
    field_uuid: str,
    field_name: str,
    farm_uuid: str,
    farm_name: str,
    requested_map_type: str,
    layer_type: str = "",
) -> List[Dict[str, Any]]:
    zone_map: Dict[Tuple[str, str], Dict[str, Any]] = {}

    for magnitude, payload in classification_outputs.items():
        classifications = ((payload.get("data") or {}).get("fieldDataLayerClassifications") or [])
        for classification in classifications:
            min_elevation = classification.get("minElevation")
            max_elevation = classification.get("maxElevation")
            elevation = classification.get("elevation")
            processing_type = str(classification.get("processingType") or "")
            unit = str(classification.get("unit") or "")
            entries = classification.get("classificationEntries") or []

            for entry in entries:
                zone_id = str(entry.get("featureId") or "")
                zone_name = str(entry.get("label") or "")
                key = (zone_id, zone_name)

                if key not in zone_map:
                    zone_map[key] = {
                        "field_uuid": field_uuid,
                        "field_name": field_name,
                        "farm_uuid": farm_uuid,
                        "farm_name": farm_name,
                        "observed_at": observed_at,
                        "observed_at_raw": observed_at_raw,
                        "zone_id": zone_id,
                        "zone_name": zone_name,
                        "map_type": requested_map_type,
                        "ndvi": None,
                        "ndvi_contrast": None,
                        "lai": None,
                        "lai_contrast": None,
                        "field_data_layer_uuid": field_data_layer_uuid,
                        "fdl_layer_type": layer_type,
                        "fdl_processing_type": processing_type,
                        "fdl_unit": unit,
                        "fdl_min_elevation": min_elevation,
                        "fdl_max_elevation": max_elevation,
                        "fdl_elevation": elevation,
                        "source_payload": entry,
                    }

                value = extract_fdl_entry_value(entry)
                metric_key = metric_key_from_magnitude(magnitude)
                if metric_key:
                    zone_map[key][metric_key] = value

    rows = list(zone_map.values())
    for row in rows:
        row["map_type"] = derive_map_type_label(row, fallback=requested_map_type)
    rows.sort(key=lambda x: (x["field_name"], x["observed_at"], x["zone_name"], x["zone_id"]))
    return rows


def normalize_iso_date(raw: Any) -> Optional[date]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    text = str(raw).strip()
    if not text:
        return None
    text = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            if fmt == "%Y%m%d":
                return datetime.strptime(text[:8], fmt).date()
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def parse_iso_datetime(raw: Any) -> Optional[datetime]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        dt = raw
    else:
        text = str(raw).strip()
        if not text:
            return None
        text = text.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(text)
        except ValueError:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def to_jst_datetime_string(raw: Any) -> Optional[str]:
    dt = parse_iso_datetime(raw)
    if dt is None:
        return None
    return dt.astimezone(JST).strftime("%Y-%m-%d %H:%M:%S JST")


def to_jst_date_string(raw: Any) -> Optional[str]:
    dt = parse_iso_datetime(raw)
    if dt is None:
        return None
    return dt.astimezone(JST).date().isoformat()


def to_float_or_none(value: Any) -> Optional[float]:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def sanitize_sheet_name(raw: str) -> str:
    invalid_chars = set(r'[]:*?/\ ')
    cleaned = "".join("_" if c in invalid_chars else c for c in raw).strip()
    if not cleaned:
        cleaned = "Field"
    return cleaned[:31]


def safe_avg(values: Sequence[float]) -> Optional[float]:
    return sum(values) / len(values) if values else None


def safe_max(values: Sequence[float]) -> Optional[float]:
    return max(values) if values else None


def safe_min(values: Sequence[float]) -> Optional[float]:
    return min(values) if values else None


def safe_range(values: Sequence[float]) -> Optional[float]:
    return (max(values) - min(values)) if values else None


def none_if_nan(value: Any) -> Any:
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def normalize_zone_label(row: Dict[str, Any]) -> str:
    zone_name = str(row.get("zone_name") or "").strip()
    zone_id = str(row.get("zone_id") or "").strip()
    if zone_name:
        return zone_name
    if zone_id:
        return f"class_{zone_id}"
    return "unclassified"


def classify_record_type(row: Dict[str, Any]) -> str:
    payload = row.get("source_payload") or {}
    has_range = (
        payload.get("minValue") is not None
        or payload.get("maxValue") is not None
        or payload.get("area") is not None
        or payload.get("percentage") is not None
    )
    zone_name = str(row.get("zone_name") or "").strip()
    zone_id = str(row.get("zone_id") or "").strip()
    if has_range and (zone_name or zone_id):
        return "zone"
    if has_range:
        return "class_range"
    return "metric_only"


def extract_metric_code_from_row(row: Dict[str, Any]) -> str:
    map_type = str(row.get("map_type") or "").lower()
    if row.get("ndvi_contrast") is not None:
        return "NDVI_CONTRAST"
    if row.get("ndvi") is not None and "ndvi" in map_type:
        return "NDVI"
    if row.get("lai_contrast") is not None:
        return "LAI_CONTRAST"
    if row.get("lai") is not None:
        return "LAI"
    if "ndvi" in map_type and "相対" in map_type:
        return "NDVI_CONTRAST"
    if "ndvi" in map_type:
        return "NDVI"
    if "相対" in map_type:
        return "LAI_CONTRAST"
    return "LAI"


def extract_primary_metric_value(row: Dict[str, Any]) -> Optional[float]:
    metric_code = extract_metric_code_from_row(row)
    if metric_code == "NDVI":
        return to_float_or_none(row.get("ndvi"))
    if metric_code == "NDVI_CONTRAST":
        return to_float_or_none(row.get("ndvi_contrast"))
    if metric_code == "LAI":
        return to_float_or_none(row.get("lai"))
    if metric_code == "LAI_CONTRAST":
        return to_float_or_none(row.get("lai_contrast"))
    return None


def map_type_code(map_type: str) -> str:
    mt = (map_type or "").strip()
    mapping = {
        "地力マップ": "soil_map",
        "地力マップ（LAIのみ）": "soil_map_lai",
        "生育マップ（絶対表示）": "lai_absolute",
        "生育マップ（相対表示）": "lai_relative",
        "生育マップ（絶対＋相対）": "lai_absolute_relative",
        "生育マップ（LAI + LAI_CONTRAST）": "lai_absolute_relative",
        "植生マップ（NDVIのみ）": "ndvi_absolute",
        "NDVIマップ（絶対表示）": "ndvi_absolute",
        "NDVIマップ（相対表示）": "ndvi_relative",
        "NDVIマップ（絶対＋相対）": "ndvi_absolute_relative",
        "植生マップ（NDVI + NDVI_CONTRAST）": "ndvi_absolute_relative",
        "全マップ（LAI/NDVI）": "all_maps",
        "全マップ（LAI/NDVI すべて）": "all_maps",
        "平均植生のマップ": "ndvi_field_average",
        "NDVI平均植生": "ndvi_average",
    }
    return mapping.get(mt, re.sub(r"[^a-z0-9_]+", "_", mt.lower()).strip("_") or "unknown")


def to_csv(data: List[Dict[str, Any]]) -> str:
    if not data:
        return ""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()


def build_detail_csv_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """1行 = 1圃場 × 1日 × 1ゾーンの完全生データCSV"""
    out: List[Dict[str, Any]] = []
    for row in rows:
        observed_at_jst = to_jst_datetime_string(row.get("observed_at_raw"))
        observed_date_jst = to_jst_date_string(row.get("observed_at_raw"))
        if observed_at_jst is None and row.get("observed_at"):
            observed_at_jst = f"{row['observed_at'].isoformat()} 00:00:00 JST"
        if observed_date_jst is None and row.get("observed_at"):
            observed_date_jst = row["observed_at"].isoformat()

        payload = row.get("source_payload") or {}
        out.append({
            "observed_date_jst":      observed_date_jst,
            "observed_at_jst":        observed_at_jst,
            "observed_at_utc_raw":    row.get("observed_at_raw"),
            "farm_name":              row.get("farm_name"),
            "farm_uuid":              row.get("farm_uuid"),
            "field_name":             row.get("field_name"),
            "field_uuid":             row.get("field_uuid"),
            "zone_label":             normalize_zone_label(row),
            "zone_id":                row.get("zone_id"),
            "map_type":               row.get("map_type"),
            "ndvi":                   none_if_nan(to_float_or_none(row.get("ndvi"))),
            "ndvi_contrast":          none_if_nan(to_float_or_none(row.get("ndvi_contrast"))),
            "lai":                    none_if_nan(to_float_or_none(row.get("lai"))),
            "lai_contrast":           none_if_nan(to_float_or_none(row.get("lai_contrast"))),
            "class_area":             none_if_nan(to_float_or_none(payload.get("area"))),
            "class_percentage":       none_if_nan(to_float_or_none(payload.get("percentage"))),
            "class_avg_value":        none_if_nan(to_float_or_none(payload.get("average"))),
            "class_min_value":        none_if_nan(to_float_or_none(payload.get("minValue"))),
            "class_max_value":        none_if_nan(to_float_or_none(payload.get("maxValue"))),
            "fdl_layer_type":         row.get("fdl_layer_type"),
            "field_data_layer_uuid":  row.get("field_data_layer_uuid"),
        })

    out.sort(key=lambda x: (
        str(x.get("farm_name") or ""),
        str(x.get("field_name") or ""),
        str(x.get("observed_date_jst") or ""),
        str(x.get("zone_label") or ""),
    ))
    return out


def build_excel_report(
    rows: List[Dict[str, Any]],
    fields: List[Dict[str, str]],
    from_date: date,
    till_date: date,
) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

    # ---- スタイル定数 ----
    DARK   = "1F6B3C"
    MID    = "2E8B57"
    SUB    = "4CAF7D"
    SOIL   = "7B5E3A"   # 地力マップ用（茶系）
    ALT    = "F0F7ED"
    WHITE  = "FFFFFF"
    BC     = "AAAAAA"

    def _b():
        s = Side(style="thin", color=BC)
        return Border(left=s, right=s, top=s, bottom=s)
    def _f(c): return PatternFill("solid", fgColor=c)
    def _h(cell, val, bg=DARK, sz=10, wrap=False):
        cell.value = val
        cell.font  = Font(name="Arial", bold=True, color="FFFFFF", size=sz)
        cell.fill  = _f(bg); cell.border = _b()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def _d(cell, val, bg=WHITE, bold=False, fmt=None, align="center"):
        cell.value = val
        cell.font  = Font(name="Arial", bold=bold, size=10)
        cell.fill  = _f(bg); cell.border = _b()
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fmt: cell.number_format = fmt
    def _cw(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- マップ種類の分類 ----
    # map_type文字列 → カテゴリ
    def _classify_map(map_type: str) -> str:
        code = map_type_code(map_type)
        if code in ("ndvi_absolute", "ndvi_average", "ndvi_field_average"):
            return "ndvi_abs"
        if code == "ndvi_relative":
            return "ndvi_rel"
        if code in ("ndvi_absolute_relative",):
            return "ndvi_abs"   # 絶対値側として扱う（相対も含む）
        if code in ("lai_absolute", "lai_absolute_relative"):
            return "lai_abs"
        if code == "lai_relative":
            return "lai_rel"
        if code in ("soil_map", "soil_map_lai"):
            return "soil"
        if code == "all_maps":
            # 実データのキーで判断
            return "all"
        return "other"

    MAP_CATEGORY_LABEL = {
        "ndvi_abs": ("NDVIマップ（絶対表示）", MID,  "ndvi",          "NDVI"),
        "ndvi_rel": ("NDVIマップ（相対表示）", SUB,  "ndvi_contrast", "NDVI相対"),
        "lai_abs":  ("生育マップ（絶対表示）", MID,  "lai",           "LAI"),
        "lai_rel":  ("生育マップ（相対表示）", SUB,  "lai_contrast",  "LAI相対"),
        "soil":     ("地力マップ",            SOIL, "lai",           "LAI（地力）"),
    }

    # ---- データ集約ヘルパー ----
    def _aggregate(target_rows, value_key):
        """field_uuid → date → {avg,max,min, zone_xxx, zone_count}"""
        by_field: Dict[str, Dict[date, Dict]] = defaultdict(lambda: defaultdict(dict))
        for row in target_rows:
            fid   = row["field_uuid"]
            dt    = row["observed_at"]
            zone  = str(row.get("zone_name") or row.get("zone_id") or "").strip()
            val   = row.get(value_key)
            entry = by_field[fid][dt]
            if val is not None:
                entry.setdefault("_vals", []).append(val)
                if zone:
                    # 正規化した代表名をキーに使用（括弧付きゾーンを統合）
                    norm_zone = _normalize_zone(zone)
                    # 同じ正規化キーに複数値がある場合は平均を取るため一時リストに追加
                    entry.setdefault(f"_z_{norm_zone}", []).append(val)
        for fid, dm in by_field.items():
            for dt, entry in dm.items():
                vals = entry.pop("_vals", [])
                entry["avg"]        = safe_avg(vals)
                entry["max"]        = safe_max(vals)
                entry["min"]        = safe_min(vals)
                entry["zone_count"] = len(set(
                    _normalize_zone(r.get("zone_name") or r.get("zone_id") or "")
                    for r in target_rows
                    if r["field_uuid"] == fid and r["observed_at"] == dt
                    and (r.get("zone_name") or r.get("zone_id"))
                ))
                # ゾーン別の一時リストを平均値に確定
                for key in list(entry.keys()):
                    if key.startswith("_z_"):
                        zone_vals = entry.pop(key)
                        real_key = "z_" + key[3:]
                        entry[real_key] = safe_avg(zone_vals)
        return by_field

    def _zone_sort_key(name: str):
        """ゾーン名から数値を抽出して自然順ソートキーを返す"""
        import re as _re
        m = _re.search(r'(\d+)', name)
        return (0, int(m.group(1)), name) if m else (1, 0, name)

    def _normalize_zone(name: str) -> str:
        """括弧付き修飾語を除去して代表名に正規化する
        例: 'ゾーン 4 (低い)' → 'ゾーン 4', 'ゾーン 1 (高い)' → 'ゾーン 1'"""
        import re as _re
        m = _re.match(r'^(ゾーン\s*\d+)\s*\(.*\)$', name.strip())
        return m.group(1) if m else name

    def _zone_names(target_rows):
        """重複を統合して自然順ソートしたゾーン名リストを返す"""
        raw = set(
            r.get("zone_name") or r.get("zone_id") or ""
            for r in target_rows if r.get("zone_name") or r.get("zone_id")
        )
        # 正規化した代表名でユニーク化
        seen = {}
        for zn in raw:
            key = _normalize_zone(zn)
            if key not in seen:
                seen[key] = key
        return sorted(seen.keys(), key=_zone_sort_key)

    # ---- セクションブロック書き込みヘルパー ----
    def _write_section(ws, start_row, sorted_dates, by_field, fid, zone_names,
                       value_label, header_bg, n_zones, title_width=None):
        """1つのマップ種類分のヘッダー+データを書く。最終行を返す。
        全データ行で値が空のゾーン列は出力しない。
        title_width: セクションタイトル行のマージ列数（省略時は実データ列数）"""

        # 事前に値が存在するゾーンだけ絞り込む
        date_map = by_field.get(fid, {})
        active_zones = [
            zn for zn in zone_names
            if any(
                entry.get(f"z_{_normalize_zone(zn)}") is not None
                for entry in date_map.values()
            )
        ]
        n_active = len(active_zones)

        # セクションタイトル行（title_widthが指定された場合はそちらに合わせる）
        n_cols = 4 + n_active
        merge_width = title_width if title_width and title_width > n_cols else n_cols
        last_c = get_column_letter(merge_width)
        ws.merge_cells(f"A{start_row}:{last_c}{start_row}")
        _h(ws[f"A{start_row}"], value_label, bg=header_bg, sz=11)
        ws.row_dimensions[start_row].height = 20
        r = start_row + 1

        # ヘッダー行1（セクション）
        ws.merge_cells(f"A{r}:A{r+1}")
        ws.merge_cells(f"B{r}:D{r}")
        if n_active > 0:
            ws.merge_cells(f"E{r}:{get_column_letter(4+n_active)}{r}")
        _h(ws[f"A{r}"], "日付")
        _h(ws[f"B{r}"], value_label, bg=header_bg)
        if n_active > 0:
            _h(ws.cell(r, 5), f"ゾーン別 {value_label}", bg=header_bg)
        ws.row_dimensions[r].height = 16
        r += 1

        # ヘッダー行2
        _h(ws.cell(r, 2), "平均", bg=header_bg)
        _h(ws.cell(r, 3), "最大", bg=header_bg)
        _h(ws.cell(r, 4), "最小", bg=header_bg)
        for zi, zn in enumerate(active_zones):
            _h(ws.cell(r, 5+zi), zn, bg=header_bg, sz=9)
        ws.row_dimensions[r].height = 14
        r += 1

        # データ行
        for i, dt in enumerate(sorted_dates):
            entry = date_map.get(dt)
            bg = ALT if i % 2 == 1 else WHITE
            _d(ws.cell(r, 1), datetime(dt.year, dt.month, dt.day), bg=bg, fmt="yyyy/mm/dd")
            _d(ws.cell(r, 2), entry.get("avg") if entry else None, bg=bg, fmt="0.000", bold=True)
            _d(ws.cell(r, 3), entry.get("max") if entry else None, bg=bg, fmt="0.000")
            _d(ws.cell(r, 4), entry.get("min") if entry else None, bg=bg, fmt="0.000")
            for zi, zn in enumerate(active_zones):
                _d(ws.cell(r, 5+zi),
                   entry.get(f"z_{_normalize_zone(zn)}") if entry else None, bg=bg, fmt="0.000")
            r += 1

        return r  # 次の空き行

    def _add_chart(ws, data_col, data_start_row, data_end_row,
                   cat_start_row, chart_anchor, title, header_bg):
        """平均（data_col）・最大（data_col+1）・最小（data_col+2）の3系列を描画"""
        chart = LineChart()
        chart.title  = title
        chart.style  = 10
        chart.y_axis.title        = "値"
        chart.y_axis.numFmt       = "0.000"
        chart.y_axis.tickLblPos   = "nextTo"   # Y軸ラベルを軸の隣に表示
        chart.y_axis.delete       = False       # Y軸ラベルを明示的に表示
        chart.x_axis.title        = "日付"
        chart.x_axis.numFmt       = "yyyy/mm/dd"
        chart.x_axis.majorTimeUnit = "days"
        chart.height = 12; chart.width = 22

        # 平均・最大・最小の3列を一括追加（ヘッダー行をタイトルとして使用）
        data_ref = Reference(ws, min_col=data_col, max_col=data_col + 2,
                             min_row=data_start_row, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1,
                             min_row=cat_start_row, max_row=data_end_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # 色・線スタイル設定
        # 平均: 太い実線
        chart.series[0].graphicalProperties.line.solidFill = header_bg
        chart.series[0].graphicalProperties.line.width = 22000
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size   = 5
        # 最大: 細い破線（同系色・薄め）
        if len(chart.series) > 1:
            chart.series[1].graphicalProperties.line.solidFill = "A8D5B5"
            chart.series[1].graphicalProperties.line.width = 12000
            chart.series[1].graphicalProperties.line.dashDot = "dash"
            chart.series[1].marker.symbol = "none"
        # 最小: 細い破線（同系色・薄め）
        if len(chart.series) > 2:
            chart.series[2].graphicalProperties.line.solidFill = "A8D5B5"
            chart.series[2].graphicalProperties.line.width = 12000
            chart.series[2].graphicalProperties.line.dashDot = "dash"
            chart.series[2].marker.symbol = "none"

        ws.add_chart(chart, chart_anchor)

    # ================================================================
    # まとめシート
    # ================================================================
    wb = Workbook()
    wb.remove(wb.active)
    sorted_fields = sorted(fields, key=lambda x: x["field_name"])

    has_ndvi     = any(r.get("ndvi")          is not None for r in rows)
    has_ndvi_rel = any(r.get("ndvi_contrast") is not None for r in rows)
    has_lai      = any(r.get("lai")           is not None for r in rows)
    has_lai_rel  = any(r.get("lai_contrast")  is not None for r in rows)
    has_soil     = any(
        map_type_code(r.get("map_type","")) in ("soil_map","soil_map_lai")
        for r in rows
    )

    all_dates = sorted(set(r["observed_at"] for r in rows))
    n_fields  = len(sorted_fields)

    # まとめシート列: 日付・農場・圃場 + 指標ごとの平均
    summary_metrics = []
    if has_ndvi:     summary_metrics.append(("ndvi",          "NDVI平均",      MID))
    if has_ndvi_rel: summary_metrics.append(("ndvi_contrast", "NDVI相対平均",  SUB))
    if has_lai:      summary_metrics.append(("lai",           "LAI平均",       MID))
    if has_lai_rel:  summary_metrics.append(("lai_contrast",  "LAI相対平均",   SUB))
    if has_soil:     summary_metrics.append(("lai",           "LAI地力平均",   SOIL))

    # 集約: 指標ごと
    agg_cache: Dict[str, Any] = {}
    for key in set(m[0] for m in summary_metrics):
        agg_cache[key] = _aggregate(rows, key)

    n_metrics = len(summary_metrics)
    last_col_sum = get_column_letter(3 + n_fields * n_metrics)

    ws_sum = wb.create_sheet(title="まとめ")
    ws_sum.merge_cells(f"A1:{last_col_sum}1")
    c = ws_sum["A1"]
    c.value = f"計測データまとめ　{from_date.isoformat()} ～ {till_date.isoformat()}"
    c.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill  = _f(DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 24

    # ヘッダー行2: 基本情報 + 圃場ブロック
    ws_sum.merge_cells("A2:C2")
    _h(ws_sum["A2"], "基本情報")
    for fi, field in enumerate(sorted_fields):
        col_s = 4 + fi * n_metrics
        col_e = col_s + n_metrics - 1
        ws_sum.merge_cells(
            f"{get_column_letter(col_s)}2:{get_column_letter(col_e)}2"
        )
        _h(ws_sum.cell(2, col_s), field["field_name"], bg=MID, sz=9)
    ws_sum.row_dimensions[2].height = 16

    # ヘッダー行3: 日付・農場・圃場 + 指標名
    _h(ws_sum.cell(3, 1), "日付")
    _h(ws_sum.cell(3, 2), "農場名")
    _h(ws_sum.cell(3, 3), "圃場名")
    for fi in range(n_fields):
        for mi, (_, mlabel, mbg) in enumerate(summary_metrics):
            _h(ws_sum.cell(3, 4 + fi*n_metrics + mi), mlabel, bg=mbg, sz=9)
    ws_sum.row_dimensions[3].height = 14

    # データ行
    for i, dt in enumerate(all_dates):
        r_idx = i + 4
        bg = ALT if i % 2 == 1 else WHITE
        _d(ws_sum.cell(r_idx, 1), datetime(dt.year, dt.month, dt.day), bg=bg, fmt="yyyy/mm/dd")
        _d(ws_sum.cell(r_idx, 2), "", bg=bg)
        _d(ws_sum.cell(r_idx, 3), "全圃場", bg=bg)
        for fi, field in enumerate(sorted_fields):
            for mi, (mkey, _, _) in enumerate(summary_metrics):
                val = agg_cache[mkey].get(field["field_uuid"], {}).get(dt, {}).get("avg")
                _d(ws_sum.cell(r_idx, 4 + fi*n_metrics + mi), val, bg=bg, fmt="0.000")

    ws_sum.column_dimensions["A"].width = 13
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 14
    for j in range(4, 4 + n_fields * n_metrics):
        ws_sum.column_dimensions[get_column_letter(j)].width = 10
    ws_sum.freeze_panes = "A4"

    # ================================================================
    # 圃場別シート（マップ種類ごとにセクション分割）
    # ================================================================
    for field in sorted_fields:
        fid   = field["field_uuid"]
        fname = field["field_name"]

        # この圃場に存在するマップ種類カテゴリを確認
        field_rows  = [r for r in rows if r["field_uuid"] == fid]
        cats_present = set(_classify_map(r.get("map_type","")) for r in field_rows)

        # カテゴリと順序を定義
        sections = []
        if "ndvi_abs" in cats_present or "all" in cats_present:
            if has_ndvi:
                sections.append(("ndvi_abs",  "ndvi",         "NDVIマップ（絶対表示）", MID))
        if "ndvi_rel" in cats_present or "all" in cats_present:
            if has_ndvi_rel:
                sections.append(("ndvi_rel",  "ndvi_contrast","NDVIマップ（相対表示）", SUB))
        if "lai_abs" in cats_present or "all" in cats_present:
            if has_lai:
                sections.append(("lai_abs",   "lai",          "生育マップ（絶対表示）", MID))
        if "lai_rel" in cats_present or "all" in cats_present:
            if has_lai_rel:
                sections.append(("lai_rel",   "lai_contrast", "生育マップ（相対表示）", SUB))
        if "soil" in cats_present:
            if has_soil or has_lai:
                sections.append(("soil",      "lai",          "地力マップ", SOIL))

        if not sections:
            # フォールバック: データにあるキーで判断
            if has_ndvi:     sections.append(("ndvi_abs",  "ndvi",          "NDVI（絶対）", MID))
            if has_ndvi_rel: sections.append(("ndvi_rel",  "ndvi_contrast", "NDVI（相対）", SUB))
            if has_lai:      sections.append(("lai_abs",   "lai",           "LAI（絶対）",  MID))
            if has_lai_rel:  sections.append(("lai_rel",   "lai_contrast",  "LAI（相対）",  SUB))

        # 圃場のゾーン名
        zones = _zone_names(field_rows)
        n_zones = len(zones)
        n_cols  = max(4 + n_zones, 5)

        # 全日付
        field_dates = sorted(set(r["observed_at"] for r in field_rows))

        sheet_name = sanitize_sheet_name(fname)
        ws = wb.create_sheet(title=sheet_name)

        # シートタイトル
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        c = ws["A1"]
        c.value = f"計測データ　{fname}"
        c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill  = _f(DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        _cw(ws, [13, 9, 9, 9] + [9]*n_zones)

        current_row = 2
        chart_positions = []  # (anchor, title, section_bg, data_col, data_start, data_end, cat_start)

        for cat, value_key, section_label, section_bg in sections:
            # このカテゴリのデータのみ絞り込み
            if cat == "soil":
                sec_rows = [r for r in field_rows
                            if map_type_code(r.get("map_type","")) in ("soil_map","soil_map_lai")]
                if not sec_rows:
                    sec_rows = field_rows  # フォールバック
            elif cat == "ndvi_abs":
                sec_rows = [r for r in field_rows
                            if map_type_code(r.get("map_type","")) in
                            ("ndvi_absolute","ndvi_absolute_relative","ndvi_average","ndvi_field_average","all_maps")
                            and r.get("ndvi") is not None]
                if not sec_rows: sec_rows = [r for r in field_rows if r.get("ndvi") is not None]
            elif cat == "ndvi_rel":
                sec_rows = [r for r in field_rows if r.get("ndvi_contrast") is not None]
            elif cat == "lai_abs":
                sec_rows = [r for r in field_rows
                            if map_type_code(r.get("map_type","")) in
                            ("lai_absolute","lai_absolute_relative","all_maps")
                            and r.get("lai") is not None]
                if not sec_rows: sec_rows = [r for r in field_rows if r.get("lai") is not None]
            elif cat == "lai_rel":
                sec_rows = [r for r in field_rows if r.get("lai_contrast") is not None]
            else:
                sec_rows = field_rows

            sec_agg   = _aggregate(sec_rows, value_key)
            sec_dates = sorted(set(r["observed_at"] for r in sec_rows)) or field_dates

            data_header_row = current_row + 2   # セクション行+ヘッダー2行後
            data_start_row  = data_header_row + 1
            data_end_row    = data_start_row + len(sec_dates) - 1

            next_row = _write_section(
                ws, current_row, sec_dates, sec_agg, fid,
                zones, section_label, section_bg, n_zones
            )
            # チャート用に位置記録
            chart_positions.append((
                section_label, section_bg,
                2,                  # 平均値列(B=2)
                data_header_row,    # タイトル行（Referenceのmin_row）
                data_start_row,     # データ開始行（categories用）
                data_end_row,
            ))

            current_row = next_row + 1  # 1行空ける

        # チャートを全セクション書き終わった後に追加
        chart_row = current_row + 1
        for i, (sec_label, sec_bg, dcol, d_min_row, cat_start, d_end_row) in enumerate(chart_positions):
            col_offset = i * 12  # チャート横並び（列12個分ずつずらす）
            anchor_col = get_column_letter(n_cols + 2 + col_offset)
            _add_chart(ws, dcol, d_min_row, d_end_row, cat_start,
                       f"{anchor_col}2", sec_label, sec_bg)

        ws.freeze_panes = "A2"

    # ================================================================
    # 地力マップ専用シート（全圃場まとめ）
    # ================================================================
    soil_rows = [r for r in rows
                 if map_type_code(r.get("map_type","")) in ("soil_map","soil_map_lai")
                 or r.get("fdl_layer_type") == "BIOMASS_MULTI_IMAGE_LAI"]
    # 地力マップシートの圃場数チェック（デバッグ用）
    soil_field_ids = set(r["field_uuid"] for r in soil_rows)
    all_field_ids  = set(r["field_uuid"] for r in rows)
    missing = all_field_ids - soil_field_ids
    if missing:
        missing_names = set(r["field_name"] for r in rows if r["field_uuid"] in missing)
        # 地力マップデータなし圃場のmap_typeを確認
        for r in rows:
            if r["field_uuid"] in missing:
                mt = r.get("map_type","")
                lt = r.get("fdl_layer_type","")
                if mt or lt:
                    append_log(f"地力マップ除外: {r['field_name']} map_type={mt!r} fdl_layer_type={lt!r}")
                    break
    if soil_rows:
        ws_soil = wb.create_sheet(title="地力マップ（全圃場）")
        soil_dates = sorted(set(r["observed_at"] for r in soil_rows))
        soil_zones = _zone_names(soil_rows)
        n_sz   = len(soil_zones)
        n_scols = max(4 + n_sz, 5)

        ws_soil.merge_cells(f"A1:{get_column_letter(n_scols)}1")
        c = ws_soil["A1"]
        c.value = f"地力マップ（全圃場）　{from_date.isoformat()} ～ {till_date.isoformat()}"
        c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill  = _f(SOIL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_soil.row_dimensions[1].height = 22

        # 圃場ごとにセクション
        current_row = 2
        chart_pos_soil = []
        for field in sorted_fields:
            fid_s = field["field_uuid"]
            f_soil_rows = [r for r in soil_rows if r["field_uuid"] == fid_s]
            if not f_soil_rows:
                continue
            f_soil_agg   = _aggregate(f_soil_rows, "lai")
            f_soil_dates = sorted(set(r["observed_at"] for r in f_soil_rows))
            dhr = current_row + 2
            dsr = dhr + 1
            der = dsr + len(f_soil_dates) - 1
            next_row = _write_section(
                ws_soil, current_row, f_soil_dates, f_soil_agg, fid_s,
                soil_zones, f"地力マップ　{field['field_name']}", SOIL, n_sz,
                title_width=n_scols,
            )
            chart_pos_soil.append((field["field_name"], dhr, dsr, der))
            current_row = next_row + 1

        for i, (fname_s, dhr, dsr, der) in enumerate(chart_pos_soil):
            col_offset = i * 12
            anchor_col = get_column_letter(n_scols + 2 + col_offset)
            _add_chart(ws_soil, 2, dhr, der, dsr, f"{anchor_col}2",
                       f"地力　{fname_s}", SOIL)

        _cw(ws_soil, [13, 9, 9, 9] + [9]*n_sz)
        ws_soil.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def ensure_session_defaults() -> None:
    st.session_state.setdefault("login_token", "")
    st.session_state.setdefault("api_token", "")
    st.session_state.setdefault("farms", [])
    st.session_state.setdefault("fields", [])
    st.session_state.setdefault("metric_rows", [])
    st.session_state.setdefault("fetch_logs", [])
    st.session_state.setdefault("fdl_uuid_text", "")
    st.session_state.setdefault("pending_fdl_uuid_text", None)
    st.session_state.setdefault("fdl_uuid_field_map", {})
    st.session_state.setdefault("auto_fields_loaded_key", "")
    st.session_state.setdefault("auto_fdl_loaded_key", "")
    st.session_state.setdefault("auto_metrics_loaded_key", "")
    st.session_state.setdefault("login_email", "")
    st.session_state.setdefault("last_query_from_date", None)
    st.session_state.setdefault("last_query_till_date", None)
    st.session_state.setdefault("last_query_map_type", "")
    st.session_state.setdefault("last_selected_field_labels", [])
    st.session_state.setdefault("input_from_date", None)
    st.session_state.setdefault("input_till_date", None)
    st.session_state.setdefault("input_map_type", "全マップ（LAI/NDVI すべて）")
    st.session_state.setdefault("max_fields", 10)
    st.session_state.setdefault("pending_field_multiselect", None)
    st.session_state.setdefault("field_over_limit", False)


def append_log(message: str) -> None:
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    st.session_state.fetch_logs.append(f"[{now_text}] {message}")


def clear_result_state() -> None:
    st.session_state.metric_rows = []
    st.session_state.fetch_logs = []
    st.session_state.last_query_from_date = None
    st.session_state.last_query_till_date = None
    st.session_state.last_query_map_type = ""
    st.session_state.last_selected_field_labels = []


def reset_session_after_logout() -> None:
    st.session_state.login_token = ""
    st.session_state.api_token = ""
    st.session_state.farms = []
    st.session_state.fields = []
    st.session_state.metric_rows = []
    st.session_state.fetch_logs = []
    st.session_state.fdl_uuid_text = ""
    st.session_state.pending_fdl_uuid_text = None
    st.session_state.fdl_uuid_field_map = {}
    st.session_state.auto_fields_loaded_key = ""
    st.session_state.auto_fdl_loaded_key = ""
    st.session_state.auto_metrics_loaded_key = ""
    st.session_state.last_query_from_date = None
    st.session_state.last_query_till_date = None
    st.session_state.last_query_map_type = ""
    st.session_state.last_selected_field_labels = []


def fetch_farms_into_session(graphql_url: str) -> List[Dict[str, str]]:
    farms_out = call_xarvio_graphql(
        graphql_url,
        st.session_state.login_token,
        st.session_state.api_token,
        "FarmsOverview",
        FARMS_OVERVIEW,
        {},
    )
    farms = parse_farms(farms_out)
    st.session_state.farms = farms
    append_log(f"farm一覧取得: {len(farms)}件")
    return farms


def render_ui_theme() -> None:
    st.markdown(
        """
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@500;700;800&family=Noto+Sans+JP:wght@400;500;700&display=swap');

/* ================================================================
   ベース: フォント・メイン背景
   ================================================================ */
html, body, [class*="css"] {
    font-family: "Manrope", "Noto Sans JP", sans-serif;
}

/* メイン背景 */
[data-testid="stAppViewContainer"] {
    background:
      radial-gradient(1200px 600px at 90% -20%, rgba(38,166,154,0.18), transparent 60%),
      radial-gradient(900px 500px at -10% -20%, rgba(255,183,77,0.16), transparent 60%),
      linear-gradient(180deg, #f7faf8 0%, #f4f7f5 100%);
}

/* ================================================================
   メインエリア: 全テキストを黒系に固定
   ================================================================ */
[data-testid="stMain"] {
    color: #1a1a1a !important;
}
[data-testid="stMain"] h1,
[data-testid="stMain"] h2,
[data-testid="stMain"] h3,
[data-testid="stMain"] h4,
[data-testid="stMain"] p,
[data-testid="stMain"] span,
[data-testid="stMain"] label,
[data-testid="stMain"] div {
    color: #1a1a1a !important;
}

/* メイン: セレクトボックス・マルチセレクトの入力エリア背景と文字 */
[data-testid="stMain"] [data-baseweb="select"] > div:first-child {
    background-color: #1f6b3c !important;
    border-color: #4caf7d !important;
}
[data-testid="stMain"] [data-baseweb="select"] span,
[data-testid="stMain"] [data-baseweb="select"] [class*="placeholder"],
[data-testid="stMain"] [data-baseweb="select"] input {
    color: #ffffff !important;
}
/* プレースホルダー文字（"選択してください" / "No options to select." など） */
[data-testid="stMain"] [data-baseweb="select"] [aria-disabled="true"] span,
[data-testid="stMain"] [data-baseweb="select"] [class*="Placeholder"] {
    color: #c0e8d8 !important;
}
/* セレクトボックスの矢印アイコン */
[data-testid="stMain"] [data-baseweb="select"] svg {
    fill: #ffffff !important;
}

/* メイン: マルチセレクトのタグ */
[data-testid="stMain"] [data-baseweb="tag"] {
    background-color: #d0ede5 !important;
}
[data-testid="stMain"] [data-baseweb="tag"] span {
    color: #1a1a1a !important;
}

/* メイン: テキスト入力・テキストエリア */
[data-testid="stMain"] input,
[data-testid="stMain"] textarea {
    color: #1a1a1a !important;
    background-color: #ffffff !important;
}

/* メイン: セカンダリボタン（"入力をクリア" など） */
[data-testid="stMain"] [data-testid="stBaseButton-secondary"],
[data-testid="stMain"] button[kind="secondary"] {
    background-color: #e8f4f1 !important;
    border: 1px solid #2e8b57 !important;
    color: #0f2f2b !important;
}
[data-testid="stMain"] [data-testid="stBaseButton-secondary"] p,
[data-testid="stMain"] [data-testid="stBaseButton-secondary"] span {
    color: #0f2f2b !important;
}

/* メイン: caption */
[data-testid="stMain"] [data-testid="stCaptionContainer"] p,
[data-testid="stMain"] [data-testid="stCaptionContainer"] span {
    color: #3a5a50 !important;
}

/* メイン: metric */
[data-testid="stMetricLabel"] p,
[data-testid="stMetricLabel"] span {
    color: #3a5a50 !important;
}
[data-testid="stMetricValue"] {
    color: #0b6158 !important;
}

/* メイン: date_input */
[data-testid="stMain"] [data-testid="stDateInput"] input {
    color: #1a1a1a !important;
    background-color: #ffffff !important;
}

/* ================================================================
   サイドバー: 濃い緑背景 → 文字はすべて明るく
   ================================================================ */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0f2f2b 0%, #133a35 100%) !important;
}

/* サイドバー内の全テキストをデフォルト明るく */
[data-testid="stSidebar"],
[data-testid="stSidebar"] * {
    color: #e8f4f1 !important;
}

/* サイドバー: タイトル・見出し */
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] [data-testid="stHeadingWithActionElements"] {
    color: #e8f4f1 !important;
}

/* サイドバー: テキスト入力（メール・パスワード） */
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] textarea {
    color: #1a1a1a !important;
    background-color: #e8f4f1 !important;
    border-color: #4caf7d !important;
}
[data-testid="stSidebar"] input::placeholder {
    color: #6a9a92 !important;
}

/* サイドバー: number_input */
[data-testid="stSidebar"] [data-testid="stNumberInput"] input {
    color: #1a1a1a !important;
    background-color: #e8f4f1 !important;
}

/* サイドバー: エクスパンダー */
[data-testid="stSidebar"] [data-testid="stExpander"] details summary p,
[data-testid="stSidebar"] [data-testid="stExpander"] details summary span,
[data-testid="stSidebar"] [data-testid="stExpander"] summary {
    color: #e8f4f1 !important;
}

/* サイドバー: primaryボタン（接続する） */
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"],
[data-testid="stSidebar"] button[kind="primary"] {
    background-color: #2e8b57 !important;
    border-color: #4caf7d !important;
}
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] p,
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] span {
    color: #ffffff !important;
}

/* サイドバー: secondaryボタン（ログアウト・処理ログ） */
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"],
[data-testid="stSidebar"] button[kind="secondary"] {
    background-color: #c8e6da !important;
    border: 1px solid #4caf7d !important;
}
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] p,
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] span {
    color: #0f2f2b !important;
}

/* サイドバー: アラート・通知バナー内文字 */
[data-testid="stSidebar"] [data-testid="stAlert"] p,
[data-testid="stSidebar"] [data-testid="stAlert"] span,
[data-testid="stSidebar"] [role="alert"] p,
[data-testid="stSidebar"] [role="alert"] span {
    color: #1a1a1a !important;
}

/* サイドバー: success バナー（接続済み） */
[data-testid="stSidebar"] [data-testid="stNotification"] p,
[data-testid="stSidebar"] [data-testid="stNotification"] span {
    color: #1a1a1a !important;
}

/* サイドバー: divider */
[data-testid="stSidebar"] hr {
    border-color: rgba(232,244,241,0.3) !important;
}

/* ================================================================
   ドロップダウンメニュー（ポップアップ、全体共通）
   ================================================================ */
[data-baseweb="popover"] ul li,
[data-baseweb="popover"] [role="option"],
[data-baseweb="menu"] [role="option"] {
    color: #1a1a1a !important;
    background-color: #ffffff !important;
}
[data-baseweb="popover"] ul li:hover,
[data-baseweb="popover"] [role="option"]:hover {
    background-color: #d0ede5 !important;
}

/* ================================================================
   DataFrameスタイル
   ================================================================ */
[data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
    border: 1px solid rgba(19,74,68,0.15);
}
</style>
        """,
        unsafe_allow_html=True,
    )


# ---- アプリ初期化 ----
st.set_page_config(page_title="xarvio Data Report", layout="wide", page_icon="🌿")
render_ui_theme()
ensure_session_defaults()

if st.session_state.pending_fdl_uuid_text is not None:
    st.session_state.fdl_uuid_text = st.session_state.pending_fdl_uuid_text
    st.session_state.pending_fdl_uuid_text = None

if st.session_state.pending_field_multiselect is not None:
    st.session_state.field_multiselect = st.session_state.pending_field_multiselect
    st.session_state.pending_field_multiselect = None

gigya_base = DEFAULT_GIGYA_BASE
gigya_api_key = DEFAULT_GIGYA_API_KEY
token_url = DEFAULT_TOKEN_URL
graphql_url = DEFAULT_GRAPHQL_URL
FDL_LOCALE = "JA_JP"

# ---- サイドバー ----
with st.sidebar:
    st.title("設定")

    is_logged_in = bool(st.session_state.login_token and st.session_state.api_token)
    with st.expander("アカウント接続", expanded=not is_logged_in):
        email = st.text_input("メールアドレス", value=st.session_state.get("login_email", ""), key="login_email_input")
        password = st.text_input("パスワード", value="", type="password", key="login_password_input")

        if st.button("接続する", use_container_width=True, type="primary"):
            if not email or not password:
                st.error("Email と Password を入力してください。")
            else:
                try:
                    login_data = gigya_login(gigya_base, gigya_api_key, email, password)
                    st.session_state.login_token = login_data["login_token"]
                    st.session_state.api_token = issue_xarvio_token(token_url, login_data)
                    st.session_state.login_email = email
                    clear_result_state()
                    st.session_state.farms = []
                    st.session_state.fields = []
                    st.session_state.pending_fdl_uuid_text = ""
                    st.session_state.fdl_uuid_field_map = {}
                    fetch_farms_into_session(graphql_url)
                    st.rerun()
                except Exception as exc:
                    st.error(f"接続失敗: {exc}")
                    st.caption(f"gigya_base={gigya_base} / api_key_prefix={gigya_api_key[:8] if gigya_api_key else 'EMPTY'}")

    if is_logged_in:
        st.success("✅ xarvio API 接続済み")
        if st.button("ログアウト / 切断", use_container_width=True):
            reset_session_after_logout()
            st.rerun()

    st.divider()

    with st.expander("設定 / デバッグ", expanded=False):
        st.number_input(
            "圃場選択の上限数",
            min_value=1,
            max_value=50,
            value=st.session_state.max_fields,
            step=1,
            key="max_fields",
            help="一度に分析できる圃場数の上限。多すぎると処理に時間がかかります。",
        )
        st.text_area(
            "キャッシュ済み FDL UUID一覧",
            value=st.session_state.fdl_uuid_text,
            key="fdl_uuid_text_widget",
            height=100,
            help="自動取得されたUUIDがここに表示されます。手動で書き換えることも可能です。",
        )
        if st.button("処理ログを表示"):
            st.code("\n".join(st.session_state.fetch_logs[-200:]))

# ---- メインコンテンツ ----
st.markdown('## <span style="color:#000000;">🌿 xarvio Data Report</span>', unsafe_allow_html=True)

if not is_logged_in:
    st.markdown('<div style="background-color:#d9eaf7;border-radius:6px;padding:12px 16px;color:#000000;">👈 サイドバーから xarvio アカウントに接続してください。</div>', unsafe_allow_html=True)
    st.stop()

if not st.session_state.farms:
    try:
        fetch_farms_into_session(graphql_url)
    except Exception as exc:
        st.error(f"farm一覧の自動取得失敗: {exc}")

with st.container():
    col1, col2 = st.columns([1, 1], gap="medium")

    with col1:
        st.subheader("1. 対象選択")
        farms = st.session_state.farms
        placeholder_farm = {"name": "(選択してください)", "uuid": None}
        selected_farm = st.selectbox(
            "農場 (Farm)",
            options=[placeholder_farm] + farms,
            format_func=lambda farm: farm["name"],
        )

        selected_farm_uuids: List[str] = []
        if selected_farm and selected_farm["uuid"]:
            selected_farm_uuids = [selected_farm["uuid"]]

        selected_farm_key = "|".join(sorted(selected_farm_uuids))
        if selected_farm_uuids and selected_farm_key != st.session_state.auto_fields_loaded_key:
            try:
                with st.spinner("圃場リストを取得中..."):
                    all_fields: List[Dict[str, str]] = []
                    for chunk in chunked(selected_farm_uuids, SCAN_CHUNK_SIZE):
                        fields_out = call_xarvio_graphql(
                            graphql_url,
                            st.session_state.login_token,
                            st.session_state.api_token,
                            "FieldsByFarms",
                            FIELDS_BY_FARMS,
                            {"farmUuids": chunk},
                        )
                        all_fields.extend(parse_fields(fields_out))

                    all_fields.sort(key=lambda x: (x["farm_name"], x["field_name"], x["field_uuid"]))
                    st.session_state.fields = all_fields
                    st.session_state.auto_fields_loaded_key = selected_farm_key
                    st.session_state.metric_rows = []
                    st.session_state.pending_fdl_uuid_text = ""
                    st.session_state.fdl_uuid_field_map = {}
                    st.session_state.last_selected_field_labels = []
            except Exception as exc:
                st.error(f"Field取得エラー: {exc}")

        fields = st.session_state.fields
        max_fields = st.session_state.max_fields

        default_selected_fields = fields if 0 < len(fields) <= min(5, max_fields) else []
        selected_fields = st.multiselect(
            f"圃場 (Fields)　最大 {max_fields} 圃場",
            options=fields,
            default=default_selected_fields,
            format_func=lambda f: f["field_name"],
            placeholder="分析する圃場を選択 (複数可)",
            key="field_multiselect",
        )
        # 上限チェック
        if len(selected_fields) > max_fields:
            # 上限超え: pendingに切り詰め済みリストをセット、フラグを立てて次サイクルで反映
            st.session_state.pending_field_multiselect = selected_fields[:max_fields]
            st.session_state.field_over_limit = True
            st.rerun()
        else:
            # 通常状態: フラグリセット
            if st.session_state.field_over_limit:
                # 直前のサイクルで上限超えがあった → エラーを表示してフラグを消す
                st.error(f"⛔ 選択できる圃場は最大 {max_fields} 圃場です。{max_fields + 1} 件目以降は追加できません。")
                st.session_state.field_over_limit = False
            elif len(selected_fields) == max_fields:
                st.caption(f"✅ {max_fields} 圃場選択済み（上限）")

    with col2:
        st.subheader("2. 分析条件")
        jst_today = datetime.now(JST).date()
        if st.session_state.input_from_date is None:
            st.session_state.input_from_date = jst_today - timedelta(days=30)
        if st.session_state.input_till_date is None:
            st.session_state.input_till_date = jst_today

        d_col1, d_col2 = st.columns(2)
        from_date = d_col1.date_input(
            "開始日",
            value=st.session_state.input_from_date,
            key="input_from_date",
        )
        till_date = d_col2.date_input(
            "終了日",
            value=st.session_state.input_till_date,
            key="input_till_date",
        )

        map_type_options = [
            "全マップ（LAI/NDVI すべて）",
            "生育マップ（LAI + LAI_CONTRAST）",
            "地力マップ（LAIのみ）",
            "植生マップ（NDVI + NDVI_CONTRAST）",
            "植生マップ（NDVIのみ）",
        ]
        fdl_map_type = st.selectbox(
            "マップ種類",
            options=map_type_options,
            index=map_type_options.index(st.session_state.input_map_type)
                  if st.session_state.input_map_type in map_type_options else 0,
            key="input_map_type",
            help="NDVIは植生、LAIは葉面積指数を表します。",
        )

    st.divider()
    run_col1, run_col2 = st.columns([2, 1])
    with run_col1:
        start_analysis = st.button(
            "レポートを作成する",
            type="primary",
            use_container_width=True,
        )
    with run_col2:
        if st.button("入力をクリア", use_container_width=True):
            clear_result_state()
            st.rerun()

# ---- 分析実行 ----
if start_analysis:
    if not selected_fields:
        st.warning("⚠️ 対象の圃場を選択してください。")
    elif from_date > till_date:
        st.error("⚠️ 開始日は終了日より前に設定してください。")
    else:
        target_fields = selected_fields

        # Step 1: FDL UUID の収集
        try:
            with st.spinner("1/2: 地図データ(FDL)を検索中..."):
                targets = target_magnitudes_all(fdl_map_type)
                found_with_field: List[Tuple[str, Dict[str, str], Optional[date], Optional[str], str]] = []

                for field in target_fields:
                    variables = {"fieldUuid": field["field_uuid"]}
                    try:
                        analytics_out = call_xarvio_graphql(
                            graphql_url,
                            st.session_state.login_token,
                            st.session_state.api_token,
                            "AnalyticsMap",
                            ANALYTICS_MAP_QUERY,
                            variables,
                        )
                        uuid_date_type_triples, observed_types = extract_fdl_uuids_from_analytics_map(analytics_out, targets)
                        if observed_types:
                            append_log(f"{field['field_name']} observed magnitudes: {', '.join(observed_types)}")
                        for uuid, layer_date, layer_date_raw, layer_type in uuid_date_type_triples:
                            found_with_field.append((uuid, field, layer_date, layer_date_raw, layer_type))
                    except Exception as exc:
                        append_log(f"FDL search failed for {field['field_name']}: {exc}")

            # 期間フィルタ: layer_date が None のものは警告ログを出してスキップ (ToDo: layer_date フォールバック除去)
            filtered_found_with_field = []
            for item in found_with_field:
                uuid, field, layer_date, layer_date_raw, layer_type = item
                if layer_date is None:
                    append_log(f"WARNING: layer_date が取得できなかったため {field['field_name']} の FDL {uuid[:8]}... をスキップします")
                    continue
                if from_date <= layer_date <= till_date:
                    filtered_found_with_field.append(item)

            if not filtered_found_with_field:
                st.warning(f"⚠️ 指定された期間（{from_date} ~ {till_date}）とマップ種類に該当するデータが見つかりませんでした。")
                st.stop()

            deduped = sorted(list(set(item[0] for item in filtered_found_with_field)))
            mapped: Dict[str, Tuple[Dict[str, str], Optional[date], Optional[str], str]] = {}
            for uuid, field, layer_date, layer_date_raw, layer_type in filtered_found_with_field:
                if uuid not in mapped:
                    mapped[uuid] = (field, layer_date, layer_date_raw, layer_type)
            st.session_state.fdl_uuid_field_map = mapped
            st.session_state.pending_fdl_uuid_text = "\n".join(deduped)

        except Exception as exc:
            st.error(f"準備中にエラーが発生しました: {exc}")
            st.stop()

        rows: List[Dict[str, Any]] = []
        errors: List[str] = []

        # Step 2: FDL 分類データ取得 (FDLモード固定)
        # fdl_uuid_textはpendingのため次サイクルでしか反映されない。
        # 同サイクルで確定済みの fdl_uuid_field_map から直接UUIDリストを取得する。
        fdl_uuid_field_map_current = st.session_state.get("fdl_uuid_field_map", {})
        fdl_uuids_current = list(fdl_uuid_field_map_current.keys())

        with st.spinner("2/2: 指標データ(NDVI/LAI)を取得中..."):
            if CONCURRENT_FETCH_ENABLED:
                async def main() -> List[Any]:
                    CONCURRENCY_LIMIT = 10
                    sem = asyncio.Semaphore(CONCURRENCY_LIMIT)

                    async def limited_graphql_call(*args, **kwargs):
                        async with sem:
                            return await async_call_xarvio_graphql(*args, **kwargs)

                    async with httpx.AsyncClient() as client:
                        tasks = []
                        fdl_uuids = fdl_uuids_current
                        fdl_uuid_field_map = fdl_uuid_field_map_current
                        target_magnitudes = target_magnitudes_all(fdl_map_type)

                        for fdl_uuid in fdl_uuids:
                            field_info = fdl_uuid_field_map.get(fdl_uuid)
                            if not field_info:
                                continue
                            field, layer_date, layer_date_raw, layer_type = field_info
                            base_vars = {"fieldDataLayerUuid": fdl_uuid, "elevation": [], "locale": FDL_LOCALE}

                            for magnitude in target_magnitudes:
                                variables = {**base_vars, "magnitude": magnitude}
                                context = {
                                    "fdl_uuid": fdl_uuid,
                                    "magnitude": magnitude,
                                    "field": field,
                                    "layer_date": layer_date,
                                    "layer_date_raw": layer_date_raw,
                                    "layer_type": layer_type,
                                }
                                tasks.append(
                                    limited_graphql_call(
                                        client,
                                        graphql_url,
                                        st.session_state.login_token,
                                        st.session_state.api_token,
                                        "UseGetFieldDataLayerClassification",
                                        FDL_CLASSIFICATION_QUERY,
                                        variables,
                                        context,
                                    )
                                )

                        return await asyncio.gather(*tasks, return_exceptions=True)

                try:
                    results = asyncio.run(main())
                except RuntimeError:
                    loop = asyncio.new_event_loop()
                    try:
                        asyncio.set_event_loop(loop)
                        results = loop.run_until_complete(main())
                    finally:
                        loop.close()
                        asyncio.set_event_loop(None)

                fdl_uuid_field_map = st.session_state.get("fdl_uuid_field_map", {})
                outputs_by_fdl: Dict[str, Dict[str, Any]] = defaultdict(
                    lambda: {"outputs": {}, "field": None, "layer_date": None, "layer_date_raw": None, "layer_type": ""}
                )

                for res in results:
                    if isinstance(res, Exception):
                        errors.append(str(res))
                        continue
                    classification_out, context = res
                    fdl_uuid = context["fdl_uuid"]
                    magnitude = context["magnitude"]
                    outputs_by_fdl[fdl_uuid]["outputs"][magnitude] = classification_out

                    if not outputs_by_fdl[fdl_uuid]["field"]:
                        outputs_by_fdl[fdl_uuid]["field"] = context["field"]
                    if not outputs_by_fdl[fdl_uuid]["layer_date"]:
                        outputs_by_fdl[fdl_uuid]["layer_date"] = context["layer_date"]
                    if not outputs_by_fdl[fdl_uuid]["layer_date_raw"]:
                        outputs_by_fdl[fdl_uuid]["layer_date_raw"] = context["layer_date_raw"]
                    if not outputs_by_fdl[fdl_uuid]["layer_type"]:
                        outputs_by_fdl[fdl_uuid]["layer_type"] = context.get("layer_type", "")

                for fdl_uuid, data in outputs_by_fdl.items():
                    field = data.get("field")
                    layer_date = data.get("layer_date")
                    field_info = fdl_uuid_field_map.get(fdl_uuid)

                    if field_info and len(field_info) > 2:
                        layer_date_raw = field_info[2]
                        layer_type = field_info[3]
                    else:
                        layer_date_raw = data.get("layer_date_raw")
                        layer_type = data.get("layer_type", "")

                    if not field or not layer_date:
                        if not layer_date:
                            append_log(f"WARNING: layer_date なしのため {fdl_uuid[:8]}... をスキップ")
                        continue

                    field_rows = extract_fdl_rows(
                        data["outputs"],
                        fdl_uuid,
                        layer_date,
                        layer_date_raw,
                        field["field_uuid"],
                        field["field_name"],
                        field["farm_uuid"],
                        field["farm_name"],
                        fdl_map_type,
                        layer_type=layer_type,
                    )
                    rows.extend(field_rows)

            else:
                # 同期フォールバック (httpx 未インストール時)
                fdl_uuids = fdl_uuids_current
                fdl_uuid_field_map = fdl_uuid_field_map_current
                target_magnitudes = target_magnitudes_all(fdl_map_type)

                for fdl_uuid in fdl_uuids:
                    if fdl_uuid not in fdl_uuid_field_map:
                        continue
                    field, layer_date, layer_date_raw, layer_type = fdl_uuid_field_map[fdl_uuid]

                    if not layer_date:
                        append_log(f"WARNING: layer_date なしのため {fdl_uuid[:8]}... をスキップ")
                        continue

                    base_vars = {"fieldDataLayerUuid": fdl_uuid, "elevation": [], "locale": FDL_LOCALE}

                    try:
                        classification_outputs = {}
                        for magnitude in target_magnitudes:
                            out = call_xarvio_graphql(
                                graphql_url,
                                st.session_state.login_token,
                                st.session_state.api_token,
                                "UseGetFieldDataLayerClassification",
                                FDL_CLASSIFICATION_QUERY,
                                {**base_vars, "magnitude": magnitude},
                            )
                            classification_outputs[magnitude] = out

                        field_rows = extract_fdl_rows(
                            classification_outputs,
                            fdl_uuid,
                            layer_date,
                            layer_date_raw,
                            field["field_uuid"],
                            field["field_name"],
                            field["farm_uuid"],
                            field["farm_name"],
                            fdl_map_type,
                            layer_type=layer_type,
                        )
                        rows.extend(field_rows)
                    except Exception as exc:
                        errors.append(f"{field['field_name']}: {exc}")

        st.session_state.metric_rows = sorted(
            rows,
            key=lambda x: (x["farm_name"], x["field_name"], x["observed_at"], x["zone_name"], x["zone_id"]),
        )
        st.session_state.last_query_from_date = from_date
        st.session_state.last_query_till_date = till_date
        st.session_state.last_query_map_type = fdl_map_type
        st.session_state.last_selected_field_labels = selected_fields

        if st.session_state.metric_rows:
            st.success(f"完了: {len(st.session_state.metric_rows)} 件のデータを取得しました。")
        elif not errors:
            st.warning("データが見つかりませんでした。期間や条件を変更して再度お試しください。")

        if errors:
            st.error(f"{len(errors)} 件のエラーが発生しました。")
            for err in errors:
                append_log(err)

# ---- レポート表示 ----
rows = st.session_state.metric_rows
if rows:
    st.divider()
    st.subheader("3. レポート結果")

    kpi_col1, kpi_col2, kpi_col3 = st.columns(3)
    kpi_col1.metric("取得データ数", f"{len(rows)} 行")

    unique_fields = len(set(r["field_uuid"] for r in rows))
    kpi_col2.metric("対象圃場数", f"{unique_fields} 圃場")

    latest_date = max((r["observed_at"] for r in rows), default=None)
    kpi_col3.metric("最新データ日付", latest_date.isoformat() if latest_date else "-")

    # ---- チャート ----
    st.subheader("📊 推移チャート")
    unique_fields_in_results = sorted(
        list(set((r["field_name"], r["field_uuid"]) for r in rows)), key=lambda x: x[0]
    )
    field_display_map = {f"{name} | {uuid[:8]}": uuid for name, uuid in unique_fields_in_results}

    chart_field_label = "全圃場の平均"
    if len(unique_fields_in_results) > 1:
        chart_field_label = st.selectbox(
            "表示する圃場を選択",
            options=["全圃場の平均"] + list(field_display_map.keys()),
        )
    elif field_display_map:
        chart_field_label = list(field_display_map.keys())[0]

    if chart_field_label == "全圃場の平均":
        chart_rows = rows
        st.caption("全圃場の平均値の推移")
    else:
        selected_field_uuid = field_display_map[chart_field_label]
        chart_rows = [r for r in rows if r["field_uuid"] == selected_field_uuid]
        st.caption(f"{chart_field_label} の推移")

    aggregated_by_day: Dict[str, Dict[str, List[float]]] = defaultdict(
        lambda: {"ndvi": [], "lai": [], "ndvi_contrast": [], "lai_contrast": []}
    )
    for row in chart_rows:
        jst_date_str = to_jst_date_string(row.get("observed_at_raw"))
        if not jst_date_str:
            if row.get("observed_at"):
                jst_date_str = row["observed_at"].isoformat()
            else:
                continue
        if row.get("ndvi") is not None:
            aggregated_by_day[jst_date_str]["ndvi"].append(row["ndvi"])
        if row.get("lai") is not None:
            aggregated_by_day[jst_date_str]["lai"].append(row["lai"])
        if row.get("ndvi_contrast") is not None:
            aggregated_by_day[jst_date_str]["ndvi_contrast"].append(row["ndvi_contrast"])
        if row.get("lai_contrast") is not None:
            aggregated_by_day[jst_date_str]["lai_contrast"].append(row["lai_contrast"])

    chart_data = []
    for day in sorted(aggregated_by_day.keys()):
        chart_data.append(
            {
                "date": day,
                "NDVI (平均)": safe_avg(aggregated_by_day[day]["ndvi"]),
                "LAI (平均)": safe_avg(aggregated_by_day[day]["lai"]),
                "NDVI相対 (平均)": safe_avg(aggregated_by_day[day]["ndvi_contrast"]),
                "LAI相対 (平均)": safe_avg(aggregated_by_day[day]["lai_contrast"]),
            }
        )

    if chart_data:
        candidate_cols = ["NDVI (平均)", "LAI (平均)", "NDVI相対 (平均)", "LAI相対 (平均)"]
        active_cols = [
            col for col in candidate_cols
            if any(row.get(col) is not None for row in chart_data)
        ]
        if active_cols:
            st.line_chart(
                chart_data,
                x="date",
                y=active_cols,
                use_container_width=True,
            )
        else:
            st.caption("表示可能な数値データがありません")
    else:
        st.caption("表示可能なデータがありません")

    st.divider()

    # ---- ダウンロード ----
    st.subheader("💾 ダウンロード")
    field_lookup: Dict[str, Dict[str, str]] = {}
    for r in rows:
        if r["field_uuid"] not in field_lookup:
            field_lookup[r["field_uuid"]] = {
                "field_uuid": r["field_uuid"],
                "field_name": r["field_name"],
                "farm_uuid":  r.get("farm_uuid", ""),
                "farm_name":  r.get("farm_name", ""),
            }
    export_fields = list(field_lookup.values())

    export_from_date = st.session_state.last_query_from_date or from_date
    export_till_date = st.session_state.last_query_till_date or till_date

    dl_col1, dl_col2 = st.columns(2)

    with dl_col1:
        export_bytes = build_excel_report(
            rows, export_fields, from_date=export_from_date, till_date=export_till_date
        )
        st.download_button(
            label="📊 Excelレポート (.xlsx)",
            data=export_bytes,
            file_name=f"xarvio_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            help="まとめ・圃場別シート・グラフを含む集計レポート",
        )

    with dl_col2:
        detail_rows = build_detail_csv_rows(rows)
        csv_bytes = to_csv(detail_rows).encode("utf-8-sig")
        st.download_button(
            label="📋 分析用CSV（全詳細）",
            data=csv_bytes,
            file_name=f"xarvio_detail_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True,
            help="1行=1圃場×1日×1ゾーン の完全生データ（BIツール・分析向け）",
        )

if st.session_state.fetch_logs:
    st.caption(f"Last Log: {st.session_state.fetch_logs[-1]}")
